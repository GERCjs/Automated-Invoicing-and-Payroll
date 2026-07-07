from io import BytesIO
from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core import mail
from django.db import IntegrityError, connection, transaction
from django.test import TestCase
from django.test.utils import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from accounts.models import UserRole
from accounts.roles import ADMIN, CUSTOMER, FINANCE, HR, STAFF, SUPERADMIN
from core.models import AuditLog
from notifications.models import EmailDeliveryLog
from payroll.models import Employee, PayrollRecord


class PayrollTestEnvironmentTests(TestCase):
    def test_test_database_uses_migrated_payroll_tables(self):
        user_model = get_user_model()
        hr_user = user_model.objects.create_user(username="payroll_fixture_hr", password="TestOnlyPass123!")
        UserRole.objects.filter(user=hr_user).update(role=HR)

        table_names = set(connection.introspection.table_names())
        self.assertIn("employee", table_names)
        self.assertIn("payslip_record", table_names)

        employee = Employee.objects.create(
            employee_code="STF-999999",
            first_name="Fixture",
            last_name="Tester",
            email="fixture.payroll@example.com",
            hire_date=date(2026, 1, 1),
            base_salary=1000,
            created_by=hr_user,
        )
        payroll_record = PayrollRecord.objects.create(
            employee_name="Fixture Tester",
            employee_id=employee.employee_code,
            basic_salary=1000,
            allowances=100,
            deductions=50,
            cpf_contribution=200,
            net_salary=850,
            payment_date=date(2026, 6, 1),
            created_by=hr_user,
        )

        self.assertEqual(Employee.objects.get(pk=employee.pk).employee_code, "STF-999999")
        self.assertEqual(PayrollRecord.objects.get(pk=payroll_record.pk).employee_id, "STF-999999")


class PayrollUploadPreviewTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(username="hr1", password="pass12345")
        UserRole.objects.filter(user=self.user).update(role=HR)
        self.client.login(username="hr1", password="pass12345")
        Employee.objects.create(
            employee_code="STF-000001",
            first_name="Alex",
            last_name="Tan",
            email="alex@example.com",
            date_of_birth=date(1990, 1, 1),
            hire_date=date(2025, 1, 1),
            base_salary=3000,
        )
        Employee.objects.create(
            employee_code="STF-000002",
            first_name="Jamie",
            last_name="Lim",
            email="jamie@example.com",
            date_of_birth=date(1995, 1, 1),
            hire_date=date(2025, 1, 1),
            base_salary=3200,
        )

    def _build_upload_file(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "employee_code",
                "employee_name",
                "employee_birthofdate",
                "working_days",
                "no_pay_leave_days",
                "basic_salary",
                "physical_products_commission",
                "credit_commission",
                "services_commission",
                "loan_deduction",
                "other_deductions",
                "notes",
            ]
        )
        sheet.append(["STF-000001", "Alex Tan", "01-01-1990", 27, 0, 3000, 15.9, 325, 700, 139.45, 0, "Test"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            "payroll.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _build_upload_file_multiple(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "employee_code",
                "employee_name",
                "employee_birthofdate",
                "working_days",
                "no_pay_leave_days",
                "basic_salary",
                "physical_products_commission",
                "credit_commission",
                "services_commission",
                "loan_deduction",
                "other_deductions",
                "notes",
            ]
        )
        sheet.append(["STF-000001", "Alex Tan", "01-01-1990", 27, 0, 3000, 15.9, 325, 700, 139.45, 0, "Row 1"])
        sheet.append(["STF-000002", "Jamie Lim", "01-01-1995", 26, 0, 3200, 50, 120, 450, 100, 20, "Row 2"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            "payroll_multi.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _build_upload_file_with_invalid_numeric(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "employee_code",
                "employee_name",
                "employee_birthofdate",
                "working_days",
                "no_pay_leave_days",
                "basic_salary",
                "physical_products_commission",
                "credit_commission",
                "services_commission",
                "loan_deduction",
                "other_deductions",
                "notes",
            ]
        )
        sheet.append(["STF-000001", "Alex Tan", "01-01-1990", 27, 0, "abc", 15.9, 325, 700, 139.45, 0, "Bad salary"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            "payroll_invalid_numeric.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _build_upload_file_with_duplicate_employee_code(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "employee_code",
                "employee_name",
                "employee_birthofdate",
                "working_days",
                "no_pay_leave_days",
                "basic_salary",
                "physical_products_commission",
                "credit_commission",
                "services_commission",
                "loan_deduction",
                "other_deductions",
                "notes",
            ]
        )
        sheet.append(["STF-000001", "Alex Tan", "01-01-1990", 27, 0, 3000, 10, 20, 30, 5, 0, "Row 1"])
        sheet.append(["STF-000001", "Alex Tan", "01-01-1990", 27, 0, 3000, 10, 20, 30, 5, 0, "Row 2"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            "payroll_duplicate_employee.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _build_upload_file_with_mismatched_employee_details(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "employee_code",
                "employee_name",
                "employee_birthofdate",
                "working_days",
                "no_pay_leave_days",
                "basic_salary",
                "physical_products_commission",
                "credit_commission",
                "services_commission",
                "loan_deduction",
                "other_deductions",
                "notes",
            ]
        )
        sheet.append(["STF-000001", "Wrong Name", "02-02-1990", 27, 0, 3000, 10, 20, 30, 5, 0, "Mismatch"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            "payroll_mismatch.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_payroll_preview_displays_cpf(self):
        response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file(), "payment_date": "2026-05-01"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Upload Preview")
        self.assertContains(response, "CPF employee share (20%)")
        self.assertContains(response, "808.18")
        self.assertContains(response, "Total rows:")
        self.assertContains(response, "Valid rows:")
        self.assertContains(response, "Invalid rows:")

    def test_confirm_save_creates_payroll_record(self):
        preview_response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file(), "payment_date": "2026-05-24"},
        )
        self.assertEqual(preview_response.status_code, 200)

        save_response = self.client.post(reverse("payroll-upload-confirm-save"))
        self.assertEqual(save_response.status_code, 302)
        self.assertEqual(PayrollRecord.objects.count(), 1)
        record = PayrollRecord.objects.first()
        self.assertEqual(record.employee_id, "STF-000001")
        self.assertEqual(record.payment_date, date(2026, 5, 24))

    def test_confirm_save_creates_multiple_payroll_records(self):
        preview_response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file_multiple(), "payment_date": "2026-05-24"},
        )
        self.assertEqual(preview_response.status_code, 200)

        save_response = self.client.post(reverse("payroll-upload-confirm-save"))
        self.assertEqual(save_response.status_code, 302)
        self.assertEqual(PayrollRecord.objects.count(), 2)

    def test_confirm_save_skips_duplicate_employee_and_payment_date(self):
        PayrollRecord.objects.create(
            employee_name="Alex Tan",
            employee_id="STF-000001",
            basic_salary=3000,
            allowances=1040.90,
            deductions=139.45,
            cpf_contribution=808.18,
            net_salary=3093.27,
            payment_date=date(2026, 5, 24),
        )
        preview_response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file(), "payment_date": "2026-05-24"},
        )
        self.assertEqual(preview_response.status_code, 200)

        save_response = self.client.post(reverse("payroll-upload-confirm-save"), follow=True)
        self.assertEqual(save_response.status_code, 200)
        self.assertEqual(PayrollRecord.objects.count(), 1)
        self.assertContains(save_response, "duplicate record(s) were skipped")

    def test_upload_preview_invalid_numeric_cell_is_row_level_error(self):
        response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file_with_invalid_numeric(), "payment_date": "2026-05-24"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Invalid Rows")
        self.assertContains(response, "Basic salary must be a valid number.")
        self.assertContains(response, "2")

    def test_upload_preview_rejects_duplicate_employee_code_within_file(self):
        response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file_with_duplicate_employee_code(), "payment_date": "2026-05-24"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Employee code is duplicated in this upload file.")

    def test_upload_preview_rejects_mismatched_employee_details_against_employee_record(self):
        response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file_with_mismatched_employee_details(), "payment_date": "2026-05-24"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Employee name does not match employee records.")
        self.assertContains(response, "Employee birthofdate does not match employee records.")

    def test_template_download(self):
        response = self.client.get(reverse("payroll-template-download"))
        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment; filename=", response["Content-Disposition"])

    def test_payroll_list_includes_date_range_hooks_and_shared_script(self):
        response = self.client.get(reverse("payroll-list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "data-date-range-form")
        self.assertContains(response, "data-date-from")
        self.assertContains(response, "data-date-to")
        self.assertContains(response, "data-date-error")
        self.assertContains(response, "js/date-range-filters.js")
        self.assertNotContains(response, 'type="month"', html=False)

    def test_payroll_list_rejects_invalid_date_range_and_preserves_values(self):
        response = self.client.get(
            reverse("payroll-list"),
            {"date_from": "2026-08-02", "date_to": "2025-08-12"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "From date cannot be later than To date.")
        self.assertContains(response, 'name="date_from"', html=False)
        self.assertContains(response, 'value="2026-08-02"', html=False)
        self.assertContains(response, 'name="date_to"', html=False)
        self.assertContains(response, 'value="2025-08-12"', html=False)

    def test_payroll_list_filters_by_payment_date_range(self):
        PayrollRecord.objects.create(
            employee_name="Alex Tan",
            employee_id="EMP001",
            basic_salary=3000,
            allowances=100,
            deductions=50,
            cpf_contribution=600,
            net_salary=2450,
            payment_date=date(2026, 8, 2),
        )
        PayrollRecord.objects.create(
            employee_name="Jamie Lim",
            employee_id="EMP002",
            basic_salary=3200,
            allowances=100,
            deductions=50,
            cpf_contribution=640,
            net_salary=2610,
            payment_date=date(2025, 8, 12),
        )

        response = self.client.get(
            reverse("payroll-list"),
            {"date_from": "2026-08-01", "date_to": "2026-08-31"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "EMP001")
        self.assertNotContains(response, "EMP002")


class PayrollInvalidRowDownloadTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.hr_user = user_model.objects.create_user(username="payroll_invalid_hr", password="pass12345")
        UserRole.objects.filter(user=self.hr_user).update(role=HR)
        self.admin_user = user_model.objects.create_user(username="payroll_invalid_admin", password="pass12345")
        UserRole.objects.filter(user=self.admin_user).update(role=ADMIN)
        self.superadmin_user = user_model.objects.create_user(username="payroll_invalid_super", password="pass12345")
        UserRole.objects.filter(user=self.superadmin_user).update(role=SUPERADMIN)
        self.staff_user = user_model.objects.create_user(username="payroll_invalid_staff", password="pass12345")
        UserRole.objects.filter(user=self.staff_user).update(role=STAFF)
        self.customer_user = user_model.objects.create_user(username="payroll_invalid_customer", password="pass12345")
        UserRole.objects.filter(user=self.customer_user).update(role=CUSTOMER)
        self.finance_user = user_model.objects.create_user(username="payroll_invalid_finance", password="pass12345")
        UserRole.objects.filter(user=self.finance_user).update(role=FINANCE)

        Employee.objects.create(
            employee_code="STF-000001",
            first_name="Alex",
            last_name="Tan",
            email="alex-invalid@example.com",
            hire_date=date(2025, 1, 1),
            base_salary=3000,
        )

    def _build_upload_file(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "employee_code",
                "employee_name",
                "employee_birthofdate",
                "working_days",
                "no_pay_leave_days",
                "basic_salary",
                "physical_products_commission",
                "credit_commission",
                "services_commission",
                "loan_deduction",
                "other_deductions",
                "notes",
            ]
        )
        sheet.append(["STF-000001", "Alex Tan", "01-01-1990", 27, 0, 3000, 15.9, 325, 700, 139.45, 0, "Valid"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            "payroll_valid.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _build_upload_file_with_invalid_numeric(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "employee_code",
                "employee_name",
                "employee_birthofdate",
                "working_days",
                "no_pay_leave_days",
                "basic_salary",
                "physical_products_commission",
                "credit_commission",
                "services_commission",
                "loan_deduction",
                "other_deductions",
                "notes",
            ]
        )
        sheet.append(["STF-000001", "Alex Tan", "01-01-1990", 27, 0, "abc", 15.9, 325, 700, 139.45, 0, "Bad salary"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            "payroll_invalid_numeric.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _build_upload_file_with_multiple_errors(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "employee_code",
                "employee_name",
                "employee_birthofdate",
                "working_days",
                "no_pay_leave_days",
                "basic_salary",
                "physical_products_commission",
                "credit_commission",
                "services_commission",
                "loan_deduction",
                "other_deductions",
                "notes",
            ]
        )
        sheet.append(["", "", "01-01-1990", 27, 0, "abc", 15.9, 325, 700, 139.45, 0, "Bad row"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            "payroll_invalid_multiple.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _upload_invalid_preview(self, upload_file=None):
        self.client.login(username="payroll_invalid_hr", password="pass12345")
        return self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": upload_file or self._build_upload_file_with_invalid_numeric(), "payment_date": "2026-05-24"},
        )

    def _download_workbook(self, response):
        return load_workbook(BytesIO(response.content))

    def test_hr_can_download_invalid_payroll_rows(self):
        preview_response = self._upload_invalid_preview()
        self.assertEqual(preview_response.status_code, 200)

        response = self.client.get(reverse("payroll-download-invalid-rows"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(".xlsx", response["Content-Disposition"])
        self.assertEqual(len(mail.outbox), 0)

    def test_admin_can_download_invalid_payroll_rows(self):
        self.client.force_login(self.admin_user)
        preview_response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file_with_invalid_numeric(), "payment_date": "2026-05-24"},
        )
        self.assertEqual(preview_response.status_code, 200)

        response = self.client.get(reverse("payroll-download-invalid-rows"))

        self.assertEqual(response.status_code, 200)

    def test_superadmin_can_download_invalid_payroll_rows(self):
        self.client.force_login(self.superadmin_user)
        preview_response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file_with_invalid_numeric(), "payment_date": "2026-05-24"},
        )
        self.assertEqual(preview_response.status_code, 200)

        response = self.client.get(reverse("payroll-download-invalid-rows"))

        self.assertEqual(response.status_code, 200)

    def test_staff_customer_and_finance_cannot_download_invalid_rows(self):
        self.client.force_login(self.hr_user)
        self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file_with_invalid_numeric(), "payment_date": "2026-05-24"},
        )

        for user in [self.staff_user, self.customer_user, self.finance_user]:
            with self.subTest(user=user.username):
                self.client.force_login(user)
                response = self.client.get(reverse("payroll-download-invalid-rows"))
                self.assertEqual(response.status_code, 403)

    def test_downloaded_workbook_contains_expected_columns_and_only_invalid_rows(self):
        self._upload_invalid_preview()

        response = self.client.get(reverse("payroll-download-invalid-rows"))
        workbook = self._download_workbook(response)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]

        self.assertEqual(headers[0], "Original Excel Row Number")
        self.assertEqual(headers[-1], "Error Reason")
        self.assertIn("Employee Code", headers)
        self.assertIn("Basic Salary", headers)
        self.assertEqual(worksheet.max_row, 2)

    def test_original_invalid_values_are_preserved_in_download(self):
        self._upload_invalid_preview()

        response = self.client.get(reverse("payroll-download-invalid-rows"))
        workbook = self._download_workbook(response)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        row = [cell.value for cell in worksheet[2]]
        row_map = dict(zip(headers, row))

        self.assertEqual(row_map["Original Excel Row Number"], 2)
        self.assertEqual(row_map["Employee Code"], "STF-000001")
        self.assertEqual(row_map["Basic Salary"], "abc")
        self.assertEqual(row_map["Notes"], "Bad salary")
        self.assertIn("Basic salary must be a valid number.", row_map["Error Reason"])

    def test_multiple_error_messages_are_included_clearly(self):
        self._upload_invalid_preview(self._build_upload_file_with_multiple_errors())

        response = self.client.get(reverse("payroll-download-invalid-rows"))
        workbook = self._download_workbook(response)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        row_map = dict(zip(headers, [cell.value for cell in worksheet[2]]))

        self.assertIn("Basic salary must be a valid number.", row_map["Error Reason"])
        self.assertIn("Employee code is required.", row_map["Error Reason"])
        self.assertIn("Employee name is required.", row_map["Error Reason"])

    def test_no_invalid_rows_redirects_with_clear_message(self):
        self.client.login(username="payroll_invalid_hr", password="pass12345")
        preview_response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file(), "payment_date": "2026-05-24"},
        )
        self.assertEqual(preview_response.status_code, 200)

        response = self.client.get(reverse("payroll-download-invalid-rows"), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "There are no invalid payroll rows to download.")

    def test_missing_or_expired_session_redirects_safely(self):
        self.client.login(username="payroll_invalid_hr", password="pass12345")

        response = self.client.get(reverse("payroll-download-invalid-rows"), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Payroll upload session expired. Please upload the file again.")

    def test_download_invalid_rows_button_appears_only_when_invalid_rows_exist(self):
        self.client.login(username="payroll_invalid_hr", password="pass12345")
        invalid_preview = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file_with_invalid_numeric(), "payment_date": "2026-05-24"},
        )
        valid_preview = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file(), "payment_date": "2026-05-24"},
        )

        self.assertContains(invalid_preview, reverse("payroll-download-invalid-rows"))
        self.assertNotContains(valid_preview, reverse("payroll-download-invalid-rows"))

    def test_corrected_downloaded_workbook_can_be_uploaded_again(self):
        self._upload_invalid_preview()

        download_response = self.client.get(reverse("payroll-download-invalid-rows"))
        workbook = self._download_workbook(download_response)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        basic_salary_col = headers.index("Basic Salary") + 1
        worksheet.cell(row=2, column=basic_salary_col, value=3000)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        corrected_file = SimpleUploadedFile(
            "payroll_invalid_rows_corrected.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": corrected_file, "payment_date": "2026-05-24"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Valid rows:</strong> 1")
        self.assertContains(response, "Invalid rows:</strong> 0")


class EmployeeUploadPreviewValidationTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.hr_user = user_model.objects.create_user(username="employee_upload_hr", password="pass12345")
        UserRole.objects.filter(user=self.hr_user).update(role=HR)
        Employee.objects.create(
            employee_code="STF-000020",
            first_name="Existing",
            last_name="Employee",
            email="existing.employee@example.com",
            hire_date=date(2025, 1, 1),
            base_salary=3000,
        )

    def _build_employee_upload_file(self, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "employee_code",
                "nric",
                "first_name",
                "last_name",
                "date_of_birth",
                "date_of_appointment",
                "legal_status",
                "gender",
                "race",
                "religion",
                "sdl_exempt",
                "cpf_exempt",
                "job_title",
                "email",
                "payment_method",
                "bank_name",
                "bank_account_number",
                "bank_branch_code",
            ]
        )
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            "employee_upload.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_employee_upload_preview_rejects_duplicate_email_and_invalid_date_order(self):
        self.client.force_login(self.hr_user)
        upload_file = self._build_employee_upload_file(
            [
                [
                    "STF-000021",
                    "S1234567A",
                    "Alex",
                    "Tan",
                    "01-01-1995",
                    "01-01-2024",
                    "citizen",
                    "male",
                    "Chinese",
                    "Buddhist",
                    "FALSE",
                    "FALSE",
                    "Therapist",
                    "duplicate@example.com",
                    "cash",
                    "",
                    "",
                    "",
                ],
                [
                    "STF-000022",
                    "S7654321B",
                    "Jamie",
                    "Lim",
                    "01-01-1996",
                    "01-01-1990",
                    "citizen",
                    "female",
                    "Chinese",
                    "Christian",
                    "FALSE",
                    "FALSE",
                    "Manager",
                    "duplicate@example.com",
                    "cash",
                    "",
                    "",
                    "",
                ],
            ]
        )

        response = self.client.post(reverse("employee-upload-preview"), {"employee_file": upload_file})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Email is duplicated in this upload file.")
        self.assertContains(response, "Date of appointment cannot be earlier than date of birth.")

    def test_employee_upload_preview_requires_giro_bank_details(self):
        self.client.force_login(self.hr_user)
        upload_file = self._build_employee_upload_file(
            [
                [
                    "STF-000023",
                    "S2222222C",
                    "Taylor",
                    "Ng",
                    "01-01-1994",
                    "01-01-2024",
                    "citizen",
                    "female",
                    "Chinese",
                    "Buddhist",
                    "FALSE",
                    "FALSE",
                    "Therapist",
                    "taylor.ng@example.com",
                    "giro",
                    "",
                    "",
                    "",
                ]
            ]
        )

        response = self.client.post(reverse("employee-upload-preview"), {"employee_file": upload_file})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bank name is required when payment method is GIRO.")
        self.assertContains(response, "Bank account number is required when payment method is GIRO.")
        self.assertContains(response, "Bank branch code is required when payment method is GIRO.")


class PayrollDuplicatePreventionTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.hr_user = user_model.objects.create_user(username="payroll_hr", password="pass12345")
        UserRole.objects.filter(user=self.hr_user).update(role=HR)
        self.admin_user = user_model.objects.create_user(username="payroll_admin", password="pass12345")
        UserRole.objects.filter(user=self.admin_user).update(role=ADMIN)
        self.superadmin_user = user_model.objects.create_user(username="payroll_super", password="pass12345")
        UserRole.objects.filter(user=self.superadmin_user).update(role=SUPERADMIN)
        self.staff_user = user_model.objects.create_user(username="payroll_staff", password="pass12345")
        UserRole.objects.filter(user=self.staff_user).update(role=STAFF)
        self.customer_user = user_model.objects.create_user(username="payroll_customer", password="pass12345")
        UserRole.objects.filter(user=self.customer_user).update(role=CUSTOMER)

        self.employee = Employee.objects.create(
            employee_code="STF-000010",
            first_name="Alex",
            last_name="Tan",
            email="alex.duplicate@example.com",
            date_of_birth=date(1990, 1, 1),
            hire_date=date(2024, 1, 1),
            base_salary=3000,
        )
        self.other_employee = Employee.objects.create(
            employee_code="STF-000011",
            first_name="Jamie",
            last_name="Lim",
            email="jamie.duplicate@example.com",
            date_of_birth=date(1992, 2, 2),
            hire_date=date(2024, 1, 1),
            base_salary=3200,
        )

    def _payroll_form_data(self, employee, payment_date, **overrides):
        data = {
            "employee_name": f"{employee.first_name} {employee.last_name}",
            "employee_id": employee.employee_code,
            "basic_salary": "3000.00",
            "physical_products_commission": "10.00",
            "credit_commission": "20.00",
            "services_commission": "30.00",
            "loan_deduction": "5.00",
            "other_deductions": "10.00",
            "cpf_contribution": "0.00",
            "payment_date": payment_date.isoformat(),
        }
        data.update(overrides)
        return data

    def _create_record(self, employee, payment_date, **overrides):
        values = {
            "employee_name": f"{employee.first_name} {employee.last_name}",
            "employee_id": employee.employee_code,
            "basic_salary": 3000,
            "allowances": 60,
            "physical_products_commission": 10,
            "credit_commission": 20,
            "services_commission": 30,
            "deductions": 15,
            "loan_deduction": 5,
            "other_deductions": 10,
            "cpf_contribution": 600,
            "net_salary": 2445,
            "payment_date": payment_date,
        }
        values.update(overrides)
        return PayrollRecord.objects.create(**values)

    def _build_upload_file_multiple(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "employee_code",
                "employee_name",
                "employee_birthofdate",
                "working_days",
                "no_pay_leave_days",
                "basic_salary",
                "physical_products_commission",
                "credit_commission",
                "services_commission",
                "loan_deduction",
                "other_deductions",
                "notes",
            ]
        )
        sheet.append(["STF-000010", "Alex Tan", "01-01-1990", 27, 0, 3000, 15.9, 325, 700, 139.45, 0, "Row 1"])
        sheet.append(["STF-000011", "Jamie Lim", "02-02-1992", 26, 0, 3200, 50, 120, 450, 100, 20, "Row 2"])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            "payroll_multi_duplicates.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_manual_creation_succeeds_for_new_employee_and_payment_date(self):
        self.client.login(username="payroll_hr", password="pass12345")

        response = self.client.post(
            reverse("payroll-create"),
            data=self._payroll_form_data(self.employee, date(2026, 6, 30)),
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(PayrollRecord.objects.count(), 1)
        record = PayrollRecord.objects.get()
        self.assertEqual(record.employee_id, self.employee.employee_code)
        self.assertEqual(record.payment_date, date(2026, 6, 30))

    def test_manual_creation_rejects_same_employee_and_payment_date(self):
        self._create_record(self.employee, date(2026, 6, 30))
        self.client.login(username="payroll_hr", password="pass12345")

        response = self.client.post(
            reverse("payroll-create"),
            data=self._payroll_form_data(self.employee, date(2026, 6, 30)),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "A payroll record already exists for this employee and payment date.")
        self.assertEqual(PayrollRecord.objects.count(), 1)

    def test_same_employee_can_have_payroll_on_different_payment_date(self):
        self.client.login(username="payroll_hr", password="pass12345")

        first = self.client.post(
            reverse("payroll-create"),
            data=self._payroll_form_data(self.employee, date(2026, 6, 30)),
        )
        second = self.client.post(
            reverse("payroll-create"),
            data=self._payroll_form_data(self.employee, date(2026, 7, 31)),
        )

        self.assertEqual(first.status_code, 302)
        self.assertEqual(second.status_code, 302)
        self.assertEqual(PayrollRecord.objects.count(), 2)

    def test_different_employees_can_have_payroll_on_same_payment_date(self):
        self.client.login(username="payroll_hr", password="pass12345")

        first = self.client.post(
            reverse("payroll-create"),
            data=self._payroll_form_data(self.employee, date(2026, 6, 30)),
        )
        second = self.client.post(
            reverse("payroll-create"),
            data=self._payroll_form_data(
                self.other_employee,
                date(2026, 6, 30),
                employee_name="Jamie Lim",
                basic_salary="3200.00",
            ),
        )

        self.assertEqual(first.status_code, 302)
        self.assertEqual(second.status_code, 302)
        self.assertEqual(PayrollRecord.objects.count(), 2)

    def test_editing_record_without_changing_employee_and_payment_date_succeeds(self):
        record = self._create_record(self.employee, date(2026, 6, 30))
        self.client.login(username="payroll_hr", password="pass12345")

        response = self.client.post(
            reverse("payroll-edit", args=[record.pk]),
            data=self._payroll_form_data(
                self.employee,
                date(2026, 6, 30),
                basic_salary="3100.00",
            ),
        )

        self.assertEqual(response.status_code, 302)
        record.refresh_from_db()
        self.assertEqual(record.employee_id, self.employee.employee_code)
        self.assertEqual(record.payment_date, date(2026, 6, 30))
        self.assertEqual(record.basic_salary, 3100)

    def test_editing_record_to_conflict_with_another_record_is_rejected(self):
        existing = self._create_record(self.employee, date(2026, 6, 30))
        editable = self._create_record(self.employee, date(2026, 7, 31), basic_salary=3200, net_salary=2645)
        self.client.login(username="payroll_hr", password="pass12345")

        response = self.client.post(
            reverse("payroll-edit", args=[editable.pk]),
            data=self._payroll_form_data(
                self.employee,
                date(2026, 6, 30),
                basic_salary="3200.00",
            ),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "A payroll record already exists for this employee and payment date.")
        editable.refresh_from_db()
        self.assertEqual(editable.payment_date, date(2026, 7, 31))
        self.assertEqual(PayrollRecord.objects.filter(employee_id=self.employee.employee_code).count(), 2)
        self.assertTrue(PayrollRecord.objects.filter(pk=existing.pk).exists())

    def test_excel_confirmation_skips_duplicate_and_saves_other_non_duplicate_rows(self):
        self._create_record(self.employee, date(2026, 5, 24))
        self.client.login(username="payroll_hr", password="pass12345")

        preview_response = self.client.post(
            reverse("payroll-upload-preview"),
            {"payroll_file": self._build_upload_file_multiple(), "payment_date": "2026-05-24"},
        )
        self.assertEqual(preview_response.status_code, 200)
        self.assertContains(
            preview_response,
            "A payroll record already exists for this employee and payment date.",
        )

        save_response = self.client.post(reverse("payroll-upload-confirm-save"), follow=True)

        self.assertEqual(save_response.status_code, 200)
        self.assertEqual(PayrollRecord.objects.count(), 2)
        self.assertTrue(
            PayrollRecord.objects.filter(
                employee_id=self.other_employee.employee_code,
                payment_date=date(2026, 5, 24),
            ).exists()
        )
        self.assertContains(save_response, "Payroll upload saved successfully. 1 record(s) created.")

    def test_direct_database_duplicate_create_raises_integrity_error(self):
        self._create_record(self.employee, date(2026, 6, 30))

        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self._create_record(self.employee, date(2026, 6, 30), basic_salary=3500, net_salary=2845)

    def test_hr_admin_and_superadmin_can_access_manual_payroll_creation(self):
        for username, role in [
            ("payroll_hr", HR),
            ("payroll_admin", ADMIN),
            ("payroll_super", SUPERADMIN),
        ]:
            with self.subTest(role=role):
                self.client.force_login(get_user_model().objects.get(username=username))
                create_response = self.client.get(reverse("payroll-create"))
                upload_response = self.client.post(reverse("payroll-upload-confirm-save"))
                self.assertEqual(create_response.status_code, 200)
                self.assertEqual(upload_response.status_code, 302)
                self.client.logout()

    def test_manual_payroll_creation_uses_employee_dropdown(self):
        self.client.force_login(get_user_model().objects.get(username="payroll_hr"))

        response = self.client.get(reverse("payroll-create"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<select name="employee_id"', html=False)
        self.assertContains(response, "Select an employee ID")
        self.assertContains(response, f"{self.employee.employee_code} - Alex Tan")

    def test_staff_and_customer_cannot_use_manual_payroll_creation_or_upload_confirmation(self):
        for username, role in [
            ("payroll_staff", STAFF),
            ("payroll_customer", CUSTOMER),
        ]:
            with self.subTest(role=role):
                self.client.force_login(get_user_model().objects.get(username=username))
                create_response = self.client.get(reverse("payroll-create"))
                upload_response = self.client.post(reverse("payroll-upload-confirm-save"))
                self.assertEqual(create_response.status_code, 403)
                self.assertEqual(upload_response.status_code, 403)
                self.client.logout()


class PayslipPdfAccessTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.staff_user = user_model.objects.create_user(
            username="staff1", password="pass12345", email="staff1@example.com"
        )
        UserRole.objects.filter(user=self.staff_user).update(role=STAFF)
        self.other_staff_user = user_model.objects.create_user(
            username="staff2", password="pass12345", email="staff2@example.com"
        )
        UserRole.objects.filter(user=self.other_staff_user).update(role=STAFF)
        self.hr_user = user_model.objects.create_user(username="hr2", password="pass12345")
        UserRole.objects.filter(user=self.hr_user).update(role=HR)

        self.staff_employee = Employee.objects.create(
            user=self.staff_user,
            employee_code="EMP100",
            first_name="Staff",
            last_name="One",
            email="staff1@example.com",
            hire_date=date(2024, 1, 1),
            base_salary=2000,
        )
        self.other_employee = Employee.objects.create(
            user=self.other_staff_user,
            employee_code="EMP200",
            first_name="Staff",
            last_name="Two",
            email="staff2@example.com",
            hire_date=date(2024, 1, 1),
            base_salary=2200,
        )

        self.staff_record = PayrollRecord.objects.create(
            employee_name="Staff One",
            employee_id=self.staff_employee.employee_code,
            basic_salary=2000,
            allowances=100,
            deductions=50,
            cpf_contribution=400,
            net_salary=1650,
            payment_date=date(2026, 5, 24),
        )
        self.other_record = PayrollRecord.objects.create(
            employee_name="Staff Two",
            employee_id=self.other_employee.employee_code,
            basic_salary=2200,
            allowances=100,
            deductions=50,
            cpf_contribution=440,
            net_salary=1810,
            payment_date=date(2026, 5, 24),
        )

    def test_hr_can_download_any_payslip_pdf(self):
        self.client.login(username="hr2", password="pass12345")
        response = self.client.get(reverse("payslip-pdf-download", args=[self.other_record.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")

    def test_staff_can_download_own_payslip_pdf_only(self):
        self.client.login(username="staff1", password="pass12345")
        own_response = self.client.get(reverse("payslip-pdf-download", args=[self.staff_record.pk]))
        self.assertEqual(own_response.status_code, 200)
        self.assertEqual(own_response["Content-Type"], "application/pdf")

        denied_response = self.client.get(reverse("payslip-pdf-download", args=[self.other_record.pk]))
        self.assertEqual(denied_response.status_code, 403)


class PayrollReportPlacementAndAccessTests(TestCase):
    def setUp(self):
        self.report_month = "2026-06"
        self.report_payment_date = date(2026, 6, 30)
        self.employee_one = Employee.objects.create(
            employee_code="STF-200001",
            first_name="Report",
            last_name="One",
            email="report.one@example.com",
            hire_date=date(2025, 1, 1),
            date_of_birth=date(1995, 6, 1),
            base_salary=3000,
        )
        self.employee_two = Employee.objects.create(
            employee_code="STF-200002",
            first_name="Report",
            last_name="Two",
            email="report.two@example.com",
            hire_date=date(2025, 1, 1),
            date_of_birth=date(1992, 3, 14),
            base_salary=2800,
        )
        self.employee_missing = Employee.objects.create(
            employee_code="STF-200003",
            first_name="Missing",
            last_name="Record",
            email="missing.record@example.com",
            hire_date=date(2025, 1, 1),
            date_of_birth=date(1990, 11, 2),
            base_salary=2600,
        )
        Employee.objects.create(
            employee_code="STF-200004",
            first_name="Inactive",
            last_name="Employee",
            email="inactive.employee@example.com",
            hire_date=date(2025, 1, 1),
            date_of_birth=date(1991, 8, 8),
            base_salary=2400,
            status=Employee.STATUS_INACTIVE,
        )
        self.record_one = PayrollRecord.objects.create(
            employee_name="Report One",
            employee_id=self.employee_one.employee_code,
            basic_salary=3000,
            allowances=150,
            deductions=50,
            cpf_contribution=600,
            net_salary=2500,
            payment_date=self.report_payment_date,
        )
        self.record_two = PayrollRecord.objects.create(
            employee_name="Report Two",
            employee_id=self.employee_two.employee_code,
            basic_salary=2800,
            allowances=100,
            deductions=25,
            cpf_contribution=560,
            net_salary=2315,
            payment_date=self.report_payment_date,
        )
        EmailDeliveryLog.objects.create(
            recipient_email="report.two@example.com",
            subject="Payslip",
            template_key="payroll_payslip",
            status=EmailDeliveryLog.STATUS_FAILED,
            related_object_type="payroll_record",
            related_object_id=str(self.record_two.pk),
            error_message="SMTP timeout",
        )
        AuditLog.objects.create(
            action="payroll.upload.previewed",
            metadata={
                "row_count": 3,
                "valid_row_count": 1,
                "invalid_row_count": 2,
                "source_file_name": "payroll_june.xlsx",
            },
        )
        AuditLog.objects.create(
            action="payroll.upload.saved",
            metadata={
                "saved_count": 2,
                "skipped_duplicate_count": 1,
                "payment_date": self.report_payment_date.isoformat(),
            },
        )

    def _make_user(self, username, role):
        user = get_user_model().objects.create_user(username=username, password="pass12345")
        UserRole.objects.filter(user=user).update(role=role)
        return user

    def test_main_navbar_does_not_show_payroll_report_link(self):
        user = self._make_user("hr_nav_user", HR)
        self.client.force_login(user)
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("payroll-dashboard"))

    def test_payroll_dashboard_shows_payroll_report_link_for_hr(self):
        user = self._make_user("hr_dashboard_user", HR)
        self.client.force_login(user)
        response = self.client.get(reverse("payroll-dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("payroll-report"))

    def test_staff_and_customer_cannot_access_overall_payroll_report(self):
        staff_user = self._make_user("staff_no_payroll_report", STAFF)
        self.client.force_login(staff_user)
        staff_response = self.client.get(reverse("payroll-report"))
        self.assertEqual(staff_response.status_code, 403)
        self.client.logout()

        customer_user = self._make_user("customer_no_payroll_report", CUSTOMER)
        self.client.force_login(customer_user)
        customer_response = self.client.get(reverse("payroll-report"))
        self.assertEqual(customer_response.status_code, 403)

    def test_finance_cannot_access_overall_payroll_report(self):
        finance_user = self._make_user("finance_no_payroll_report", FINANCE)
        self.client.force_login(finance_user)
        response = self.client.get(reverse("payroll-report"))
        self.assertEqual(response.status_code, 403)

    def test_admin_and_superadmin_can_access_payroll_report(self):
        admin_user = self._make_user("admin_payroll_report", ADMIN)
        self.client.force_login(admin_user)
        admin_response = self.client.get(reverse("payroll-report"))
        self.assertEqual(admin_response.status_code, 200)
        self.client.logout()

        superadmin_user = self._make_user("super_payroll_report", SUPERADMIN)
        self.client.force_login(superadmin_user)
        super_response = self.client.get(reverse("payroll-report"))
        self.assertEqual(super_response.status_code, 200)

    def test_hr_can_access_payroll_report(self):
        hr_user = self._make_user("hr_payroll_report", HR)
        self.client.force_login(hr_user)
        response = self.client.get(reverse("payroll-report"), data={"month": self.report_month})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Payroll Report")

    def test_payroll_report_shows_dashboard_sections_and_kpis(self):
        hr_user = self._make_user("hr_payroll_report_ui", HR)
        self.client.force_login(hr_user)
        response = self.client.get(reverse("payroll-report"), data={"month": self.report_month})

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Use this report to review payroll costs, CPF amounts, processing issues and individual employee payroll records.",
        )
        self.assertContains(response, "Reporting period: June 2026")
        self.assertContains(response, "Monthly Filters")
        self.assertContains(response, "Total Payroll Cost")
        self.assertContains(response, "Total Net Pay")
        self.assertContains(response, "Total CPF")
        self.assertContains(response, "Employees Paid")
        self.assertContains(response, "Employee CPF")
        self.assertContains(response, "Employer CPF")
        self.assertContains(response, "Total Deductions")
        self.assertContains(response, "Failed Payslip Emails")
        self.assertContains(response, "Payroll Processing Issues")
        self.assertContains(response, "Invalid upload rows")
        self.assertContains(response, "Duplicate rows skipped")
        self.assertContains(response, "Failed payslip emails")
        self.assertContains(response, "Missing payroll records")
        self.assertContains(response, "Monthly Payroll Cost Trend")
        self.assertContains(response, "CPF Contribution Split")
        self.assertContains(response, "Detailed Payroll Records")
        self.assertContains(response, "View Record")
        self.assertContains(response, "View Payslip")
        self.assertContains(response, "S$6,050.00")
        self.assertContains(response, "S$4,815.00")
        self.assertContains(response, "STF-200003")
        self.assertContains(response, "js/report_charts.js")
        self.assertNotContains(response, ">Open<")
        self.assertEqual(response.context["total_payroll_amount_month"], Decimal("6050"))
        self.assertEqual(response.context["total_net_pay_month"], Decimal("4815"))
        self.assertEqual(response.context["employees_paid_month"], 2)
        self.assertEqual(response.context["failed_payslip_email_count"], 1)
        self.assertEqual(response.context["invalid_upload_rows_count"], 2)
        self.assertEqual(response.context["duplicate_rows_skipped_count"], 1)
        self.assertEqual(response.context["missing_payroll_records_count"], 1)
        self.assertEqual(
            response.context["total_cpf_month"],
            response.context["employee_cpf_total_month"] + response.context["employer_cpf_total_month"],
        )

    def test_payroll_report_preserves_selected_month_and_employee_filters(self):
        hr_user = self._make_user("hr_payroll_report_filter", HR)
        self.client.force_login(hr_user)

        response = self.client.get(
            reverse("payroll-report"),
            data={"month": "2026-06", "employee": "STF-200001"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="month"', html=False)
        self.assertContains(response, 'value="2026-06"', html=False)
        self.assertContains(response, 'name="employee"', html=False)
        self.assertContains(response, 'value="STF-200001"', html=False)
        self.assertContains(response, "Showing payroll records for June 2026 and employee search")

    def test_payroll_report_empty_state_shows_when_no_records_match(self):
        hr_user = self._make_user("hr_payroll_report_empty", HR)
        self.client.force_login(hr_user)

        response = self.client.get(
            reverse("payroll-report"),
            data={"month": "2026-08", "employee": "NO-MATCH"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No payroll records were found for the selected month.")


class MyPayslipsViewTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.staff_user = user_model.objects.create_user(
            username="staff_list_1", password="pass12345", email="stafflist1@example.com"
        )
        UserRole.objects.filter(user=self.staff_user).update(role=STAFF)
        self.other_staff_user = user_model.objects.create_user(
            username="staff_list_2", password="pass12345", email="stafflist2@example.com"
        )
        UserRole.objects.filter(user=self.other_staff_user).update(role=STAFF)

        self.staff_employee = Employee.objects.create(
            user=self.staff_user,
            employee_code="EMP300",
            first_name="List",
            last_name="One",
            email="stafflist1@example.com",
            hire_date=date(2024, 1, 1),
            base_salary=2100,
        )
        self.other_employee = Employee.objects.create(
            user=self.other_staff_user,
            employee_code="EMP400",
            first_name="List",
            last_name="Two",
            email="stafflist2@example.com",
            hire_date=date(2024, 1, 1),
            base_salary=2300,
        )

        self.own_record = PayrollRecord.objects.create(
            employee_name="List One",
            employee_id=self.staff_employee.employee_code,
            basic_salary=2100,
            allowances=100,
            deductions=50,
            cpf_contribution=420,
            net_salary=1730,
            payment_date=date(2026, 5, 24),
        )
        self.other_record = PayrollRecord.objects.create(
            employee_name="List Two",
            employee_id=self.other_employee.employee_code,
            basic_salary=2300,
            allowances=100,
            deductions=50,
            cpf_contribution=460,
            net_salary=1890,
            payment_date=date(2026, 5, 24),
        )

    def test_staff_my_payslips_shows_only_own_records_without_page_view_log(self):
        self.client.login(username="staff_list_1", password="pass12345")
        response = self.client.get(reverse("my-payslips"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.own_record.employee_id)
        self.assertNotContains(response, self.other_record.employee_id)
        self.assertFalse(
            AuditLog.objects.filter(
                user=self.staff_user,
                action="payroll.my_payslips.viewed",
            ).exists()
        )


@override_settings(
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
    DEFAULT_FROM_EMAIL="noreply@example.com",
)
class PayslipEmailSendTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.hr_user = user_model.objects.create_user(username="hr_email_test", password="pass12345")
        UserRole.objects.filter(user=self.hr_user).update(role=HR)

        self.employee = Employee.objects.create(
            employee_code="EMP500",
            first_name="Mail",
            last_name="Tester",
            email="mail.tester@example.com",
            hire_date=date(2024, 1, 1),
            base_salary=2400,
        )
        self.record = PayrollRecord.objects.create(
            employee_name="Mail Tester",
            employee_id=self.employee.employee_code,
            basic_salary=2400,
            allowances=100,
            deductions=50,
            cpf_contribution=480,
            net_salary=1970,
            payment_date=date(2026, 5, 24),
        )

    def test_payslip_email_includes_pdf_attachment(self):
        self.client.login(username="hr_email_test", password="pass12345")
        response = self.client.post(reverse("payslip-email-send", args=[self.record.pk]))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(len(mail.outbox), 1)
        sent_message = mail.outbox[0]
        self.assertTrue(sent_message.attachments)
        self.assertTrue(any(attachment[2] == "application/pdf" for attachment in sent_message.attachments))
        self.assertTrue(
            EmailDeliveryLog.objects.filter(
                related_object_type="payroll_record",
                related_object_id=str(self.record.id),
                status=EmailDeliveryLog.STATUS_SENT,
            ).exists()
        )
