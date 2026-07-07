from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import date, datetime
from typing import Any
import re

from openpyxl import load_workbook


MONEY_PLACES = Decimal("0.01")

TEMPLATE_HEADERS = [
    "employee_code",
    "employee_name",
    "employee_birthofdate",
    "working_days",
    "no_pay_leave_days",
    "basic_salary",
    "physical_products_commission",
    "credit_commission",
    "services_commission",
    "loan_deduction",
    "other_deductions",
    "notes",
]

PAYROLL_UPLOAD_COLUMN_LABELS = {
    "employee_code": "Employee Code",
    "employee_name": "Employee Name",
    "employee_birthofdate": "Employee Birthofdate",
    "working_days": "Working Days",
    "no_pay_leave_days": "No Pay Leave Days",
    "basic_salary": "Basic Salary",
    "physical_products_commission": "Physical Products Commission",
    "credit_commission": "Credit Commission",
    "services_commission": "Services Commission",
    "loan_deduction": "Loan Deduction",
    "other_deductions": "Other Deductions",
    "notes": "Notes",
}

PAYROLL_UPLOAD_HEADER_ALIASES = {
    label.strip().lower(): header
    for header, label in PAYROLL_UPLOAD_COLUMN_LABELS.items()
}

EMPLOYEE_CODE_PATTERN = re.compile(r"^STF-[0-9]{6}$")


@dataclass
class CPFResult:
    employee_rate: Decimal
    employer_rate: Decimal
    employee_amount: Decimal
    employer_amount: Decimal


def as_decimal(value: Any, default: str = "0") -> Decimal:
    if value in (None, ""):
        return Decimal(default)
    return Decimal(str(value))


def money(value: Decimal) -> Decimal:
    return value.quantize(MONEY_PLACES, rounding=ROUND_HALF_UP)


def cpf_rates_for_age(age: int) -> tuple[Decimal, Decimal]:
    if age <= 55:
        return Decimal("20"), Decimal("17")
    if age <= 60:
        return Decimal("18"), Decimal("16")
    if age <= 65:
        return Decimal("12.5"), Decimal("12.5")
    if age <= 70:
        return Decimal("7.5"), Decimal("9")
    return Decimal("5"), Decimal("7.5")


def cpf_for_2026(monthly_wage: Decimal, age: int) -> CPFResult:
    employee_rate, employer_rate = cpf_rates_for_age(age)

    if monthly_wage <= Decimal("50"):
        employee_amount = Decimal("0")
        employer_amount = Decimal("0")
    elif monthly_wage <= Decimal("500"):
        employee_amount = Decimal("0")
        employer_amount = monthly_wage * (employer_rate / Decimal("100"))
    elif monthly_wage <= Decimal("750"):
        # CPF Board graduated employee share for >$500 to $750 follows:
        # 0.6 / 0.54 / 0.375 / 0.225 / 0.15 x (TW - 500) by age band in 2026.
        graduated_multiplier = (employee_rate / Decimal("100")) * Decimal("3")
        employee_amount = graduated_multiplier * (monthly_wage - Decimal("500"))
        total_amount = (monthly_wage * (employer_rate / Decimal("100"))) + employee_amount
        employer_amount = total_amount - employee_amount
    else:
        employee_amount = monthly_wage * (employee_rate / Decimal("100"))
        employer_amount = monthly_wage * (employer_rate / Decimal("100"))

    return CPFResult(
        employee_rate=employee_rate,
        employer_rate=employer_rate,
        employee_amount=money(employee_amount),
        employer_amount=money(employer_amount),
    )


def calculate_age_on(dob: date, on_date: date) -> int:
    years = on_date.year - dob.year
    if (on_date.month, on_date.day) < (dob.month, dob.day):
        years -= 1
    return years


def employee_cpf_contribution_2026_from_basic_salary(
    basic_salary: Decimal, dob: date, payment_date: date
) -> Decimal:
    age = calculate_age_on(dob, payment_date)
    cpf = cpf_for_2026(basic_salary, age)
    return cpf.employee_amount


def parse_payroll_excel(uploaded_file, payment_date: date) -> list[dict[str, Any]]:
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Uploaded file is empty.")

    headers = [
        PAYROLL_UPLOAD_HEADER_ALIASES.get(str(h).strip().lower(), str(h).strip().lower())
        if h is not None
        else ""
        for h in rows[0]
    ]
    required = set(TEMPLATE_HEADERS)
    missing = required - set(headers)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    index_map = {name: headers.index(name) for name in required}
    parsed: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if row is None or all(v in (None, "") for v in row):
            continue
        row_errors: list[str] = []
        raw_values = {
            header: _stringify_uploaded_value(row[index_map[header]])
            for header in TEMPLATE_HEADERS
        }
        employee_code = str(row[index_map["employee_code"]] or "").strip()
        employee_name = str(row[index_map["employee_name"]] or "").strip()

        employee_birthofdate = _parse_date_field(
            row[index_map["employee_birthofdate"]],
            "Employee birthofdate",
            row_errors,
        )
        age = None
        if employee_birthofdate is not None:
            age = calculate_age_on(employee_birthofdate, payment_date)
            if age < 0:
                row_errors.append("Employee birthofdate cannot be after payment date.")

        working_days = _parse_decimal_field(
            row[index_map["working_days"]],
            "Working days",
            row_errors,
            default="0",
        )
        no_pay_leave_days = _parse_decimal_field(
            row[index_map["no_pay_leave_days"]],
            "No pay leave days",
            row_errors,
            default="0",
        )
        basic_salary = _parse_decimal_field(row[index_map["basic_salary"]], "Basic salary", row_errors)
        physical_products_commission = _parse_decimal_field(
            row[index_map["physical_products_commission"]],
            "Physical products commission",
            row_errors,
            default="0",
        )
        credit_commission = _parse_decimal_field(
            row[index_map["credit_commission"]],
            "Credit commission",
            row_errors,
            default="0",
        )
        services_commission = _parse_decimal_field(
            row[index_map["services_commission"]],
            "Services commission",
            row_errors,
            default="0",
        )
        loan_deduction = _parse_decimal_field(row[index_map["loan_deduction"]], "Loan deduction", row_errors, default="0")
        other_deductions = _parse_decimal_field(
            row[index_map["other_deductions"]],
            "Other deductions",
            row_errors,
            default="0",
        )

        if row_errors:
            parsed.append(
                {
                    "row_number": row_number,
                    "employee_code": employee_code,
                    "employee_name": employee_name,
                    "__parse_errors": row_errors,
                    "__raw_values": raw_values,
                }
            )
            continue

        total_earnings = basic_salary + physical_products_commission + credit_commission + services_commission
        cpf = cpf_for_2026(total_earnings, age)
        total_deductions = loan_deduction + other_deductions + cpf.employee_amount
        net_pay = total_earnings - total_deductions

        parsed.append(
            {
                "row_number": row_number,
                "employee_code": employee_code,
                "employee_name": employee_name,
                "employee_birthofdate": employee_birthofdate.strftime("%d-%m-%Y"),
                "working_days": working_days,
                "no_pay_leave_days": no_pay_leave_days,
                "basic_salary": money(basic_salary),
                "physical_products_commission": money(physical_products_commission),
                "credit_commission": money(credit_commission),
                "services_commission": money(services_commission),
                "loan_deduction": money(loan_deduction),
                "other_deductions": money(other_deductions),
                "notes": row[index_map["notes"]] or "",
                "total_earnings": money(total_earnings),
                "total_deductions": money(total_deductions),
                "net_pay": money(net_pay),
                "cpf_employee_rate": cpf.employee_rate,
                "cpf_employer_rate": cpf.employer_rate,
                "cpf_employee_amount": cpf.employee_amount,
                "cpf_employer_amount": cpf.employer_amount,
                "__raw_values": raw_values,
            }
        )
    return parsed


def parse_and_validate_payroll_excel(uploaded_file, payment_date: date) -> dict[str, Any]:
    parsed_rows = parse_payroll_excel(uploaded_file, payment_date)
    invalid_rows: list[dict[str, Any]] = []
    valid_rows: list[dict[str, Any]] = []
    seen_employee_rows: dict[str, int] = {}

    for row in parsed_rows:
        reasons: list[str] = list(row.get("__parse_errors", []))
        employee_code = str(row.get("employee_code") or "").strip()
        if not employee_code:
            reasons.append("Employee code is required.")
        elif not EMPLOYEE_CODE_PATTERN.fullmatch(employee_code):
            reasons.append("Employee code must follow STF-000000 format.")
        elif employee_code in seen_employee_rows:
            reasons.append(
                f"Employee code is duplicated in this upload file. First duplicate appears on row {seen_employee_rows[employee_code]}."
            )
        else:
            seen_employee_rows[employee_code] = int(row.get("row_number") or 0)
        if not str(row.get("employee_name") or "").strip():
            reasons.append("Employee name is required.")
        if row.get("basic_salary") is not None and row.get("basic_salary", Decimal("0")) < 0:
            reasons.append("Basic salary cannot be negative.")
        if row.get("working_days") is not None and row.get("working_days", Decimal("0")) < 0:
            reasons.append("Working days cannot be negative.")
        if row.get("no_pay_leave_days") is not None and row.get("no_pay_leave_days", Decimal("0")) < 0:
            reasons.append("No pay leave days cannot be negative.")
        if (
            row.get("working_days") is not None
            and row.get("no_pay_leave_days") is not None
            and row.get("working_days", Decimal("0")) >= 0
            and row.get("no_pay_leave_days", Decimal("0")) >= 0
            and row.get("no_pay_leave_days", Decimal("0")) > row.get("working_days", Decimal("0"))
        ):
            reasons.append("No pay leave days cannot be more than working days.")
        if row.get("physical_products_commission") is not None and row.get("physical_products_commission", Decimal("0")) < 0:
            reasons.append("Physical products commission cannot be negative.")
        if row.get("credit_commission") is not None and row.get("credit_commission", Decimal("0")) < 0:
            reasons.append("Credit commission cannot be negative.")
        if row.get("services_commission") is not None and row.get("services_commission", Decimal("0")) < 0:
            reasons.append("Services commission cannot be negative.")
        if row.get("loan_deduction") is not None and row.get("loan_deduction", Decimal("0")) < 0:
            reasons.append("Loan deduction cannot be negative.")
        if row.get("other_deductions") is not None and row.get("other_deductions", Decimal("0")) < 0:
            reasons.append("Other deductions cannot be negative.")
        if row.get("net_pay") is not None and row.get("net_pay", Decimal("0")) < 0:
            reasons.append("Net pay cannot be negative. Check salary, commissions, and deductions.")

        if reasons:
            invalid_rows.append(
                {
                    "row_number": row.get("row_number"),
                    "employee_code": row.get("employee_code", ""),
                    "employee_name": row.get("employee_name", ""),
                    "errors": reasons,
                    "raw_values": row.get("__raw_values", {}),
                }
            )
        else:
            valid_rows.append(row)

    return {
        "total_rows": len(parsed_rows),
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
    }


def default_template_row() -> list[Any]:
    return [
        "STF-000001",
        "Alex Tan",
        "01-01-1990",
        27,
        0,
        3000,
        15.9,
        325,
        700,
        139.45,
        0,
        "Sample row",
    ]


def _parse_decimal_field(value: Any, label: str, row_errors: list[str], default: str | None = None) -> Decimal | None:
    try:
        if default is None:
            return as_decimal(value)
        return as_decimal(value, default=default)
    except (InvalidOperation, ValueError):
        row_errors.append(f"{label} must be a valid number.")
        return None


def _parse_date_field(value: Any, label: str, row_errors: list[str]) -> date | None:
    if value in (None, ""):
        row_errors.append(f"{label} is required.")
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    row_errors.append(f"{label} must be DD-MM-YYYY.")
    return None


def _stringify_uploaded_value(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    return str(value).strip()
