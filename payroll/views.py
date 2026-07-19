from io import BytesIO
from decimal import Decimal, InvalidOperation
from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import Workbook
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from django.db import IntegrityError, transaction
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
from reportlab.pdfgen import canvas

from accounts.permissions import get_user_role, role_required
from accounts.roles import ADMIN, HR, STAFF, SUPERADMIN
from core.audit import get_client_ip, log_event
from core.models import AuditLog
from notifications.models import EmailDeliveryLog

from .forms import EmployeeForm, EmployeeUploadForm, PayrollRecordForm, PayrollUploadForm
from .models import Employee, PayrollRecord
from .services import (
    PAYROLL_UPLOAD_COLUMN_LABELS,
    TEMPLATE_HEADERS,
    cpf_for_2026,
    parse_and_validate_payroll_excel,
)

EMPLOYEE_CODE_PATTERN = re.compile(r"^STF-[0-9]{6}$")


def _safe_sum(queryset, field_name):
    return queryset.aggregate(total=Sum(field_name))["total"] or Decimal("0")


def _to_float(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _build_chart_summary(labels, values):
    numeric_values = [float(value or 0) for value in values]
    total_value = sum(numeric_values)
    has_data = any(value > 0 for value in numeric_values)
    peak_label = ""
    peak_value = 0.0
    if has_data and labels and numeric_values:
        peak_index = max(range(len(numeric_values)), key=numeric_values.__getitem__)
        peak_label = labels[peak_index]
        peak_value = numeric_values[peak_index]
    return {
        "has_data": has_data,
        "six_month_total": total_value,
        "peak_label": peak_label,
        "peak_value": peak_value,
    }


def _recent_month_starts(today, total_months=6):
    month_start = today.replace(day=1)
    month_starts = []
    for offset in range(total_months - 1, -1, -1):
        year = month_start.year
        month = month_start.month - offset
        while month <= 0:
            month += 12
            year -= 1
        month_starts.append(month_start.replace(year=year, month=month, day=1))
    return month_starts


def _month_starts_from_anchor(anchor_month, end_month):
    start_month = anchor_month.replace(day=1)
    final_month = end_month.replace(day=1)
    month_starts = []
    current_month = start_month

    while current_month <= final_month:
        month_starts.append(current_month)
        if current_month.month == 12:
            current_month = current_month.replace(year=current_month.year + 1, month=1, day=1)
        else:
            current_month = current_month.replace(month=current_month.month + 1, day=1)

    return month_starts


def _parse_iso_date(raw_value: str):
    value = (raw_value or "").strip()
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _parse_date_range(raw_date_from: str, raw_date_to: str):
    date_filter_error = ""
    date_from = None
    date_to = None

    if raw_date_from:
        date_from = _parse_iso_date(raw_date_from)
        if date_from is None:
            date_filter_error = "From date is invalid. Use YYYY-MM-DD."

    if raw_date_to and not date_filter_error:
        date_to = _parse_iso_date(raw_date_to)
        if date_to is None:
            date_filter_error = "To date is invalid. Use YYYY-MM-DD."

    if date_from and date_to and date_from > date_to:
        date_filter_error = "From date cannot be later than To date."

    filter_date_from = None if date_filter_error else date_from
    filter_date_to = None if date_filter_error else date_to
    return date_from, date_to, filter_date_from, filter_date_to, date_filter_error


def _parse_month_filter(raw_value: str):
    value = (raw_value or "").strip()
    if not value:
        return "", None, None
    try:
        month_start = date.fromisoformat(f"{value}-01")
    except ValueError:
        return "", None, None
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1, day=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1, day=1)
    month_end = next_month - timezone.timedelta(days=1)
    return value, month_start, month_end


def _generate_next_employee_code():
    latest_code = (
        Employee.objects.filter(employee_code__regex=r"^STF-[0-9]{6}$")
        .order_by("-id")
        .values_list("employee_code", flat=True)
        .first()
    )
    if not latest_code:
        return "STF-000001"

    number = int(latest_code.split("-")[1]) + 1
    return f"STF-{number:06d}"


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payroll_employee_lookup(request):
    employee_id = (request.GET.get("employee_id") or "").strip()
    if not employee_id:
        return JsonResponse({"ok": False, "employee_name": "", "reason": "missing_employee_id"})

    employee = Employee.objects.filter(employee_code=employee_id).first()
    if employee is None:
        return JsonResponse({"ok": False, "employee_name": "", "reason": "employee_not_found"})

    employee_name = f"{employee.first_name} {employee.last_name}".strip()
    return JsonResponse({"ok": True, "employee_name": employee_name, "reason": "found"})


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payroll_cpf_preview(request):
    employee_id = (request.GET.get("employee_id") or "").strip()
    basic_salary_raw = (request.GET.get("basic_salary") or "").strip()
    physical_products_commission_raw = (request.GET.get("physical_products_commission") or "").strip()
    credit_commission_raw = (request.GET.get("credit_commission") or "").strip()
    services_commission_raw = (request.GET.get("services_commission") or "").strip()
    payment_date_raw = (request.GET.get("payment_date") or "").strip()

    if not employee_id or not basic_salary_raw or not payment_date_raw:
        return JsonResponse(
            {
                "ok": False,
                "cpf_contribution": "",
                "employer_cpf_contribution": "",
                "reason": "missing_inputs",
            }
        )

    try:
        basic_salary = Decimal(basic_salary_raw)
        physical_products_commission = Decimal(physical_products_commission_raw or "0")
        credit_commission = Decimal(credit_commission_raw or "0")
        services_commission = Decimal(services_commission_raw or "0")
    except (InvalidOperation, ValueError):
        return JsonResponse(
            {
                "ok": False,
                "cpf_contribution": "",
                "employer_cpf_contribution": "",
                "reason": "invalid_salary",
            }
        )

    payment_date = _parse_date_ddmmyyyy(payment_date_raw)
    if payment_date is None:
        return JsonResponse(
            {
                "ok": False,
                "cpf_contribution": "",
                "employer_cpf_contribution": "",
                "reason": "invalid_payment_date",
            }
        )

    employee = Employee.objects.filter(employee_code=employee_id).first()
    if employee is None:
        return JsonResponse(
            {
                "ok": False,
                "cpf_contribution": "",
                "employer_cpf_contribution": "",
                "reason": "employee_not_found",
            }
        )
    if employee.cpf_exempt:
        return JsonResponse(
            {
                "ok": True,
                "cpf_contribution": "0.00",
                "employer_cpf_contribution": "0.00",
                "reason": "cpf_exempt",
            }
        )
    if not employee.date_of_birth:
        return JsonResponse(
            {
                "ok": False,
                "cpf_contribution": "",
                "employer_cpf_contribution": "",
                "reason": "missing_dob",
            }
        )

    total_earnings = basic_salary + physical_products_commission + credit_commission + services_commission
    age = payment_date.year - employee.date_of_birth.year - (
        (payment_date.month, payment_date.day) < (employee.date_of_birth.month, employee.date_of_birth.day)
    )
    cpf = cpf_for_2026(total_earnings, age)
    return JsonResponse(
        {
            "ok": True,
            "cpf_contribution": str(cpf.employee_amount),
            "employer_cpf_contribution": str(cpf.employer_amount),
            "reason": "calculated",
        }
    )


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payroll_dashboard(request):
    today = timezone.localdate()
    selected_month, month_start, month_end = _parse_month_filter(request.GET.get("month"))
    if not selected_month:
        month_start = today.replace(day=1)
        if month_start.month == 12:
            next_month = month_start.replace(year=month_start.year + 1, month=1, day=1)
        else:
            next_month = month_start.replace(month=month_start.month + 1, day=1)
        month_end = next_month - timezone.timedelta(days=1)
        selected_month = month_start.strftime("%Y-%m")

    month_records = PayrollRecord.objects.filter(payment_date__gte=month_start, payment_date__lte=month_end)
    total_records = month_records.count()
    total_net_salary = _safe_sum(month_records, "net_salary")
    total_payroll_cost = _safe_sum(month_records, "basic_salary") + _safe_sum(month_records, "allowances")
    total_cpf = _safe_sum(month_records, "cpf_contribution")
    employees_paid = month_records.values("employee_id").distinct().count()
    report_generated_at = timezone.now()

    month_record_id_set = set(month_records.values_list("id", flat=True))
    email_logs = EmailDeliveryLog.objects.filter(
        related_object_type="payroll_record",
        related_object_id__in=[str(record_id) for record_id in month_record_id_set],
    ).values("related_object_id", "status")
    emailed_ids = {
        int(log["related_object_id"])
        for log in email_logs
        if str(log["related_object_id"]).isdigit() and log["status"] == EmailDeliveryLog.STATUS_SENT
    }
    failed_email_ids = {
        int(log["related_object_id"])
        for log in email_logs
        if str(log["related_object_id"]).isdigit() and log["status"] == EmailDeliveryLog.STATUS_FAILED
    }
    downloaded_ids = {
        int(target_id)
        for target_id in AuditLog.objects.filter(
            action="payroll.pdf.downloaded",
            target_type="payroll_record",
            target_id__in=[str(record_id) for record_id in month_record_id_set],
        ).values_list("target_id", flat=True)
        if str(target_id).isdigit()
    }

    delivery_status_totals = {
        "emailed": 0,
        "downloaded": 0,
        "failed": 0,
        "action_required": 0,
    }
    for record_id in month_record_id_set:
        if record_id in emailed_ids:
            delivery_status_totals["emailed"] += 1
        elif record_id in downloaded_ids:
            delivery_status_totals["downloaded"] += 1
        elif record_id in failed_email_ids:
            delivery_status_totals["failed"] += 1
            delivery_status_totals["action_required"] += 1
        else:
            delivery_status_totals["action_required"] += 1

    trend_anchor_month = date(2026, 1, 1)
    month_starts = _month_starts_from_anchor(trend_anchor_month, month_start)
    payroll_trend_labels = [month.strftime("%b %Y") for month in month_starts]
    month_keys = [month.strftime("%Y-%m") for month in month_starts]
    monthly_rows = list(
        PayrollRecord.objects.annotate(month=TruncMonth("payment_date"))
        .values("month")
        .annotate(total_basic=Sum("basic_salary"), total_allowances=Sum("allowances"))
        .order_by("month")
    )
    payroll_amount_map = {}
    for row in monthly_rows:
        month_value = row.get("month")
        if not month_value:
            continue
        month_key = month_value.strftime("%Y-%m")
        payroll_amount_map[month_key] = _to_float(row.get("total_basic")) + _to_float(row.get("total_allowances"))
    payroll_trend_values = [payroll_amount_map.get(month_key, 0.0) for month_key in month_keys]

    pay_mix_labels = ["Basic Salary", "Allowances", "CPF", "Net Salary"]
    pay_mix_values = [
        _to_float(_safe_sum(month_records, "basic_salary")),
        _to_float(_safe_sum(month_records, "allowances")),
        _to_float(total_cpf),
        _to_float(total_net_salary),
    ]

    latest_upload_preview = (
        AuditLog.objects.filter(action="payroll.upload.previewed")
        .only("metadata", "created_at")
        .first()
    )
    invalid_upload_rows_count = 0
    if latest_upload_preview:
        invalid_upload_rows_count = int((latest_upload_preview.metadata or {}).get("invalid_row_count") or 0)

    recent_saved_upload_logs = list(
        AuditLog.objects.filter(action="payroll.upload.saved")
        .only("metadata", "created_at")[:20]
    )
    saved_upload_for_selected_month = next(
        (
            log
            for log in recent_saved_upload_logs
            if str((log.metadata or {}).get("payment_date") or "").startswith(selected_month)
        ),
        None,
    )
    duplicate_rows_skipped_count = 0
    if saved_upload_for_selected_month:
        duplicate_rows_skipped_count = int((saved_upload_for_selected_month.metadata or {}).get("skipped_duplicate_count") or 0)

    active_employees = Employee.objects.filter(status=Employee.STATUS_ACTIVE)
    paid_employee_codes = list(month_records.values_list("employee_id", flat=True).distinct())
    missing_payroll_records = active_employees.exclude(employee_code__in=paid_employee_codes).order_by("employee_code")
    missing_payroll_records_count = missing_payroll_records.count()
    missing_payroll_sample_codes = list(missing_payroll_records.values_list("employee_code", flat=True)[:3])

    employees = Employee.objects.all()
    inactive_employees = employees.filter(status=Employee.STATUS_INACTIVE)
    new_employees_this_month = employees.filter(hire_date__gte=month_start, hire_date__lte=month_end)
    missing_payment_setup = active_employees.filter(
        Q(payment_method="")
        | Q(bank_name="")
        | Q(bank_account_number="")
    )

    payment_method_counts = {
        row["payment_method"] or "unset": row["total"]
        for row in employees.values("payment_method").annotate(total=Count("id"))
    }
    payment_method_labels = ["GIRO", "Cash", "Cheque", "Not Set"]
    payment_method_values = [
        payment_method_counts.get(Employee.PAYMENT_METHOD_GIRO, 0),
        payment_method_counts.get(Employee.PAYMENT_METHOD_CASH, 0),
        payment_method_counts.get(Employee.PAYMENT_METHOD_CHEQUE, 0),
        payment_method_counts.get("unset", 0),
    ]
    giro_employees = employees.filter(payment_method=Employee.PAYMENT_METHOD_GIRO)
    cash_employees = employees.filter(payment_method=Employee.PAYMENT_METHOD_CASH)
    cheque_employees = employees.filter(payment_method=Employee.PAYMENT_METHOD_CHEQUE)
    unset_payment_employees = employees.filter(Q(payment_method="") | Q(payment_method__isnull=True))

    employee_status_labels = ["Active", "Inactive"]
    employee_status_values = [
        active_employees.count(),
        inactive_employees.count(),
    ]

    hiring_month_starts = _recent_month_starts(month_start, total_months=6)
    hiring_trend_labels = [month.strftime("%b %Y") for month in hiring_month_starts]
    hiring_month_keys = [month.strftime("%Y-%m") for month in hiring_month_starts]
    hiring_rows = list(
        employees.annotate(month=TruncMonth("hire_date"))
        .values("month")
        .annotate(total=Count("id"))
        .order_by("month")
    )
    hiring_map = {}
    for row in hiring_rows:
        month_value = row.get("month")
        if not month_value:
            continue
        hiring_map[month_value.strftime("%Y-%m")] = int(row.get("total") or 0)
    hiring_trend_values = [hiring_map.get(month_key, 0) for month_key in hiring_month_keys]

    recent_payslip_records = list(month_records.only(
        "employee_name",
        "employee_id",
        "payment_date",
        "net_salary",
        "deductions",
        "created_at",
    )[:8])

    recent_action_records = [{"record": record} for record in recent_payslip_records]

    payroll_chart_summary = _build_chart_summary(payroll_trend_labels, payroll_trend_values)
    reporting_period_label = month_start.strftime("%B %Y")
    secondary_summary_items = [
        {
            "label": "Emailed Payslips",
            "value": str(delivery_status_totals["emailed"]),
            "note": "Payslip records with a successful email delivery log in the selected month.",
        },
        {
            "label": "Downloaded Payslips",
            "value": str(delivery_status_totals["downloaded"]),
            "note": "Payroll records downloaded as PDF for follow-up or manual distribution.",
        },
        {
            "label": "Failed Emails",
            "value": str(delivery_status_totals["failed"]),
            "note": "Payslip emails that failed and may need corrected addresses or a resend.",
        },
    ]
    attention_items = []
    if invalid_upload_rows_count:
        attention_items.append(
            {
                "priority_label": "Review",
                "priority_class": "status-warning",
                "issue": "Invalid upload rows",
                "count": invalid_upload_rows_count,
                "scope": "Latest upload preview",
                "detail": "Rows failed validation before payroll records could be saved.",
                "action_label": "Upload payroll file",
                "action_url": reverse("payroll-upload-preview"),
            }
        )
    if duplicate_rows_skipped_count:
        attention_items.append(
            {
                "priority_label": "Review",
                "priority_class": "status-neutral",
                "issue": "Duplicate rows skipped",
                "count": duplicate_rows_skipped_count,
                "scope": reporting_period_label,
                "detail": "Rows matched payroll records already saved for the employee and selected payroll month.",
                "action_label": "View existing records",
                "action_url": f'{reverse("payroll-list")}?month={selected_month}',
            }
        )
    if delivery_status_totals["failed"]:
        attention_items.append(
            {
                "priority_label": "High",
                "priority_class": "status-danger",
                "issue": "Failed payslip emails",
                "count": delivery_status_totals["failed"],
                "scope": reporting_period_label,
                "detail": "Employees may not have received their payslips because email delivery failed.",
                "action_label": "Open payroll records",
                "action_url": f'{reverse("payroll-list")}?month={selected_month}',
            }
        )
    if missing_payroll_records_count:
        attention_items.append(
            {
                "priority_label": "Review",
                "priority_class": "status-warning",
                "issue": "Missing payroll records",
                "count": missing_payroll_records_count,
                "scope": reporting_period_label,
                "detail": "Active employees without a payroll record this month: "
                + ", ".join(missing_payroll_sample_codes)
                + ("..." if missing_payroll_records_count > len(missing_payroll_sample_codes) else ""),
                "action_label": "Review employees",
                "action_url": reverse("employee-list"),
            }
        )

    return render(
        request,
        "payroll/payroll_dashboard.html",
        {
            "selected_month": selected_month,
            "reporting_period_label": reporting_period_label,
            "report_generated_at": report_generated_at,
            "dashboard_title": "Payroll Officer Dashboard",
            "total_records": total_records,
            "total_net_salary": total_net_salary,
            "total_payroll_cost": total_payroll_cost,
            "total_cpf": total_cpf,
            "employees_paid": employees_paid,
            "active_employee_count": active_employees.count(),
            "new_employee_count": new_employees_this_month.count(),
            "missing_payment_setup_count": missing_payment_setup.count(),
            "employees_without_payroll_count": missing_payroll_records_count,
            "payroll_trend_labels": payroll_trend_labels,
            "payroll_trend_values": payroll_trend_values,
            "payroll_chart_summary": payroll_chart_summary,
            "pay_mix_labels": pay_mix_labels,
            "pay_mix_values": pay_mix_values,
            "payment_method_labels": payment_method_labels,
            "payment_method_values": payment_method_values,
            "giro_employees": giro_employees.order_by("first_name", "last_name", "employee_code"),
            "cash_employees": cash_employees.order_by("first_name", "last_name", "employee_code"),
            "cheque_employees": cheque_employees.order_by("first_name", "last_name", "employee_code"),
            "unset_payment_employees": unset_payment_employees.order_by("first_name", "last_name", "employee_code"),
            "employee_status_labels": employee_status_labels,
            "employee_status_values": employee_status_values,
            "active_employees": active_employees.order_by("first_name", "last_name", "employee_code"),
            "inactive_employees": inactive_employees.order_by("first_name", "last_name", "employee_code"),
            "hiring_trend_labels": hiring_trend_labels,
            "hiring_trend_values": hiring_trend_values,
            "delivery_status_totals": delivery_status_totals,
            "secondary_summary_items": secondary_summary_items,
            "attention_items": attention_items,
            "payroll_list_query": f'?month={selected_month}',
            "payroll_report_query": f'?month={selected_month}',
            "recent_action_records": recent_action_records,
        },
    )


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payroll_list(request):
    search_query = request.GET.get("q", "").strip()
    raw_date_from = (request.GET.get("date_from") or "").strip()
    raw_date_to = (request.GET.get("date_to") or "").strip()
    date_from, date_to, filter_date_from, filter_date_to, date_filter_error = _parse_date_range(
        raw_date_from,
        raw_date_to,
    )
    payslip_records = PayrollRecord.objects.all()
    if search_query:
        payslip_records = payslip_records.filter(
            Q(employee_name__icontains=search_query)
            | Q(employee_id__icontains=search_query)
        )
    if not date_filter_error:
        if filter_date_from:
            payslip_records = payslip_records.filter(payment_date__gte=filter_date_from)
        if filter_date_to:
            payslip_records = payslip_records.filter(payment_date__lte=filter_date_to)

    return render(
        request,
        "payroll/payroll_list.html",
        {
            "payslip_records": payslip_records,
            "search_query": search_query,
            "date_from": date_from.isoformat() if date_from else raw_date_from,
            "date_to": date_to.isoformat() if date_to else raw_date_to,
            "date_filter_error": date_filter_error,
            "result_count": payslip_records.count(),
        },
    )


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payroll_create(request):
    if request.method == "POST":
        form = PayrollRecordForm(request.POST)
        if form.is_valid():
            payslip_record = form.save(commit=False)
            employee = Employee.objects.get(employee_code=payslip_record.employee_id)
            if not employee.date_of_birth:
                form.add_error(
                    "employee_id",
                    "Selected employee is missing Date of Birth. CPF cannot be auto-calculated.",
                )
                return render(
                    request,
                    "payroll/payroll_form.html",
                    {"form": form, "is_edit": False, "payslip_record": None},
                )
            payslip_record.employee_name = f"{employee.first_name} {employee.last_name}".strip()
            payslip_record.nric = (employee.nric or "")[:9]
            payslip_record.cpf_exempted = employee.cpf_exempt
            payslip_record.sdl_exempted = employee.sdl_exempt
            payslip_record.physical_products_commission = (
                form.cleaned_data.get("physical_products_commission") or Decimal("0")
            )
            payslip_record.credit_commission = form.cleaned_data.get("credit_commission") or Decimal("0")
            payslip_record.services_commission = form.cleaned_data.get("services_commission") or Decimal("0")
            payslip_record.loan_deduction = form.cleaned_data.get("loan_deduction") or Decimal("0")
            payslip_record.other_deductions = form.cleaned_data.get("other_deductions") or Decimal("0")
            payslip_record.allowances = (
                payslip_record.physical_products_commission
                + payslip_record.credit_commission
                + payslip_record.services_commission
            )
            payslip_record.deductions = payslip_record.loan_deduction + payslip_record.other_deductions
            total_earnings = payslip_record.basic_salary + payslip_record.allowances
            if employee.cpf_exempt:
                employee_cpf = Decimal("0")
            else:
                age = payslip_record.payment_date.year - employee.date_of_birth.year - (
                    (payslip_record.payment_date.month, payslip_record.payment_date.day)
                    < (employee.date_of_birth.month, employee.date_of_birth.day)
                )
                employee_cpf = cpf_for_2026(total_earnings, age).employee_amount
            payslip_record.cpf_contribution = employee_cpf
            payslip_record.net_salary = (
                total_earnings
                - payslip_record.deductions
                - employee_cpf
            )
            payslip_record.created_by = request.user
            try:
                with transaction.atomic():
                    payslip_record.save()
            except IntegrityError:
                form.add_error(None, PayrollRecordForm.DUPLICATE_ERROR)
                return render(
                    request,
                    "payroll/payroll_form.html",
                    {"form": form, "is_edit": False, "payslip_record": None},
                )
            log_event(
                action="payroll.record.created",
                user=request.user,
                target_type="payroll_record",
                target_id=str(payslip_record.id),
                metadata={"employee_id": payslip_record.employee_id},
                ip_address=get_client_ip(request),
            )
            messages.success(request, "Payslip record created.")
            return redirect("payroll-detail", pk=payslip_record.pk)
    else:
        form = PayrollRecordForm()

    return render(
        request,
        "payroll/payroll_form.html",
        {"form": form, "is_edit": False, "payslip_record": None},
    )


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payroll_edit(request, pk):
    payslip_record = get_object_or_404(PayrollRecord, pk=pk)
    if request.method == "POST":
        form = PayrollRecordForm(request.POST, instance=payslip_record)
        if form.is_valid():
            payslip_record = form.save(commit=False)
            employee = Employee.objects.get(employee_code=payslip_record.employee_id)
            if not employee.date_of_birth:
                form.add_error(
                    "employee_id",
                    "Selected employee is missing Date of Birth. CPF cannot be auto-calculated.",
                )
                return render(
                    request,
                    "payroll/payroll_form.html",
                    {"form": form, "is_edit": True, "payslip_record": payslip_record},
                )
            payslip_record.employee_name = f"{employee.first_name} {employee.last_name}".strip()
            payslip_record.nric = (employee.nric or "")[:9]
            payslip_record.cpf_exempted = employee.cpf_exempt
            payslip_record.sdl_exempted = employee.sdl_exempt
            payslip_record.physical_products_commission = (
                form.cleaned_data.get("physical_products_commission") or Decimal("0")
            )
            payslip_record.credit_commission = form.cleaned_data.get("credit_commission") or Decimal("0")
            payslip_record.services_commission = form.cleaned_data.get("services_commission") or Decimal("0")
            payslip_record.loan_deduction = form.cleaned_data.get("loan_deduction") or Decimal("0")
            payslip_record.other_deductions = form.cleaned_data.get("other_deductions") or Decimal("0")
            payslip_record.allowances = (
                payslip_record.physical_products_commission
                + payslip_record.credit_commission
                + payslip_record.services_commission
            )
            payslip_record.deductions = payslip_record.loan_deduction + payslip_record.other_deductions
            total_earnings = payslip_record.basic_salary + payslip_record.allowances
            if employee.cpf_exempt:
                employee_cpf = Decimal("0")
            else:
                age = payslip_record.payment_date.year - employee.date_of_birth.year - (
                    (payslip_record.payment_date.month, payslip_record.payment_date.day)
                    < (employee.date_of_birth.month, employee.date_of_birth.day)
                )
                employee_cpf = cpf_for_2026(total_earnings, age).employee_amount
            payslip_record.cpf_contribution = employee_cpf
            payslip_record.net_salary = (
                total_earnings
                - payslip_record.deductions
                - employee_cpf
            )
            try:
                with transaction.atomic():
                    payslip_record.save()
            except IntegrityError:
                form.add_error(None, PayrollRecordForm.DUPLICATE_ERROR)
                return render(
                    request,
                    "payroll/payroll_form.html",
                    {"form": form, "is_edit": True, "payslip_record": payslip_record},
                )
            log_event(
                action="payroll.record.updated",
                user=request.user,
                target_type="payroll_record",
                target_id=str(payslip_record.id),
                metadata={"employee_id": payslip_record.employee_id},
                ip_address=get_client_ip(request),
            )
            messages.success(request, "Payslip record updated.")
            return redirect("payroll-detail", pk=payslip_record.pk)
    else:
        form = PayrollRecordForm(instance=payslip_record)

    return render(
        request,
        "payroll/payroll_form.html",
        {"form": form, "is_edit": True, "payslip_record": payslip_record},
    )


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payroll_detail(request, pk):
    payslip_record = get_object_or_404(PayrollRecord, pk=pk)
    employer_cpf_contribution = _calculate_employer_cpf_for_record(payslip_record)
    return render(
        request,
        "payroll/payroll_detail.html",
        {
            "payslip_record": payslip_record,
            "employer_cpf_contribution": employer_cpf_contribution,
        },
    )


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payroll_delete(request, pk):
    if request.method != "POST":
        return redirect("payroll-detail", pk=pk)

    payslip_record = get_object_or_404(PayrollRecord, pk=pk)
    record_id = payslip_record.id
    employee_id = payslip_record.employee_id
    payslip_record.delete()
    log_event(
        action="payroll.record.deleted",
        user=request.user,
        target_type="payroll_record",
        target_id=str(record_id),
        metadata={"employee_id": employee_id},
        ip_address=get_client_ip(request),
    )
    messages.success(request, "Payslip record deleted permanently.")
    return redirect("payroll-list")


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payroll_upload_preview(request):
    form = PayrollUploadForm(request.POST or None, request.FILES or None)
    preview_rows = []
    invalid_rows = []
    total_rows = 0
    valid_count = 0
    invalid_count = 0

    if request.method == "POST" and form.is_valid():
        try:
            upload_result = parse_and_validate_payroll_excel(
                form.cleaned_data["payroll_file"],
                form.cleaned_data["payment_date"],
            )
            preview_rows = upload_result["valid_rows"]
            invalid_rows = upload_result["invalid_rows"]
            employee_lookup = {
                employee.employee_code: employee
                for employee in Employee.objects.filter(
                    employee_code__in=[str(row.get("employee_code") or "").strip() for row in preview_rows]
                )
            }
            filtered_valid_rows = []
            for row in preview_rows:
                employee_code = str(row.get("employee_code") or "").strip()
                employee = employee_lookup.get(employee_code)
                row_errors = []
                if employee is None:
                    row_errors.append("Employee code not found in employee records.")
                else:
                    expected_name = f"{employee.first_name} {employee.last_name}".strip()
                    if expected_name and str(row.get("employee_name") or "").strip() != expected_name:
                        row_errors.append("Employee name does not match employee records.")
                    uploaded_dob = _parse_date_ddmmyyyy(row.get("employee_birthofdate"))
                    if employee.date_of_birth and uploaded_dob and employee.date_of_birth != uploaded_dob:
                        row_errors.append("Employee birthofdate does not match employee records.")
                    if PayrollRecord.objects.filter(
                        employee_id=employee_code,
                        payment_date=form.cleaned_data["payment_date"],
                    ).exists():
                        row_errors.append("A payroll record already exists for this employee and payment date.")

                if row_errors:
                    invalid_rows.append(
                        {
                            "row_number": row.get("row_number"),
                            "employee_code": employee_code,
                            "employee_name": row.get("employee_name", ""),
                            "errors": row_errors,
                            "raw_values": row.get("__raw_values", {}),
                        }
                    )
                else:
                    filtered_valid_rows.append(row)
            preview_rows = filtered_valid_rows
            total_rows = upload_result["total_rows"]
            valid_count = len(preview_rows)
            invalid_count = len(invalid_rows)

            if not preview_rows and not invalid_rows:
                messages.warning(request, "No data rows were found in the uploaded file.")
            else:
                request.session["payroll_upload_preview"] = {
                    "payment_date": form.cleaned_data["payment_date"].isoformat(),
                    "rows": [_serialize_preview_row(r) for r in preview_rows],
                    "invalid_rows": [_serialize_invalid_preview_row(r) for r in invalid_rows],
                }
            log_event(
                action="payroll.upload.previewed",
                user=request.user,
                metadata={
                    "path": request.path,
                    "row_count": total_rows,
                    "valid_row_count": valid_count,
                    "invalid_row_count": invalid_count,
                    "source_file_name": form.cleaned_data["payroll_file"].name,
                },
                ip_address=get_client_ip(request),
            )
        except Exception as exc:
            log_event(
                action="payroll.upload.failed",
                user=request.user,
                metadata={"path": request.path, "error_message": str(exc)},
                ip_address=get_client_ip(request),
            )
            messages.error(request, f"Unable to process file: {exc}")
            return redirect("payroll-upload-preview")

    return render(
        request,
        "payroll/upload_preview.html",
        {
            "form": form,
            "preview_rows": preview_rows,
            "invalid_rows": invalid_rows,
            "total_rows": total_rows,
            "valid_count": valid_count,
            "invalid_count": invalid_count,
        },
    )


def _serialize_preview_row(row):
    return {
        "employee_code": str(row["employee_code"]),
        "employee_name": str(row["employee_name"]),
        "basic_salary": str(row["basic_salary"]),
        "physical_products_commission": str(row["physical_products_commission"]),
        "credit_commission": str(row["credit_commission"]),
        "services_commission": str(row["services_commission"]),
        "loan_deduction": str(row["loan_deduction"]),
        "other_deductions": str(row["other_deductions"]),
        "cpf_employee_amount": str(row["cpf_employee_amount"]),
        "net_pay": str(row["net_pay"]),
    }


def _serialize_invalid_preview_row(row):
    raw_values = row.get("raw_values") or {}
    return {
        "row_number": row.get("row_number"),
        "employee_code": str(row.get("employee_code", "")),
        "employee_name": str(row.get("employee_name", "")),
        "errors": [str(error) for error in row.get("errors", [])],
        "raw_values": {
            header: str(raw_values.get(header, ""))
            for header in TEMPLATE_HEADERS
        },
    }


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def download_invalid_rows(request):
    upload_data = request.session.get("payroll_upload_preview") or {}
    invalid_rows = upload_data.get("invalid_rows") or []

    if not upload_data:
        messages.error(request, "Payroll upload session expired. Please upload the file again.")
        return redirect("payroll-upload-preview")

    if not invalid_rows:
        messages.warning(request, "There are no invalid payroll rows to download.")
        return redirect("payroll-upload-preview")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invalid Payroll Rows"
    worksheet.append(
        ["Original Excel Row Number"]
        + [PAYROLL_UPLOAD_COLUMN_LABELS[header] for header in TEMPLATE_HEADERS]
        + ["Error Reason"]
    )

    for row in invalid_rows:
        raw_values = row.get("raw_values") or {}
        worksheet.append(
            [row.get("row_number", "")]
            + [raw_values.get(header, "") for header in TEMPLATE_HEADERS]
            + [", ".join(row.get("errors", []))]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    log_event(
        action="payroll.upload.invalid_rows_downloaded",
        user=request.user,
        metadata={"invalid_row_count": len(invalid_rows)},
        ip_address=get_client_ip(request),
    )

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="payroll_invalid_rows.xlsx"'
    return response


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payroll_upload_confirm_save(request):
    if request.method != "POST":
        return redirect("payroll-upload-preview")

    upload_data = request.session.get("payroll_upload_preview") or {}
    payment_date_raw = upload_data.get("payment_date")
    serialized_rows = upload_data.get("rows") or []
    serialized_invalid_rows = upload_data.get("invalid_rows") or []
    duplicate_error_message = "A payroll record already exists for this employee and payment date."

    duplicate_only_preview = (
        bool(payment_date_raw)
        and not serialized_rows
        and bool(serialized_invalid_rows)
        and all(
            (row.get("errors") or []) == [duplicate_error_message]
            for row in serialized_invalid_rows
        )
    )

    if not payment_date_raw or (not serialized_rows and not duplicate_only_preview):
        messages.error(request, "No valid upload preview found. Please upload and preview first.")
        return redirect("payroll-upload-preview")

    payment_date = date.fromisoformat(payment_date_raw)
    if duplicate_only_preview:
        skipped_count = len(serialized_invalid_rows)
        request.session.pop("payroll_upload_preview", None)
        log_event(
            action="payroll.upload.saved",
            user=request.user,
            metadata={
                "saved_count": 0,
                "skipped_duplicate_count": skipped_count,
                "payment_date": payment_date.isoformat(),
            },
            ip_address=get_client_ip(request),
        )
        messages.warning(
            request,
            f"No new payroll records were created. {skipped_count} duplicate record(s) were skipped.",
        )
        return redirect("payroll-list")

    saved_count = 0
    skipped_count = 0
    invalid_format_count = 0

    for row in serialized_rows:
        employee_code = str(row["employee_code"]).strip()
        if not EMPLOYEE_CODE_PATTERN.fullmatch(employee_code):
            invalid_format_count += 1
            continue

        try:
            with transaction.atomic():
                if PayrollRecord.objects.filter(employee_id=employee_code, payment_date=payment_date).exists():
                    skipped_count += 1
                    continue

                allowances = (
                    Decimal(row["physical_products_commission"])
                    + Decimal(row["credit_commission"])
                    + Decimal(row["services_commission"])
                )
                deductions = Decimal(row["loan_deduction"]) + Decimal(row["other_deductions"])
                record = PayrollRecord(
                    employee_name=row["employee_name"],
                    employee_id=employee_code,
                    basic_salary=Decimal(row["basic_salary"]),
                    physical_products_commission=Decimal(row["physical_products_commission"]),
                    credit_commission=Decimal(row["credit_commission"]),
                    services_commission=Decimal(row["services_commission"]),
                    allowances=allowances,
                    loan_deduction=Decimal(row["loan_deduction"]),
                    other_deductions=Decimal(row["other_deductions"]),
                    deductions=deductions,
                    cpf_contribution=Decimal(row["cpf_employee_amount"]),
                    net_salary=Decimal(row["net_pay"]),
                    payment_date=payment_date,
                    created_by=request.user,
                )
                employee = Employee.objects.filter(employee_code=record.employee_id).first()
                if employee:
                    record.nric = (employee.nric or "")[:9]
                    record.cpf_exempted = employee.cpf_exempt
                    record.sdl_exempted = employee.sdl_exempt
                record.save()
        except IntegrityError:
            skipped_count += 1
            continue

        saved_count += 1

    request.session.pop("payroll_upload_preview", None)
    log_event(
        action="payroll.upload.saved",
        user=request.user,
        metadata={
            "saved_count": saved_count,
            "skipped_duplicate_count": skipped_count,
            "payment_date": payment_date.isoformat(),
        },
        ip_address=get_client_ip(request),
    )
    if saved_count > 0 and skipped_count > 0:
        messages.success(
            request,
            f"Payroll upload saved successfully. {saved_count} record(s) created, {skipped_count} duplicate record(s) skipped.",
        )
    elif saved_count > 0:
        messages.success(request, f"Payroll upload saved successfully. {saved_count} record(s) created.")
    elif skipped_count > 0:
        messages.warning(
            request,
            f"No new payroll records were created. {skipped_count} duplicate record(s) were skipped.",
        )
    else:
        messages.warning(request, "No payroll records were created from this upload.")
    if invalid_format_count:
        messages.warning(
            request,
            f"{invalid_format_count} row(s) were skipped because employee code was not in STF-000000 format.",
        )
    return redirect("payroll-list")


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payroll_template_download(request):
    template_path = Path(__file__).resolve().parent / "payroll_upload_template.xlsx"
    response = FileResponse(
        template_path.open("rb"),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="payroll_upload_template.xlsx"'
    return response


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def employee_dashboard(request):
    month_query = (request.GET.get("month") or "").strip()
    dashboard_url = reverse("payroll-dashboard")
    if month_query:
        return redirect(f"{dashboard_url}?month={month_query}")
    return redirect("payroll-dashboard")


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def employee_list(request):
    search_query = request.GET.get("q", "").strip()
    selected_status = request.GET.get("status", "").strip()
    selected_payment_method = request.GET.get("payment_method", "").strip()
    employees = Employee.objects.all()
    if selected_status in {Employee.STATUS_ACTIVE, Employee.STATUS_INACTIVE}:
        employees = employees.filter(status=selected_status)
    valid_payment_methods = {
        Employee.PAYMENT_METHOD_GIRO,
        Employee.PAYMENT_METHOD_CASH,
        Employee.PAYMENT_METHOD_CHEQUE,
        "unset",
    }
    if selected_payment_method in valid_payment_methods:
        if selected_payment_method == "unset":
            employees = employees.filter(Q(payment_method="") | Q(payment_method__isnull=True))
        else:
            employees = employees.filter(payment_method=selected_payment_method)
    if search_query:
        employees = employees.filter(
            Q(employee_code__icontains=search_query)
            | Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(email__icontains=search_query)
        )
    return render(
        request,
        "payroll/employee_list.html",
        {
            "employees": employees,
            "search_query": search_query,
            "result_count": employees.count(),
            "selected_status": selected_status,
            "selected_payment_method": selected_payment_method,
            "active_status_value": Employee.STATUS_ACTIVE,
            "inactive_status_value": Employee.STATUS_INACTIVE,
            "giro_payment_value": Employee.PAYMENT_METHOD_GIRO,
            "cash_payment_value": Employee.PAYMENT_METHOD_CASH,
            "cheque_payment_value": Employee.PAYMENT_METHOD_CHEQUE,
            "unset_payment_value": "unset",
        },
    )


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def employee_create(request):
    if request.method == "POST":
        post_data = request.POST.copy()
        if not (post_data.get("employee_code") or "").strip():
            post_data["employee_code"] = _generate_next_employee_code()
        form = EmployeeForm(post_data)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.hire_date = employee.date_of_appointment or employee.hire_date
            employee.status = Employee.STATUS_ACTIVE
            employee.created_by = request.user
            employee.save()
            log_event(
                action="payroll.employee.created",
                user=request.user,
                target_type="employee",
                target_id=str(employee.id),
                metadata={"employee_code": employee.employee_code},
                ip_address=get_client_ip(request),
            )
            messages.success(request, "Employee saved successfully.")
            return redirect("employee-list")
    else:
        form = EmployeeForm(initial={"employee_code": _generate_next_employee_code()})
    return render(request, "payroll/employee_form.html", {"form": form, "is_edit": False})


def _serialize_employee_preview_row(row):
    return {
        "row_number": row["row_number"],
        "employee_code": row["employee_code"],
        "nric": row["nric"],
        "first_name": row["first_name"],
        "last_name": row["last_name"],
        "date_of_birth": row["date_of_birth"],
        "date_of_appointment": row["date_of_appointment"],
        "legal_status": row["legal_status"],
        "gender": row["gender"],
        "race": row["race"],
        "religion": row["religion"],
        "sdl_exempt": row["sdl_exempt"],
        "cpf_exempt": row["cpf_exempt"],
        "job_title": row["job_title"],
        "email": row["email"],
        "payment_method": row["payment_method"],
        "bank_name": row["bank_name"],
        "bank_account_number": row["bank_account_number"],
        "bank_branch_code": row["bank_branch_code"],
    }


def _deserialize_employee_preview_row(row):
    parsed = dict(row)
    for key in ("date_of_birth", "date_of_appointment"):
        if parsed.get(key):
            parsed_date = _parse_date_ddmmyyyy(parsed[key])
            parsed[key] = parsed_date if parsed_date else None
    return parsed


def _parse_date_ddmmyyyy(raw_value):
    if raw_value in (None, ""):
        return None
    if isinstance(raw_value, datetime):
        return raw_value.date()
    if isinstance(raw_value, date):
        return raw_value
    raw_text = str(raw_value).strip()
    if not raw_text:
        return None
    try:
        return datetime.strptime(raw_text, "%d-%m-%Y").date()
    except ValueError:
        return None


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def employee_upload_preview(request):
    form = EmployeeUploadForm(request.POST or None, request.FILES or None)
    preview_rows = []
    invalid_rows = []
    total_rows = 0
    valid_count = 0
    invalid_count = 0

    if request.method == "POST" and form.is_valid():
        try:
            uploaded_file = form.cleaned_data["employee_file"]
            from openpyxl import load_workbook

            worksheet = load_workbook(uploaded_file, data_only=True).active
            raw_rows = list(worksheet.iter_rows(values_only=True))
            if not raw_rows:
                raise ValueError("Uploaded file is empty.")

            headers = [str(h).strip().lower() if h is not None else "" for h in raw_rows[0]]
            required_headers = [
                "employee_code",
                "nric",
                "first_name",
                "last_name",
                "date_of_birth",
                "date_of_appointment",
                "legal_status",
                "gender",
                "race",
                "religion",
                "sdl_exempt",
                "cpf_exempt",
                "job_title",
                "email",
                "payment_method",
                "bank_name",
                "bank_account_number",
                "bank_branch_code",
            ]
            missing = sorted(set(required_headers) - set(headers))
            if missing:
                raise ValueError(f"Missing required columns: {', '.join(missing)}")

            index_map = {name: headers.index(name) for name in required_headers}
            valid_legal_status = {choice[0] for choice in Employee.LEGAL_STATUS_CHOICES}
            valid_gender = {choice[0] for choice in Employee.GENDER_CHOICES}
            valid_payment_method = {choice[0] for choice in Employee.PAYMENT_METHOD_CHOICES}
            used_employee_codes = set(Employee.objects.values_list("employee_code", flat=True))
            seen_upload_employee_codes = {}
            seen_upload_emails = {}

            for row_number, row in enumerate(raw_rows[1:], start=2):
                if row is None or all(value in (None, "") for value in row):
                    continue

                row_errors = []
                parsed_row = {}
                for header in required_headers:
                    value = row[index_map[header]]
                    if header in ("date_of_birth", "date_of_appointment"):
                        parsed_row[header] = value
                    else:
                        parsed_row[header] = str(value).strip() if value is not None else ""

                if not parsed_row["first_name"]:
                    row_errors.append("First name is required.")
                if not parsed_row["last_name"]:
                    row_errors.append("Last name is required.")
                if not parsed_row["email"]:
                    row_errors.append("Email is required.")
                elif "@" not in parsed_row["email"] or "." not in parsed_row["email"].split("@")[-1]:
                    row_errors.append("Email must be a valid email address.")

                for date_field in ("date_of_birth", "date_of_appointment"):
                    raw_value = parsed_row[date_field]
                    if not raw_value:
                        if date_field == "date_of_appointment":
                            row_errors.append("Date of appointment is required.")
                        parsed_row[date_field] = ""
                        continue
                    parsed_date = _parse_date_ddmmyyyy(raw_value)
                    if parsed_date is None:
                        row_errors.append(f"{date_field.replace('_', ' ').title()} must be DD-MM-YYYY.")
                    else:
                        parsed_row[date_field] = parsed_date.strftime("%d-%m-%Y")

                parsed_birth_date = _parse_date_ddmmyyyy(parsed_row.get("date_of_birth"))
                parsed_appointment_date = _parse_date_ddmmyyyy(parsed_row.get("date_of_appointment"))
                if parsed_birth_date and parsed_appointment_date and parsed_appointment_date < parsed_birth_date:
                    row_errors.append("Date of appointment cannot be earlier than date of birth.")

                for bool_field in ("sdl_exempt", "cpf_exempt"):
                    raw_value = parsed_row[bool_field].lower()
                    if raw_value in ("true", "1", "yes", "y"):
                        parsed_row[bool_field] = True
                    elif raw_value in ("false", "0", "no", "n", ""):
                        parsed_row[bool_field] = False
                    else:
                        row_errors.append(f"{bool_field} must be TRUE/FALSE.")

                if parsed_row["legal_status"] and parsed_row["legal_status"] not in valid_legal_status:
                    row_errors.append("Invalid legal_status value.")
                if parsed_row["gender"] and parsed_row["gender"] not in valid_gender:
                    row_errors.append("Invalid gender value.")
                if parsed_row["payment_method"] and parsed_row["payment_method"] not in valid_payment_method:
                    row_errors.append("Invalid payment_method value.")
                if parsed_row["employee_code"] and not EMPLOYEE_CODE_PATTERN.fullmatch(parsed_row["employee_code"]):
                    row_errors.append("Employee code must follow STF-000000 format.")
                if parsed_row["employee_code"] and Employee.objects.filter(employee_code=parsed_row["employee_code"]).exists():
                    row_errors.append("Employee code already exists.")
                if parsed_row["employee_code"]:
                    first_seen_row = seen_upload_employee_codes.get(parsed_row["employee_code"])
                    if first_seen_row is not None:
                        row_errors.append(
                            f"Employee code is duplicated in this upload file. First duplicate appears on row {first_seen_row}."
                        )
                    else:
                        seen_upload_employee_codes[parsed_row["employee_code"]] = row_number
                if Employee.objects.filter(email__iexact=parsed_row["email"]).exists():
                    row_errors.append("Email already exists in employee records.")
                if parsed_row["email"]:
                    email_key = parsed_row["email"].lower()
                    first_seen_row = seen_upload_emails.get(email_key)
                    if first_seen_row is not None:
                        row_errors.append(
                            f"Email is duplicated in this upload file. First duplicate appears on row {first_seen_row}."
                        )
                    else:
                        seen_upload_emails[email_key] = row_number
                if parsed_row["payment_method"] == Employee.PAYMENT_METHOD_GIRO:
                    if not parsed_row["bank_name"]:
                        row_errors.append("Bank name is required when payment method is GIRO.")
                    if not parsed_row["bank_account_number"]:
                        row_errors.append("Bank account number is required when payment method is GIRO.")
                    if not parsed_row["bank_branch_code"]:
                        row_errors.append("Bank branch code is required when payment method is GIRO.")

                if row_errors:
                    invalid_rows.append(
                        {
                            "row_number": row_number,
                            "employee_code": parsed_row["employee_code"],
                            "employee_name": f"{parsed_row['first_name']} {parsed_row['last_name']}".strip(),
                            "errors": row_errors,
                        }
                    )
                    continue

                parsed_row["row_number"] = row_number
                if not parsed_row["employee_code"]:
                    generated_code = _generate_next_employee_code()
                    while generated_code in used_employee_codes:
                        numeric_part = int(generated_code.split("-")[1]) + 1
                        generated_code = f"STF-{numeric_part:06d}"
                    parsed_row["employee_code"] = generated_code
                used_employee_codes.add(parsed_row["employee_code"])
                preview_rows.append(parsed_row)

            total_rows = len(preview_rows) + len(invalid_rows)
            valid_count = len(preview_rows)
            invalid_count = len(invalid_rows)

            if preview_rows:
                request.session["employee_upload_preview"] = {
                    "rows": [_serialize_employee_preview_row(r) for r in preview_rows],
                }

            log_event(
                action="employee.upload.previewed",
                user=request.user,
                metadata={
                    "path": request.path,
                    "row_count": total_rows,
                    "valid_row_count": valid_count,
                    "invalid_row_count": invalid_count,
                    "source_file_name": uploaded_file.name,
                },
                ip_address=get_client_ip(request),
            )
        except Exception as exc:
            log_event(
                action="employee.upload.failed",
                user=request.user,
                metadata={"path": request.path, "error_message": str(exc)},
                ip_address=get_client_ip(request),
            )
            messages.error(request, f"Unable to process file: {exc}")
            return redirect("employee-upload-preview")

    return render(
        request,
        "payroll/employee_upload_preview.html",
        {
            "form": form,
            "preview_rows": preview_rows,
            "invalid_rows": invalid_rows,
            "total_rows": total_rows,
            "valid_count": valid_count,
            "invalid_count": invalid_count,
        },
    )


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def employee_upload_confirm_save(request):
    if request.method != "POST":
        return redirect("employee-upload-preview")

    upload_data = request.session.get("employee_upload_preview") or {}
    serialized_rows = upload_data.get("rows") or []
    if not serialized_rows:
        messages.error(request, "No valid employee upload preview found. Please upload and preview first.")
        return redirect("employee-upload-preview")

    saved_count = 0
    skipped_count = 0
    invalid_format_count = 0
    with transaction.atomic():
        for row in serialized_rows:
            parsed = _deserialize_employee_preview_row(row)
            if not EMPLOYEE_CODE_PATTERN.fullmatch(parsed["employee_code"]):
                invalid_format_count += 1
                continue
            if Employee.objects.filter(employee_code=parsed["employee_code"]).exists() or Employee.objects.filter(
                email__iexact=parsed["email"]
            ).exists():
                skipped_count += 1
                continue

            employee = Employee(
                employee_code=parsed["employee_code"],
                nric=parsed["nric"],
                first_name=parsed["first_name"],
                last_name=parsed["last_name"],
                date_of_birth=parsed["date_of_birth"] or None,
                date_of_appointment=parsed["date_of_appointment"],
                legal_status=parsed["legal_status"],
                gender=parsed["gender"],
                race=parsed["race"],
                religion=parsed["religion"],
                sdl_exempt=parsed["sdl_exempt"],
                cpf_exempt=parsed["cpf_exempt"],
                job_title=parsed["job_title"],
                email=parsed["email"],
                payment_method=parsed["payment_method"],
                bank_name=parsed["bank_name"],
                bank_account_number=parsed["bank_account_number"],
                bank_branch_code=parsed["bank_branch_code"],
                hire_date=parsed["date_of_appointment"],
                status=Employee.STATUS_ACTIVE,
                created_by=request.user,
            )
            employee.save()
            saved_count += 1

    request.session.pop("employee_upload_preview", None)
    if invalid_format_count:
        messages.warning(
            request,
            f"{invalid_format_count} row(s) were skipped because employee code was not in STF-000000 format.",
        )
    messages.success(
        request,
        f"Employee upload saved. {saved_count} created, {skipped_count} skipped due to duplicates."
        if skipped_count
        else f"Employee upload saved. {saved_count} employee record(s) created.",
    )
    return redirect("employee-list")


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def employee_template_download(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Employee Template"
    headers = [
        "employee_code",
        "nric",
        "first_name",
        "last_name",
        "date_of_birth",
        "date_of_appointment",
        "legal_status",
        "gender",
        "race",
        "religion",
        "sdl_exempt",
        "cpf_exempt",
        "job_title",
        "email",
        "payment_method",
        "bank_name",
        "bank_account_number",
        "bank_branch_code",
    ]
    worksheet.append(headers)
    worksheet.append(
        [
            "STF-000001",
            "S1234567A",
            "Alex",
            "Tan",
            "01-01-1990",
            "10-01-2024",
            "citizen",
            "male",
            "Chinese",
            "Buddhist",
            "FALSE",
            "FALSE",
            "Therapist",
            "alex.tan@example.com",
            "giro",
            "DBS Bank",
            "123-456-789",
            "001",
        ]
    )
    # Keep date columns explicitly in DD-MM-YYYY format.
    worksheet["E2"].number_format = "DD-MM-YYYY"
    worksheet["F2"].number_format = "DD-MM-YYYY"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="employee_upload_template.xlsx"'
    return response


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def employee_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    return render(request, "payroll/employee_detail.html", {"employee": employee})


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def employee_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == "POST":
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.hire_date = employee.date_of_appointment or employee.hire_date
            employee.save()
            log_event(
                action="payroll.employee.updated",
                user=request.user,
                target_type="employee",
                target_id=str(employee.id),
                metadata={"employee_code": employee.employee_code},
                ip_address=get_client_ip(request),
            )
            messages.success(request, "Employee updated successfully.")
            return redirect("employee-list")
    else:
        form = EmployeeForm(instance=employee)
    return render(request, "payroll/employee_form.html", {"form": form, "is_edit": True, "employee": employee})


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def employee_delete(request, pk):
    if request.method != "POST":
        return redirect("employee-list")

    employee = get_object_or_404(Employee, pk=pk)
    employee_id = employee.id
    employee_code = employee.employee_code
    employee.delete()
    log_event(
        action="payroll.employee.deleted",
        user=request.user,
        target_type="employee",
        target_id=str(employee_id),
        metadata={"employee_code": employee_code},
        ip_address=get_client_ip(request),
    )
    messages.success(request, "Employee deleted permanently.")
    return redirect("employee-list")


@login_required
@role_required(STAFF)
def my_payslips(request):
    employee = _resolve_staff_employee(request.user)

    if employee is None:
        messages.error(
            request,
            "Your account is not linked to an employee profile yet. Ask admin to link your account in Employees.",
        )
        return render(
            request,
            "payroll/my_payslips.html",
            {
                "employee": None,
                "payslip_records": PayrollRecord.objects.none(),
            },
        )

    payslip_records = PayrollRecord.objects.filter(employee_id=employee.employee_code)
    return render(
        request,
        "payroll/my_payslips.html",
        {
            "employee": employee,
            "payslip_records": payslip_records,
        },
    )


def _resolve_staff_employee(user):
    employee = getattr(user, "employee_profile", None)
    if employee is not None:
        return employee
    user_email = (user.email or "").strip()
    if not user_email:
        return None
    email_matches = Employee.objects.filter(email__iexact=user_email)
    if email_matches.count() != 1:
        return None
    employee = email_matches.first()
    employee.user = user
    employee.save(update_fields=["user", "updated_at"])
    return employee


def _build_payslip_pdf(payslip_record: PayrollRecord) -> bytes:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    page_width, page_height = A4

    pdf.setTitle(f"Payslip-{payslip_record.employee_id}-{payslip_record.payment_date}")
    margin = 12 * mm
    width = page_width - (2 * margin)
    top = page_height - margin

    basic_salary = payslip_record.basic_salary or Decimal("0.00")
    physical_products_commission = payslip_record.physical_products_commission or Decimal("0.00")
    credit_commission = payslip_record.credit_commission or Decimal("0.00")
    services_commission = payslip_record.services_commission or Decimal("0.00")
    allowances = payslip_record.allowances or Decimal("0.00")
    loan_deduction = payslip_record.loan_deduction or Decimal("0.00")
    other_deductions = payslip_record.other_deductions or Decimal("0.00")
    deductions = payslip_record.deductions or Decimal("0.00")
    cpf_contribution = payslip_record.cpf_contribution or Decimal("0.00")
    net_salary = payslip_record.net_salary or Decimal("0.00")

    if (
        physical_products_commission == Decimal("0.00")
        and credit_commission == Decimal("0.00")
        and services_commission == Decimal("0.00")
    ):
        services_commission = allowances
    if loan_deduction == Decimal("0.00") and other_deductions == Decimal("0.00"):
        other_deductions = deductions

    total_earnings = basic_salary + allowances
    total_deductions = deductions + cpf_contribution
    employer_cpf_contribution = _calculate_employer_cpf_for_record(payslip_record)
    month_year = payslip_record.payment_date.strftime("%B %Y")
    border_color = colors.HexColor("#243041")
    divider_color = colors.HexColor("#CBD5E1")
    muted_text = colors.HexColor("#5B6677")
    heading_text = colors.HexColor("#132238")
    accent_fill = colors.HexColor("#F7F9FC")
    net_fill = colors.HexColor("#EEF6F1")
    currency_prefix = "S$"

    # Header
    header_h = 34 * mm
    pdf.setStrokeColor(colors.white)
    pdf.setLineWidth(0)
    pdf.roundRect(margin, top - header_h, width, header_h, 4 * mm, stroke=0, fill=0)
    logo_dir = settings.BASE_DIR / "media" / "invoice_branding" / "logos"
    logo_files = sorted(path for path in logo_dir.iterdir() if path.is_file()) if logo_dir.exists() else []
    logo_path = logo_files[0] if logo_files else None
    text_x = margin + 4 * mm
    if logo_path is not None:
        try:
            logo_reader = ImageReader(str(logo_path))
            image_width, image_height = logo_reader.getSize()
            logo_height = 14 * mm
            scale_ratio = logo_height / float(image_height)
            logo_width = image_width * scale_ratio
            logo_x = margin + 4 * mm
            logo_y = top - 16 * mm
            pdf.drawImage(
                logo_reader,
                logo_x,
                logo_y,
                width=logo_width,
                height=logo_height,
                preserveAspectRatio=True,
                mask="auto",
            )
            text_x = logo_x + logo_width + (5 * mm)
        except Exception:
            text_x = margin + 4 * mm
    pdf.setFillColor(heading_text)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(text_x, top - 10 * mm, "Vaniday Pte Ltd - Payslip")
    pdf.setFillColor(muted_text)
    pdf.setFont("Helvetica", 8.5)
    pdf.drawString(text_x, top - 15.8 * mm, "Automated Invoicing & Payroll")
    pdf.setFillColor(heading_text)
    pdf.setFont("Helvetica-Bold", 13.5)
    pdf.drawString(text_x, top - 22.5 * mm, f"{payslip_record.employee_name} for {month_year}")

    # Employee block
    info_top = top - header_h
    info_h = 24 * mm
    pdf.setFillColor(accent_fill)
    pdf.roundRect(margin, info_top - info_h, width, info_h, 3 * mm, stroke=0, fill=1)
    label_x = margin + 4 * mm
    value_x = margin + 32 * mm
    employee_y = info_top - 8.5 * mm
    payment_y = info_top - 17 * mm
    pdf.setFillColor(muted_text)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(label_x, employee_y, "Employee:")
    pdf.drawString(label_x, payment_y, "Payment Date:")
    pdf.setFillColor(heading_text)
    pdf.setFont("Helvetica", 11)
    pdf.drawString(value_x, employee_y, payslip_record.employee_name)
    pdf.drawString(value_x, payment_y, payslip_record.payment_date.strftime("%d-%m-%Y"))

    # Table
    table_top = info_top - info_h - 4 * mm
    table_h = 86 * mm
    half = width / 2
    pdf.setFillColor(colors.white)
    pdf.roundRect(margin, table_top - table_h, width, table_h, 3 * mm, stroke=0, fill=1)
    pdf.setStrokeColor(divider_color)
    pdf.setLineWidth(0.9)
    pdf.line(margin + half, table_top, margin + half, table_top - table_h)
    head_h = 14 * mm
    pdf.setFillColor(accent_fill)
    pdf.roundRect(margin, table_top - head_h, width, head_h, 0, stroke=0, fill=1)
    pdf.line(margin, table_top - head_h, margin + width, table_top - head_h)

    pdf.setFillColor(heading_text)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(margin + 3 * mm, table_top - 9 * mm, "Earnings")
    pdf.drawString(margin + half - 28 * mm, table_top - 9 * mm, "Amount")
    pdf.drawString(margin + half + 3 * mm, table_top - 9 * mm, "Deductions")
    pdf.drawString(margin + width - 28 * mm, table_top - 9 * mm, "Amount")

    left_amount_x = margin + half - 4 * mm
    right_amount_x = margin + width - 4 * mm
    y = table_top - head_h - 7.5 * mm
    pdf.setFillColor(heading_text)
    pdf.setFont("Helvetica", 10.5)

    pdf.drawString(margin + 3 * mm, y, "Basic salary")
    pdf.drawRightString(left_amount_x, y, f"{currency_prefix} {basic_salary:.2f}")
    y -= 8 * mm
    pdf.drawString(margin + 3 * mm, y, "Physical products commission")
    pdf.drawRightString(left_amount_x, y, f"{currency_prefix} {physical_products_commission:.2f}")
    y -= 8 * mm
    pdf.drawString(margin + 3 * mm, y, "Credit commission")
    pdf.drawRightString(left_amount_x, y, f"{currency_prefix} {credit_commission:.2f}")
    y -= 8 * mm
    pdf.drawString(margin + 3 * mm, y, "Services commission")
    pdf.drawRightString(left_amount_x, y, f"{currency_prefix} {services_commission:.2f}")

    y2 = table_top - head_h - 7.5 * mm
    pdf.drawString(margin + half + 3 * mm, y2, "Loan deduction")
    pdf.drawRightString(right_amount_x, y2, f"{currency_prefix} {loan_deduction:.2f}")
    y2 -= 8 * mm
    pdf.drawString(margin + half + 3 * mm, y2, "Other deductions")
    pdf.drawRightString(right_amount_x, y2, f"{currency_prefix} {other_deductions:.2f}")
    y2 -= 8 * mm
    pdf.drawString(margin + half + 3 * mm, y2, "Employee CPF")
    pdf.drawRightString(right_amount_x, y2, f"{currency_prefix} {cpf_contribution:.2f}")

    totals_line = table_top - table_h + 14 * mm
    pdf.line(margin, totals_line, margin + width, totals_line)
    pdf.setFillColor(heading_text)
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(margin + 3 * mm, totals_line - 9 * mm, "Total earnings:")
    pdf.drawRightString(left_amount_x, totals_line - 9 * mm, f"{currency_prefix} {total_earnings:.2f}")
    pdf.drawString(margin + half + 3 * mm, totals_line - 9 * mm, "Total deductions:")
    pdf.drawRightString(right_amount_x, totals_line - 9 * mm, f"{currency_prefix} {total_deductions:.2f}")

    # Net pay block
    net_top = table_top - table_h
    net_h = 20 * mm
    pdf.setFillColor(net_fill)
    pdf.roundRect(margin, net_top - net_h, width, net_h, 3 * mm, stroke=0, fill=1)
    pdf.setFillColor(heading_text)
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(margin + 5 * mm, net_top - 8 * mm, "Net pay")
    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawRightString(right_amount_x, net_top - 8 * mm, f"{currency_prefix} {net_salary:.2f}")
    pdf.setFillColor(muted_text)
    pdf.setFont("Helvetica", 9)
    pdf.drawRightString(right_amount_x, net_top - 15 * mm, f"Employer CPF: {currency_prefix} {employer_cpf_contribution:.2f}")

    # Note
    note_top = net_top - net_h
    note_h = 13 * mm
    pdf.setFillColor(colors.white)
    pdf.roundRect(margin, note_top - note_h, width, note_h, 3 * mm, stroke=0, fill=1)
    pdf.setFillColor(heading_text)
    pdf.setFont("Helvetica-Bold", 9.5)
    pdf.drawString(margin + 3 * mm, note_top - 6.5 * mm, "Note:")
    pdf.setFillColor(muted_text)
    pdf.setFont("Helvetica", 8.5)
    pdf.drawString(
        margin + 18 * mm,
        note_top - 6.5 * mm,
        "This is a computer-generated payslip. For payroll enquiries, please contact HR.",
    )

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer.getvalue()


@login_required
@role_required(SUPERADMIN, ADMIN, HR, STAFF)
def payslip_pdf_download(request, pk):
    payslip_record = get_object_or_404(PayrollRecord, pk=pk)
    role = get_user_role(request.user)

    if role == STAFF:
        employee = _resolve_staff_employee(request.user)
        if employee is None or payslip_record.employee_id != employee.employee_code:
            log_event(
                action="payroll.pdf.permission_denied",
                user=request.user,
                target_type="payroll_record",
                target_id=str(payslip_record.id),
                metadata={"employee_id": payslip_record.employee_id},
                ip_address=get_client_ip(request),
            )
            raise PermissionDenied("You can only download your own payslips.")

    pdf_bytes = _build_payslip_pdf(payslip_record)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="payslip_{payslip_record.employee_id}_{payslip_record.payment_date}.pdf"'
    )
    log_event(
        action="payroll.pdf.downloaded",
        user=request.user,
        target_type="payroll_record",
        target_id=str(payslip_record.id),
        metadata={"employee_id": payslip_record.employee_id, "role": role},
        ip_address=get_client_ip(request),
    )
    return response


def _calculate_employer_cpf_for_record(payslip_record: PayrollRecord) -> Decimal:
    employee = Employee.objects.filter(employee_code=payslip_record.employee_id).first()
    if not employee or employee.cpf_exempt or not employee.date_of_birth:
        return Decimal("0.00")
    total_earnings = payslip_record.basic_salary + payslip_record.allowances
    age = payslip_record.payment_date.year - employee.date_of_birth.year - (
        (payslip_record.payment_date.month, payslip_record.payment_date.day)
        < (employee.date_of_birth.month, employee.date_of_birth.day)
    )
    return cpf_for_2026(total_earnings, age).employer_amount


@login_required
@role_required(SUPERADMIN, ADMIN, HR)
def payslip_email_send(request, pk):
    if request.method != "POST":
        return redirect("payroll-detail", pk=pk)

    payslip_record = get_object_or_404(PayrollRecord, pk=pk)
    employee = Employee.objects.filter(employee_code=payslip_record.employee_id).first()
    recipient = ((employee.email if employee else "") or "").strip().lower()
    subject = f"Payslip for {payslip_record.payment_date:%Y-%m-%d} - {payslip_record.employee_id}"
    text_body = (
        f"Dear {payslip_record.employee_name},\n\n"
        f"Your payslip for {payslip_record.payment_date:%Y-%m-%d} is attached.\n"
        f"Net salary: SGD {payslip_record.net_salary:.2f}\n\n"
        "Regards,\nPayroll Team"
    )

    email_log = EmailDeliveryLog.objects.create(
        recipient_email=recipient,
        subject=subject,
        template_key="payroll_payslip_email_v1",
        status=EmailDeliveryLog.STATUS_PENDING,
        related_object_type="payroll_record",
        related_object_id=str(payslip_record.id),
        triggered_by=request.user,
        metadata={
            "employee_id": payslip_record.employee_id,
            "payment_date": payslip_record.payment_date.isoformat(),
        },
    )

    if not recipient:
        email_log.status = EmailDeliveryLog.STATUS_FAILED
        email_log.error_message = "Employee email is missing."
        email_log.save(update_fields=["status", "error_message"])
        log_event(
            action="payroll.email.failed",
            user=request.user,
            target_type="payroll_record",
            target_id=str(payslip_record.id),
            metadata={"reason": "missing_employee_email"},
            ip_address=get_client_ip(request),
        )
        messages.error(request, "Unable to send email: employee email is missing.")
        return redirect("payroll-detail", pk=pk)

    message = EmailMultiAlternatives(
        subject=subject,
        body=text_body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[recipient],
    )
    message.attach(
        filename=f"payslip_{payslip_record.employee_id}_{payslip_record.payment_date}.pdf",
        content=_build_payslip_pdf(payslip_record),
        mimetype="application/pdf",
    )
    try:
        sent_count = message.send()
        if sent_count < 1:
            raise RuntimeError("Email backend returned zero deliveries.")
    except Exception as exc:
        email_log.status = EmailDeliveryLog.STATUS_FAILED
        email_log.error_message = str(exc)
        email_log.save(update_fields=["status", "error_message"])
        log_event(
            action="payroll.email.failed",
            user=request.user,
            target_type="payroll_record",
            target_id=str(payslip_record.id),
            metadata={"reason": str(exc)},
            ip_address=get_client_ip(request),
        )
        messages.error(request, f"Payslip email failed: {exc}")
        return redirect("payroll-detail", pk=pk)

    email_log.status = EmailDeliveryLog.STATUS_SENT
    email_log.sent_at = timezone.now()
    email_log.save(update_fields=["status", "sent_at"])
    log_event(
        action="payroll.email.sent",
        user=request.user,
        target_type="payroll_record",
        target_id=str(payslip_record.id),
        metadata={"recipient": recipient},
        ip_address=get_client_ip(request),
    )
    messages.success(request, f"Payslip email sent to {recipient}.")
    return redirect("payroll-detail", pk=pk)
