from datetime import datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
import shutil
import tempfile
from unittest.mock import patch

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.templatetags.static import static
from django.test import TestCase
from django.test import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from PIL import Image as PilImage
from reportlab.lib.units import mm

from accounts.roles import ADMIN, CUSTOMER, FINANCE, HR, STAFF, SUPERADMIN
from core.models import AuditLog
from imports.models import ImportJob
from notifications.models import EmailDeliveryLog
from payments.models import PaymentBankDetails, PaymentRecord
from support.models import SupportTicket

from .exports import (
    COMPUTER_GENERATED_INVOICE_STATEMENT,
    _build_invoice_logo,
    _build_logo_row,
    _payment_note_lines_for_pdf,
    _registered_office_line,
    _resolve_invoice_payment_summary,
    _resolve_invoice_branding,
    _invoice_text_lines,
    build_export_context,
    generate_invoice_pdf,
)
from .models import Customer, Invoice, InvoiceItem, InvoiceSourceRow, InvoiceTemplateSettings
from .services import (
    apply_overdue_status,
    parse_invoice_csv,
    parse_invoice_excel,
    recalculate_invoice_totals,
    refresh_overdue_invoices,
    transition_invoice_status,
)

User = get_user_model()


class InvoicingMvpTests(TestCase):
    def setUp(self):
        self.finance_user = User.objects.create_user(username="finance_u", password="TempPass123!")
        self.finance_user.role_profile.role = FINANCE
        self.finance_user.role_profile.save()

        self.admin_user = User.objects.create_user(username="admin_u", password="TempPass123!")
        self.admin_user.role_profile.role = ADMIN
        self.admin_user.role_profile.save()

        self.superadmin_user = User.objects.create_user(username="superadmin_u", password="TempPass123!")
        self.superadmin_user.role_profile.role = SUPERADMIN
        self.superadmin_user.role_profile.save()

        self.hr_user = User.objects.create_user(username="hr_u", password="TempPass123!")
        self.hr_user.role_profile.role = HR
        self.hr_user.role_profile.save()

        self.staff_user = User.objects.create_user(username="staff_u", password="TempPass123!")
        self.staff_user.role_profile.role = STAFF
        self.staff_user.role_profile.save()

        self.customer_user = User.objects.create_user(
            username="customer_u",
            password="TempPass123!",
            email="billing@acme.com",
        )
        self.customer_user.role_profile.role = CUSTOMER
        self.customer_user.role_profile.save()

        self.customer = Customer.objects.create(
            name="Acme Pte Ltd",
            email="billing@acme.com",
            created_by=self.finance_user,
        )

    def _create_invoice_with_item(
        self,
        *,
        invoice_number="INV-2099-2001",
        status=Invoice.STATUS_DRAFT,
        customer=None,
    ):
        customer = customer or self.customer
        invoice = Invoice.objects.create(
            invoice_number=invoice_number,
            customer=customer,
            status=status,
            issue_date=timezone.localdate(),
            due_date=timezone.localdate() + timedelta(days=10),
            currency="SGD",
            subtotal=Decimal("200.00"),
            tax_amount=Decimal("18.00"),
            total_amount=Decimal("218.00"),
            created_by=self.finance_user,
        )
        InvoiceItem.objects.create(
            invoice=invoice,
            description="Service Fee",
            quantity=Decimal("2.00"),
            unit_price=Decimal("100.00"),
            tax_rate=Decimal("9.00"),
            line_total=Decimal("218.00"),
        )
        return invoice

    def _create_invoice_with_items(
        self,
        items,
        *,
        invoice_number="INV-2099-2101",
        status=Invoice.STATUS_DRAFT,
        customer=None,
    ):
        customer = customer or self.customer
        invoice = Invoice.objects.create(
            invoice_number=invoice_number,
            customer=customer,
            status=status,
            issue_date=timezone.localdate(),
            due_date=timezone.localdate() + timedelta(days=10),
            currency="SGD",
            created_by=self.finance_user,
        )
        for item in items:
            InvoiceItem.objects.create(
                invoice=invoice,
                description=item["description"],
                quantity=Decimal(item["quantity"]),
                unit_price=Decimal(item["unit_price"]),
                tax_rate=Decimal(item.get("tax_rate", "0.00")),
            )
        recalculate_invoice_totals(invoice)
        return invoice

    def _create_basic_invoice(
        self,
        *,
        invoice_number,
        status,
        due_date,
        issue_date=None,
        customer=None,
    ):
        customer = customer or self.customer
        return Invoice.objects.create(
            invoice_number=invoice_number,
            customer=customer,
            status=status,
            issue_date=issue_date or timezone.localdate(),
            due_date=due_date,
            currency="SGD",
            subtotal=Decimal("100.00"),
            tax_amount=Decimal("9.00"),
            total_amount=Decimal("109.00"),
            created_by=self.finance_user,
        )

    def test_finance_can_create_invoice_and_calculates_totals(self):
        self.client.login(username="finance_u", password="TempPass123!")
        issue_date = timezone.localdate()
        due_date = issue_date + timedelta(days=15)

        response = self.client.post(
            reverse("invoice-create"),
            data={
                "customer": self.customer.pk,
                "issue_date": issue_date,
                "due_date": due_date,
                "currency": "SGD",
                "notes": "Test invoice",
                "items-TOTAL_FORMS": "1",
                "items-INITIAL_FORMS": "0",
                "items-MIN_NUM_FORMS": "1",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-description": "Service Fee",
                "items-0-quantity": "2",
                "items-0-unit_price": "100.00",
                "items-0-tax_rate": "9.00",
            },
        )

        self.assertEqual(response.status_code, 302)
        invoice = Invoice.objects.first()
        self.assertIsNotNone(invoice)
        self.assertTrue(invoice.invoice_number.startswith(f"INV-{issue_date.year}-"))
        self.assertEqual(invoice.items.count(), 1)
        self.assertEqual(invoice.subtotal, Decimal("200.00"))
        self.assertEqual(invoice.tax_amount, Decimal("18.00"))
        self.assertEqual(invoice.total_amount, Decimal("218.00"))
        self.assertTrue(
            AuditLog.objects.filter(
                action="invoice.created",
                target_type="invoice",
                target_id=str(invoice.id),
            ).exists()
        )

    def test_invoice_create_rejects_zero_unit_price(self):
        self.client.login(username="finance_u", password="TempPass123!")
        issue_date = timezone.localdate()
        due_date = issue_date + timedelta(days=15)

        response = self.client.post(
            reverse("invoice-create"),
            data={
                "customer": self.customer.pk,
                "issue_date": issue_date,
                "due_date": due_date,
                "currency": "SGD",
                "notes": "Invalid invoice",
                "items-TOTAL_FORMS": "1",
                "items-INITIAL_FORMS": "0",
                "items-MIN_NUM_FORMS": "1",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-description": "Service Fee",
                "items-0-quantity": "2",
                "items-0-unit_price": "0.00",
                "items-0-tax_rate": "9.00",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Unit price must be greater than 0.")
        self.assertFalse(Invoice.objects.exists())

    def test_invoice_create_rejects_negative_unit_price(self):
        self.client.login(username="finance_u", password="TempPass123!")
        issue_date = timezone.localdate()
        due_date = issue_date + timedelta(days=15)

        response = self.client.post(
            reverse("invoice-create"),
            data={
                "customer": self.customer.pk,
                "issue_date": issue_date,
                "due_date": due_date,
                "currency": "SGD",
                "notes": "Invalid invoice",
                "items-TOTAL_FORMS": "1",
                "items-INITIAL_FORMS": "0",
                "items-MIN_NUM_FORMS": "1",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-description": "Service Fee",
                "items-0-quantity": "2",
                "items-0-unit_price": "-1.00",
                "items-0-tax_rate": "9.00",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Unit price must be greater than 0.")
        self.assertFalse(Invoice.objects.exists())

    def test_finance_can_create_invoice_with_three_items(self):
        self.client.login(username="finance_u", password="TempPass123!")
        issue_date = timezone.localdate()
        due_date = issue_date + timedelta(days=15)

        response = self.client.post(
            reverse("invoice-create"),
            data={
                "customer": self.customer.pk,
                "issue_date": issue_date,
                "due_date": due_date,
                "currency": "SGD",
                "notes": "Multi-item invoice",
                "items-TOTAL_FORMS": "3",
                "items-INITIAL_FORMS": "0",
                "items-MIN_NUM_FORMS": "1",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-description": "Service Fee",
                "items-0-quantity": "2",
                "items-0-unit_price": "100.00",
                "items-0-tax_rate": "9.00",
                "items-1-description": "Consulting",
                "items-1-quantity": "1",
                "items-1-unit_price": "50.00",
                "items-1-tax_rate": "0.00",
                "items-2-description": "Materials",
                "items-2-quantity": "3",
                "items-2-unit_price": "10.00",
                "items-2-tax_rate": "10.00",
            },
        )

        self.assertEqual(response.status_code, 302)
        invoice = Invoice.objects.get()
        self.assertEqual(invoice.items.count(), 3)
        self.assertEqual(invoice.subtotal, Decimal("280.00"))
        self.assertEqual(invoice.tax_amount, Decimal("21.00"))
        self.assertEqual(invoice.total_amount, Decimal("301.00"))

    def test_invoice_create_rejects_no_valid_item(self):
        self.client.login(username="finance_u", password="TempPass123!")
        issue_date = timezone.localdate()
        due_date = issue_date + timedelta(days=15)

        response = self.client.post(
            reverse("invoice-create"),
            data={
                "customer": self.customer.pk,
                "issue_date": issue_date,
                "due_date": due_date,
                "currency": "SGD",
                "notes": "",
                "items-TOTAL_FORMS": "1",
                "items-INITIAL_FORMS": "0",
                "items-MIN_NUM_FORMS": "1",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-description": "",
                "items-0-quantity": "",
                "items-0-unit_price": "",
                "items-0-tax_rate": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "This field is required.")
        self.assertFalse(Invoice.objects.exists())

    def test_empty_extra_invoice_item_row_does_not_create_item(self):
        self.client.login(username="finance_u", password="TempPass123!")
        issue_date = timezone.localdate()
        due_date = issue_date + timedelta(days=15)

        response = self.client.post(
            reverse("invoice-create"),
            data={
                "customer": self.customer.pk,
                "issue_date": issue_date,
                "due_date": due_date,
                "currency": "SGD",
                "notes": "Blank extra row",
                "items-TOTAL_FORMS": "2",
                "items-INITIAL_FORMS": "0",
                "items-MIN_NUM_FORMS": "1",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-description": "Service Fee",
                "items-0-quantity": "2",
                "items-0-unit_price": "100.00",
                "items-0-tax_rate": "9.00",
                "items-1-description": "",
                "items-1-quantity": "",
                "items-1-unit_price": "",
                "items-1-tax_rate": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        invoice = Invoice.objects.get()
        self.assertEqual(invoice.items.count(), 1)
        self.assertEqual(invoice.total_amount, Decimal("218.00"))

    def test_allowed_roles_can_access_manual_invoice_create_form(self):
        for username in ["finance_u", "admin_u", "superadmin_u"]:
            with self.subTest(username=username):
                self.client.login(username=username, password="TempPass123!")
                response = self.client.get(reverse("invoice-create"))
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, "Add Item")
                self.client.logout()

    def test_disallowed_roles_cannot_access_manual_invoice_create_form(self):
        for username in ["customer_u", "hr_u", "staff_u"]:
            with self.subTest(username=username):
                self.client.login(username=username, password="TempPass123!")
                response = self.client.get(reverse("invoice-create"))
                self.assertEqual(response.status_code, 403)
                self.client.logout()

    def test_finance_can_open_customer_create_page_from_invoice_flow(self):
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-customer-create"), data={"next": reverse("invoice-create")})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Add New Customer")

    def test_finance_can_create_customer_and_return_to_invoice_create_with_selection(self):
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.post(
            reverse("invoice-customer-create"),
            data={
                "name": "New Merchant Partner",
                "email": "new-merchant@example.com",
                "phone": "",
                "billing_address": "",
                "tax_number": "",
                "status": "active",
                "next": reverse("invoice-create"),
            },
        )

        customer = Customer.objects.get(email="new-merchant@example.com")
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, f"{reverse('invoice-create')}?customer={customer.id}")
        self.assertTrue(
            AuditLog.objects.filter(
                action="invoice.customer.created",
                target_type="customer",
                target_id=str(customer.id),
            ).exists()
        )

    def test_staff_cannot_access_customer_create_page(self):
        self.client.login(username="staff_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-customer-create"))

        self.assertEqual(response.status_code, 403)

    def test_staff_cannot_access_invoice_pages(self):
        self.client.login(username="staff_u", password="TempPass123!")
        response = self.client.get(reverse("invoice-list"))
        self.assertEqual(response.status_code, 403)

    def test_customer_cannot_access_invoice_pages(self):
        self.client.login(username="customer_u", password="TempPass123!")
        response = self.client.get(reverse("invoice-list"))
        self.assertEqual(response.status_code, 403)

    def test_public_view_marks_invoice_as_viewed_and_tracks_count(self):
        invoice = Invoice.objects.create(
            invoice_number="INV-2099-0001",
            customer=self.customer,
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate(),
            due_date=timezone.localdate() + timedelta(days=10),
            currency="SGD",
            created_by=self.finance_user,
        )
        public_response = self.client.get(reverse("invoice-public-view", args=[invoice.public_view_token]))
        self.assertEqual(public_response.status_code, 200)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_VIEWED)
        self.assertEqual(invoice.view_count, 1)
        self.assertIsNotNone(invoice.viewed_at)

    def test_overdue_logic_applies_when_listing(self):
        invoice = Invoice.objects.create(
            invoice_number="INV-2099-0002",
            customer=self.customer,
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=20),
            due_date=timezone.localdate() - timedelta(days=1),
            currency="SGD",
            created_by=self.finance_user,
        )
        self.client.login(username="finance_u", password="TempPass123!")
        list_response = self.client.get(reverse("invoice-list"))
        self.assertEqual(list_response.status_code, 200)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_OVERDUE)

    def test_past_due_draft_invoice_remains_draft(self):
        invoice = self._create_basic_invoice(
            invoice_number="INV-2099-0003",
            status=Invoice.STATUS_DRAFT,
            due_date=timezone.localdate() - timedelta(days=1),
            issue_date=timezone.localdate() - timedelta(days=10),
        )

        changed = apply_overdue_status(invoice)

        self.assertFalse(changed)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_DRAFT)

    def test_draft_invoice_due_today_remains_draft(self):
        invoice = self._create_basic_invoice(
            invoice_number="INV-2099-0004",
            status=Invoice.STATUS_DRAFT,
            due_date=timezone.localdate(),
            issue_date=timezone.localdate() - timedelta(days=7),
        )

        changed = refresh_overdue_invoices()

        self.assertEqual(changed, 0)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_DRAFT)

    def test_future_due_draft_invoice_remains_draft(self):
        invoice = self._create_basic_invoice(
            invoice_number="INV-2099-0005",
            status=Invoice.STATUS_DRAFT,
            due_date=timezone.localdate() + timedelta(days=5),
        )

        changed = apply_overdue_status(invoice)

        self.assertFalse(changed)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_DRAFT)

    def test_past_due_viewed_invoice_becomes_overdue(self):
        invoice = self._create_basic_invoice(
            invoice_number="INV-2099-0006",
            status=Invoice.STATUS_VIEWED,
            due_date=timezone.localdate() - timedelta(days=2),
            issue_date=timezone.localdate() - timedelta(days=14),
        )

        changed = apply_overdue_status(invoice)

        self.assertTrue(changed)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_OVERDUE)

    def test_sent_invoice_due_today_does_not_become_overdue(self):
        invoice = self._create_basic_invoice(
            invoice_number="INV-2099-0007",
            status=Invoice.STATUS_SENT,
            due_date=timezone.localdate(),
            issue_date=timezone.localdate() - timedelta(days=7),
        )

        changed = refresh_overdue_invoices()

        self.assertEqual(changed, 0)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_SENT)

    def test_future_due_sent_invoice_remains_sent(self):
        invoice = self._create_basic_invoice(
            invoice_number="INV-2099-0008",
            status=Invoice.STATUS_SENT,
            due_date=timezone.localdate() + timedelta(days=4),
        )

        changed = apply_overdue_status(invoice)

        self.assertFalse(changed)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_SENT)

    def test_future_due_viewed_invoice_remains_viewed(self):
        invoice = self._create_basic_invoice(
            invoice_number="INV-2099-0009",
            status=Invoice.STATUS_VIEWED,
            due_date=timezone.localdate() + timedelta(days=4),
        )

        changed = apply_overdue_status(invoice)

        self.assertFalse(changed)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_VIEWED)

    def test_past_due_paid_invoice_remains_paid(self):
        invoice = self._create_basic_invoice(
            invoice_number="INV-2099-0010",
            status=Invoice.STATUS_PAID,
            due_date=timezone.localdate() - timedelta(days=3),
            issue_date=timezone.localdate() - timedelta(days=10),
        )

        changed = apply_overdue_status(invoice)

        self.assertFalse(changed)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_PAID)

    def test_status_transition_rejects_manual_paid_change(self):
        invoice = self._create_basic_invoice(
            invoice_number="INV-2099-0010A",
            status=Invoice.STATUS_SENT,
            due_date=timezone.localdate() + timedelta(days=3),
        )

        success, message = transition_invoice_status(invoice, Invoice.STATUS_PAID)

        self.assertFalse(success)
        self.assertIn("Stripe confirmation or verified bank-transfer confirmation", message)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_SENT)

    def test_past_due_refunded_invoice_remains_refunded(self):
        invoice = self._create_basic_invoice(
            invoice_number="INV-2099-0011",
            status=Invoice.STATUS_REFUNDED,
            due_date=timezone.localdate() - timedelta(days=3),
            issue_date=timezone.localdate() - timedelta(days=10),
        )

        changed = apply_overdue_status(invoice)

        self.assertFalse(changed)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_REFUNDED)

    def test_existing_overdue_invoice_remains_overdue(self):
        invoice = self._create_basic_invoice(
            invoice_number="INV-2099-0012",
            status=Invoice.STATUS_OVERDUE,
            due_date=timezone.localdate() - timedelta(days=5),
            issue_date=timezone.localdate() - timedelta(days=12),
        )

        changed = apply_overdue_status(invoice)

        self.assertFalse(changed)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_OVERDUE)

    def test_finance_can_filter_invoice_list_by_status(self):
        draft_invoice = Invoice.objects.create(
            invoice_number="INV-2099-3101",
            customer=self.customer,
            status=Invoice.STATUS_DRAFT,
            issue_date=timezone.localdate(),
            due_date=timezone.localdate() + timedelta(days=5),
            currency="SGD",
            created_by=self.finance_user,
        )
        paid_invoice = Invoice.objects.create(
            invoice_number="INV-2099-3102",
            customer=self.customer,
            status=Invoice.STATUS_PAID,
            issue_date=timezone.localdate(),
            due_date=timezone.localdate() + timedelta(days=5),
            currency="SGD",
            created_by=self.finance_user,
        )
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-list"), data={"status": "draft"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, draft_invoice.invoice_number)
        self.assertNotContains(response, paid_invoice.invoice_number)

    def test_finance_can_search_invoice_list_by_number_name_and_email(self):
        acme_invoice = Invoice.objects.create(
            invoice_number="INV-ACME-5001",
            customer=self.customer,
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate(),
            due_date=timezone.localdate() + timedelta(days=5),
            currency="SGD",
            created_by=self.finance_user,
        )
        other_customer = Customer.objects.create(
            name="Beta Industries",
            email="billing@beta.com",
            created_by=self.finance_user,
        )
        beta_invoice = Invoice.objects.create(
            invoice_number="INV-BETA-5002",
            customer=other_customer,
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate(),
            due_date=timezone.localdate() + timedelta(days=5),
            currency="SGD",
            created_by=self.finance_user,
        )
        self.client.login(username="finance_u", password="TempPass123!")

        by_number = self.client.get(reverse("invoice-list"), data={"q": "ACME-5001"})
        by_name = self.client.get(reverse("invoice-list"), data={"q": "Beta Industries"})
        by_email = self.client.get(reverse("invoice-list"), data={"q": "billing@beta.com"})

        self.assertContains(by_number, acme_invoice.invoice_number)
        self.assertNotContains(by_number, beta_invoice.invoice_number)
        self.assertContains(by_name, beta_invoice.invoice_number)
        self.assertNotContains(by_name, acme_invoice.invoice_number)
        self.assertContains(by_email, beta_invoice.invoice_number)
        self.assertNotContains(by_email, acme_invoice.invoice_number)

    def test_invoice_list_filters_by_issue_date_range_when_from_is_earlier_than_to(self):
        earlier_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1001",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=10),
            due_date=timezone.localdate() + timedelta(days=5),
        )
        later_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1002",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=2),
            due_date=timezone.localdate() + timedelta(days=7),
        )
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(
            reverse("invoice-list"),
            data={
                "issue_date_from": (timezone.localdate() - timedelta(days=11)).strftime("%Y-%m-%d"),
                "issue_date_to": (timezone.localdate() - timedelta(days=8)).strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, earlier_invoice.invoice_number)
        self.assertNotContains(response, later_invoice.invoice_number)

    def test_invoice_list_allows_same_day_issue_date_range(self):
        same_day = timezone.localdate() - timedelta(days=4)
        matching_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1003",
            status=Invoice.STATUS_SENT,
            issue_date=same_day,
            due_date=timezone.localdate() + timedelta(days=5),
        )
        other_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1004",
            status=Invoice.STATUS_SENT,
            issue_date=same_day - timedelta(days=1),
            due_date=timezone.localdate() + timedelta(days=6),
        )
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(
            reverse("invoice-list"),
            data={
                "issue_date_from": same_day.strftime("%Y-%m-%d"),
                "issue_date_to": same_day.strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, matching_invoice.invoice_number)
        self.assertNotContains(response, other_invoice.invoice_number)

    def test_invoice_list_allows_issue_date_from_only(self):
        older_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1005",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=9),
            due_date=timezone.localdate() + timedelta(days=4),
        )
        newer_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1006",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=1),
            due_date=timezone.localdate() + timedelta(days=8),
        )
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(
            reverse("invoice-list"),
            data={"issue_date_from": (timezone.localdate() - timedelta(days=3)).strftime("%Y-%m-%d")},
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, older_invoice.invoice_number)
        self.assertContains(response, newer_invoice.invoice_number)

    def test_invoice_list_allows_issue_date_to_only(self):
        older_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1007",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=9),
            due_date=timezone.localdate() + timedelta(days=4),
        )
        newer_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1008",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=1),
            due_date=timezone.localdate() + timedelta(days=8),
        )
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(
            reverse("invoice-list"),
            data={"issue_date_to": (timezone.localdate() - timedelta(days=3)).strftime("%Y-%m-%d")},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, older_invoice.invoice_number)
        self.assertNotContains(response, newer_invoice.invoice_number)

    def test_invoice_list_rejects_issue_date_from_later_than_to(self):
        first_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1009",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=8),
            due_date=timezone.localdate() + timedelta(days=4),
        )
        second_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1010",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=2),
            due_date=timezone.localdate() + timedelta(days=7),
        )
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(
            reverse("invoice-list"),
            data={
                "issue_date_from": "2026-06-18",
                "issue_date_to": "2026-06-14",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "From date cannot be later than To date.")
        self.assertContains(response, first_invoice.invoice_number)
        self.assertContains(response, second_invoice.invoice_number)
        self.assertContains(response, 'name="issue_date_from"', html=False)
        self.assertContains(response, 'value="2026-06-18"', html=False)
        self.assertContains(response, 'value="2026-06-14"', html=False)

    def test_invoice_list_rejects_invalid_issue_date_from_without_crashing(self):
        matching_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1011",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=5),
            due_date=timezone.localdate() + timedelta(days=5),
        )
        later_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1012",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=1),
            due_date=timezone.localdate() + timedelta(days=7),
        )
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(
            reverse("invoice-list"),
            data={
                "issue_date_from": "2026-99-99",
                "issue_date_to": (timezone.localdate() - timedelta(days=3)).strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'From date &quot;2026-99-99&quot; is invalid. Use YYYY-MM-DD.', html=False)
        self.assertContains(response, 'value="2026-99-99"', html=False)
        self.assertContains(response, matching_invoice.invoice_number)
        self.assertNotContains(response, later_invoice.invoice_number)

    def test_invoice_list_rejects_invalid_issue_date_to_without_crashing(self):
        older_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1013",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=6),
            due_date=timezone.localdate() + timedelta(days=4),
        )
        newer_invoice = self._create_basic_invoice(
            invoice_number="INV-RANGE-1014",
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate() - timedelta(days=1),
            due_date=timezone.localdate() + timedelta(days=7),
        )
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(
            reverse("invoice-list"),
            data={
                "issue_date_from": (timezone.localdate() - timedelta(days=3)).strftime("%Y-%m-%d"),
                "issue_date_to": "2026-99-88",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'To date &quot;2026-99-88&quot; is invalid. Use YYYY-MM-DD.', html=False)
        self.assertContains(response, 'value="2026-99-88"', html=False)
        self.assertNotContains(response, older_invoice.invoice_number)
        self.assertContains(response, newer_invoice.invoice_number)

    def test_invoice_list_combines_search_status_and_valid_issue_date_range(self):
        target_customer = Customer.objects.create(
            name="Gamma Holdings",
            email="billing@gamma.com",
            created_by=self.finance_user,
        )
        matching_invoice = self._create_basic_invoice(
            invoice_number="INV-GAMMA-1015",
            status=Invoice.STATUS_SENT,
            customer=target_customer,
            issue_date=timezone.localdate() - timedelta(days=3),
            due_date=timezone.localdate() + timedelta(days=5),
        )
        other_status_invoice = self._create_basic_invoice(
            invoice_number="INV-GAMMA-1016",
            status=Invoice.STATUS_PAID,
            customer=target_customer,
            issue_date=timezone.localdate() - timedelta(days=3),
            due_date=timezone.localdate() + timedelta(days=5),
        )
        other_date_invoice = self._create_basic_invoice(
            invoice_number="INV-GAMMA-1017",
            status=Invoice.STATUS_SENT,
            customer=target_customer,
            issue_date=timezone.localdate() - timedelta(days=10),
            due_date=timezone.localdate() + timedelta(days=5),
        )
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(
            reverse("invoice-list"),
            data={
                "q": "Gamma",
                "status": "sent",
                "issue_date_from": (timezone.localdate() - timedelta(days=4)).strftime("%Y-%m-%d"),
                "issue_date_to": (timezone.localdate() - timedelta(days=2)).strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, matching_invoice.invoice_number)
        self.assertNotContains(response, other_status_invoice.invoice_number)
        self.assertNotContains(response, other_date_invoice.invoice_number)

    def test_invoice_list_reset_link_returns_base_url_without_filters(self):
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(
            reverse("invoice-list"),
            data={
                "q": "Acme",
                "status": "draft",
                "issue_date_from": "2026-06-10",
                "issue_date_to": "2026-06-12",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f'<a href="{reverse("invoice-list")}" class="btn btn-outline-secondary invoice-action-btn" id="invoiceListResetButton">Reset</a>',
            html=True,
        )

    def test_invoice_create_rejects_when_all_items_are_deleted(self):
        self.client.login(username="finance_u", password="TempPass123!")
        issue_date = timezone.localdate()
        due_date = issue_date + timedelta(days=15)

        response = self.client.post(
            reverse("invoice-create"),
            data={
                "customer": self.customer.pk,
                "issue_date": issue_date,
                "due_date": due_date,
                "currency": "SGD",
                "notes": "",
                "items-TOTAL_FORMS": "1",
                "items-INITIAL_FORMS": "0",
                "items-MIN_NUM_FORMS": "1",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-description": "",
                "items-0-quantity": "",
                "items-0-unit_price": "",
                "items-0-tax_rate": "",
                "items-0-DELETE": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Please submit at least 1 form.")
        self.assertEqual(Invoice.objects.count(), 0)

    def test_finance_can_edit_draft_invoice_and_recalculate_totals(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.post(
            reverse("invoice-edit", args=[invoice.pk]),
            data={
                "customer": self.customer.pk,
                "issue_date": invoice.issue_date,
                "due_date": invoice.due_date,
                "currency": "SGD",
                "notes": "Updated note",
                "items-TOTAL_FORMS": "1",
                "items-INITIAL_FORMS": "1",
                "items-MIN_NUM_FORMS": "1",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-id": invoice.items.first().pk,
                "items-0-description": "Updated Service Fee",
                "items-0-quantity": "3",
                "items-0-unit_price": "100.00",
                "items-0-tax_rate": "9.00",
            },
        )

        self.assertEqual(response.status_code, 302)
        invoice.refresh_from_db()
        self.assertEqual(invoice.notes, "Updated note")
        self.assertEqual(invoice.subtotal, Decimal("300.00"))
        self.assertEqual(invoice.tax_amount, Decimal("27.00"))
        self.assertEqual(invoice.total_amount, Decimal("327.00"))
        self.assertTrue(
            AuditLog.objects.filter(
                action="invoice.edited",
                target_type="invoice",
                target_id=str(invoice.id),
            ).exists()
        )

    def test_finance_can_edit_invoice_with_multiple_items(self):
        invoice = self._create_invoice_with_items(
            [
                {"description": "Service Fee", "quantity": "1.00", "unit_price": "100.00", "tax_rate": "9.00"},
                {"description": "Setup", "quantity": "2.00", "unit_price": "50.00", "tax_rate": "0.00"},
                {"description": "Materials", "quantity": "3.00", "unit_price": "10.00", "tax_rate": "10.00"},
            ]
        )
        items = list(invoice.items.order_by("id"))
        self.client.login(username="finance_u", password="TempPass123!")

        get_response = self.client.get(reverse("invoice-edit", args=[invoice.pk]))
        self.assertEqual(get_response.status_code, 200)
        self.assertContains(get_response, "Service Fee")
        self.assertContains(get_response, "Setup")
        self.assertContains(get_response, "Materials")

        response = self.client.post(
            reverse("invoice-edit", args=[invoice.pk]),
            data={
                "customer": self.customer.pk,
                "issue_date": invoice.issue_date,
                "due_date": invoice.due_date,
                "currency": "SGD",
                "notes": "Updated multi-item invoice",
                "items-TOTAL_FORMS": "3",
                "items-INITIAL_FORMS": "3",
                "items-MIN_NUM_FORMS": "1",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-id": items[0].pk,
                "items-0-description": "Service Fee",
                "items-0-quantity": "2",
                "items-0-unit_price": "100.00",
                "items-0-tax_rate": "9.00",
                "items-1-id": items[1].pk,
                "items-1-description": "Setup",
                "items-1-quantity": "2",
                "items-1-unit_price": "50.00",
                "items-1-tax_rate": "0.00",
                "items-2-id": items[2].pk,
                "items-2-description": "Materials",
                "items-2-quantity": "4",
                "items-2-unit_price": "10.00",
                "items-2-tax_rate": "10.00",
            },
        )

        self.assertEqual(response.status_code, 302)
        invoice.refresh_from_db()
        self.assertEqual(invoice.items.count(), 3)
        self.assertEqual(invoice.subtotal, Decimal("340.00"))
        self.assertEqual(invoice.tax_amount, Decimal("22.00"))
        self.assertEqual(invoice.total_amount, Decimal("362.00"))

    def test_finance_can_remove_one_invoice_item_during_edit(self):
        invoice = self._create_invoice_with_items(
            [
                {"description": "Service Fee", "quantity": "1.00", "unit_price": "100.00", "tax_rate": "9.00"},
                {"description": "Setup", "quantity": "2.00", "unit_price": "50.00", "tax_rate": "0.00"},
                {"description": "Materials", "quantity": "3.00", "unit_price": "10.00", "tax_rate": "10.00"},
            ],
            invoice_number="INV-2099-2102",
        )
        items = list(invoice.items.order_by("id"))
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.post(
            reverse("invoice-edit", args=[invoice.pk]),
            data={
                "customer": self.customer.pk,
                "issue_date": invoice.issue_date,
                "due_date": invoice.due_date,
                "currency": "SGD",
                "notes": invoice.notes,
                "items-TOTAL_FORMS": "3",
                "items-INITIAL_FORMS": "3",
                "items-MIN_NUM_FORMS": "1",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-id": items[0].pk,
                "items-0-description": "Service Fee",
                "items-0-quantity": "1",
                "items-0-unit_price": "100.00",
                "items-0-tax_rate": "9.00",
                "items-1-id": items[1].pk,
                "items-1-description": "Setup",
                "items-1-quantity": "2",
                "items-1-unit_price": "50.00",
                "items-1-tax_rate": "0.00",
                "items-1-DELETE": "on",
                "items-2-id": items[2].pk,
                "items-2-description": "Materials",
                "items-2-quantity": "3",
                "items-2-unit_price": "10.00",
                "items-2-tax_rate": "10.00",
            },
        )

        self.assertEqual(response.status_code, 302)
        invoice.refresh_from_db()
        self.assertEqual(invoice.items.count(), 2)
        self.assertFalse(invoice.items.filter(description="Setup").exists())
        self.assertEqual(invoice.subtotal, Decimal("130.00"))
        self.assertEqual(invoice.tax_amount, Decimal("12.00"))
        self.assertEqual(invoice.total_amount, Decimal("142.00"))

    def test_staff_cannot_edit_invoice(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="staff_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-edit", args=[invoice.pk]))

        self.assertEqual(response.status_code, 403)

    def test_non_draft_invoice_cannot_be_edited(self):
        invoice = self._create_invoice_with_item()
        invoice.status = Invoice.STATUS_SENT
        invoice.save(update_fields=["status", "updated_at"])
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-edit", args=[invoice.pk]), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Only Draft invoices can be edited.")

    def test_draft_invoice_detail_shows_delete_button(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-detail", args=[invoice.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("invoice-delete-draft", args=[invoice.pk]))

    def test_non_draft_invoice_detail_hides_delete_button(self):
        invoice = self._create_invoice_with_item()
        invoice.status = Invoice.STATUS_SENT
        invoice.save(update_fields=["status", "updated_at"])
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-detail", args=[invoice.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, reverse("invoice-delete-draft", args=[invoice.pk]))

    def test_finance_can_delete_draft_invoice_with_confirmation(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="finance_u", password="TempPass123!")

        confirm_response = self.client.get(reverse("invoice-delete-draft", args=[invoice.pk]))
        self.assertEqual(confirm_response.status_code, 200)
        self.assertContains(confirm_response, "Delete Draft Invoice")
        self.assertContains(confirm_response, invoice.invoice_number)

        delete_response = self.client.post(reverse("invoice-delete-draft", args=[invoice.pk]))

        self.assertRedirects(delete_response, reverse("invoice-list"))
        self.assertFalse(Invoice.objects.filter(pk=invoice.pk).exists())
        self.assertTrue(
            AuditLog.objects.filter(
                action="invoice.deleted",
                target_type="invoice",
                target_id=str(invoice.id),
            ).exists()
        )

    def test_non_draft_invoice_cannot_be_deleted(self):
        invoice = self._create_invoice_with_item()
        invoice.status = Invoice.STATUS_SENT
        invoice.save(update_fields=["status", "updated_at"])
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.post(reverse("invoice-delete-draft", args=[invoice.pk]), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Only Draft invoices can be deleted.")
        self.assertTrue(Invoice.objects.filter(pk=invoice.pk).exists())

    def test_staff_cannot_delete_draft_invoice(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="staff_u", password="TempPass123!")

        response = self.client.post(reverse("invoice-delete-draft", args=[invoice.pk]))

        self.assertEqual(response.status_code, 403)

    def test_invoice_detail_shows_resend_email_button_and_last_email_label(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-detail", args=[invoice.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Send / Resend Invoice Email")
        self.assertContains(response, "Last Invoice Email Sent:")
        self.assertContains(response, "GST %")
        self.assertContains(response, "GST</th>", html=False)

    def test_invoice_list_shows_visible_issue_date_labels_and_pending_payment_status(self):
        invoice = self._create_invoice_with_item()
        invoice.status = Invoice.STATUS_SENT
        invoice.save(update_fields=["status", "updated_at"])
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Issue Date From")
        self.assertContains(response, "Issue Date To")
        self.assertContains(response, "Pending Payment")
        self.assertContains(response, 'id="issueDateFrom"', html=False)
        self.assertContains(response, 'id="issueDateTo"', html=False)
        self.assertContains(response, 'id="invoiceDateRangeMessage"', html=False)
        self.assertContains(response, static("js/invoice_list.js"))
        self.assertContains(response, "Choose a date range to limit invoices by issue date.")

    def test_invoice_list_shows_batch_email_controls(self):
        eligible_invoice = self._create_invoice_with_item(invoice_number="INV-2099-2101", status=Invoice.STATUS_SENT)
        ineligible_invoice = self._create_invoice_with_item(invoice_number="INV-2099-2102", status=Invoice.STATUS_PAID)
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Send Selected Invoice Emails")
        self.assertContains(response, 'name="selected_invoice_ids"', html=False)
        self.assertContains(response, f'aria-label="Select {eligible_invoice.invoice_number}"', html=False)
        self.assertContains(
            response,
            f'aria-label="{ineligible_invoice.invoice_number} cannot be batch emailed"',
            html=False,
        )

    def test_finance_can_download_invoice_pdf(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-download-pdf", args=[invoice.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn(f'{invoice.invoice_number}.pdf', response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_invoice_pdf_context_uses_editable_bank_transfer_details(self):
        PaymentBankDetails.objects.update_or_create(
            pk=1,
            defaults={
                "account_name": "PDF Billing Pte Ltd",
                "bank_name": "PDF Web Bank",
                "account_number": "555-444333-2",
                "paynow_id": "PDF-UEN",
                "bic": "PDFSGSG",
                "instructions": "Use the web-managed payment reference.",
            },
        )
        invoice = self._create_invoice_with_item()

        context = build_export_context(invoice)

        self.assertEqual(context["bank_transfer_details"]["account_name"], "PDF Billing Pte Ltd")
        self.assertEqual(context["bank_transfer_details"]["bank_name"], "PDF Web Bank")
        self.assertEqual(context["bank_transfer_details"]["account_number"], "555-444333-2")
        self.assertNotIn("invoice_bank_text", context)

    def test_finance_can_download_invoice_excel_and_values_match(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.get(reverse("invoice-download-excel", args=[invoice.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(f'{invoice.invoice_number}.xlsx', response["Content-Disposition"])

        wb = load_workbook(BytesIO(response.content))
        ws = wb["Invoice"]
        self.assertEqual(ws["F1"].value, invoice.invoice_number)
        self.assertEqual(ws["D11"].value, "GST %")
        self.assertAlmostEqual(float(ws["E14"].value), float(invoice.subtotal), places=2)
        self.assertEqual(ws["D15"].value, "GST")
        self.assertAlmostEqual(float(ws["E15"].value), float(invoice.tax_amount), places=2)
        self.assertAlmostEqual(float(ws["E16"].value), float(invoice.total_amount), places=2)

    def test_staff_cannot_download_invoice_documents(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="staff_u", password="TempPass123!")

        pdf_response = self.client.get(reverse("invoice-download-pdf", args=[invoice.pk]))
        excel_response = self.client.get(reverse("invoice-download-excel", args=[invoice.pk]))

        self.assertEqual(pdf_response.status_code, 403)
        self.assertEqual(excel_response.status_code, 403)

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="billing@example.com",
    )
    def test_finance_can_batch_send_selected_invoice_emails_and_skip_paid_invoice(self):
        draft_invoice = self._create_invoice_with_item(invoice_number="INV-2099-2201", status=Invoice.STATUS_DRAFT)
        sent_invoice = self._create_invoice_with_item(invoice_number="INV-2099-2202", status=Invoice.STATUS_SENT)
        paid_invoice = self._create_invoice_with_item(invoice_number="INV-2099-2203", status=Invoice.STATUS_PAID)
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.post(
            reverse("invoice-send-email-batch"),
            data={
                "selected_invoice_ids": [str(draft_invoice.id), str(sent_invoice.id), str(paid_invoice.id)],
                "next": reverse("invoice-list"),
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        draft_invoice.refresh_from_db()
        sent_invoice.refresh_from_db()
        paid_invoice.refresh_from_db()
        self.assertEqual(draft_invoice.status, Invoice.STATUS_SENT)
        self.assertEqual(sent_invoice.status, Invoice.STATUS_SENT)
        self.assertEqual(paid_invoice.status, Invoice.STATUS_PAID)
        self.assertEqual(len(mail.outbox), 2)
        self.assertGreaterEqual(len(mail.outbox[0].attachments), 1)
        self.assertGreaterEqual(len(mail.outbox[1].attachments), 1)
        sent_logs = EmailDeliveryLog.objects.filter(
            related_object_type="invoice",
            template_key="invoice_email_v1",
            status=EmailDeliveryLog.STATUS_SENT,
        )
        self.assertEqual(sent_logs.count(), 2)
        self.assertTrue(
            AuditLog.objects.filter(action="invoice.email.sent", target_id=str(draft_invoice.id)).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(action="invoice.email.sent", target_id=str(sent_invoice.id)).exists()
        )
        self.assertContains(response, "Invoice emails sent: 2.")
        self.assertContains(response, "Skipped 1 invoice(s):")
        self.assertContains(response, paid_invoice.invoice_number)
        self.assertContains(response, "Paid invoices are excluded from batch email sending.")

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="billing@example.com",
    )
    def test_batch_send_reports_missing_customer_email_failure(self):
        ok_invoice = self._create_invoice_with_item(invoice_number="INV-2099-2301", status=Invoice.STATUS_SENT)
        missing_email_customer = Customer.objects.create(
            name="No Email Customer",
            email="temp-no-email@example.com",
            created_by=self.finance_user,
        )
        failing_invoice = self._create_invoice_with_item(
            invoice_number="INV-2099-2302",
            status=Invoice.STATUS_SENT,
            customer=missing_email_customer,
        )
        missing_email_customer.email = ""
        missing_email_customer.save(update_fields=["email", "updated_at"])
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.post(
            reverse("invoice-send-email-batch"),
            data={
                "selected_invoice_ids": [str(ok_invoice.id), str(failing_invoice.id)],
                "next": reverse("invoice-list"),
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(mail.outbox), 1)
        failed_log = EmailDeliveryLog.objects.filter(
            related_object_type="invoice",
            related_object_id=str(failing_invoice.id),
        ).latest("attempted_at")
        self.assertEqual(failed_log.status, EmailDeliveryLog.STATUS_FAILED)
        self.assertIn("Customer email is missing", failed_log.error_message)
        self.assertTrue(
            AuditLog.objects.filter(action="invoice.email.failed", target_id=str(failing_invoice.id)).exists()
        )
        self.assertContains(response, "Invoice emails sent: 1.")
        self.assertContains(response, "Failed to send 1 invoice email(s):")
        self.assertContains(response, failing_invoice.invoice_number)

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="billing@example.com",
    )
    def test_finance_can_send_invoice_email_and_log_delivery(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.post(reverse("invoice-send-email", args=[invoice.pk]))

        self.assertEqual(response.status_code, 302)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_SENT)
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn(invoice.invoice_number, mail.outbox[0].subject)
        self.assertIn("http://testserver/invoices/view/", mail.outbox[0].body)
        self.assertEqual(mail.outbox[0].to, [invoice.customer.email])
        self.assertGreaterEqual(len(mail.outbox[0].attachments), 1)
        attachment_name = mail.outbox[0].attachments[0][0]
        self.assertIn(invoice.invoice_number, attachment_name)
        log = EmailDeliveryLog.objects.latest("attempted_at")
        self.assertEqual(log.status, EmailDeliveryLog.STATUS_SENT)
        self.assertEqual(log.related_object_type, "invoice")
        self.assertEqual(log.related_object_id, str(invoice.id))
        self.assertIsNotNone(log.sent_at)
        self.assertEqual(log.metadata.get("invoice_status_after_send"), Invoice.STATUS_SENT)
        self.assertEqual(log.metadata.get("pdf_attachment_added"), True)
        self.assertTrue(
            AuditLog.objects.filter(
                action="invoice.email.sent",
                target_type="invoice",
                target_id=str(invoice.id),
            ).exists()
        )

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="billing@example.com",
    )
    def test_send_invoice_email_failure_does_not_change_invoice_status(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="finance_u", password="TempPass123!")

        with patch("notifications.services.EmailMultiAlternatives.send", side_effect=Exception("SMTP unavailable")):
            response = self.client.post(reverse("invoice-send-email", args=[invoice.pk]))

        self.assertEqual(response.status_code, 302)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_DRAFT)
        log = EmailDeliveryLog.objects.latest("attempted_at")
        self.assertEqual(log.status, EmailDeliveryLog.STATUS_FAILED)
        self.assertIn("SMTP unavailable", log.error_message)
        self.assertFalse(
            AuditLog.objects.filter(
                action="invoice.email.sent",
                target_type="invoice",
                target_id=str(invoice.id),
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                action="invoice.email.failed",
                target_type="invoice",
                target_id=str(invoice.id),
            ).exists()
        )

    def test_staff_cannot_send_invoice_email(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="staff_u", password="TempPass123!")

        response = self.client.post(reverse("invoice-send-email", args=[invoice.pk]))

        self.assertEqual(response.status_code, 403)

    def test_staff_cannot_batch_send_invoice_email(self):
        invoice = self._create_invoice_with_item()
        self.client.login(username="staff_u", password="TempPass123!")

        response = self.client.post(
            reverse("invoice-send-email-batch"),
            data={"selected_invoice_ids": [str(invoice.id)], "next": reverse("invoice-list")},
        )

        self.assertEqual(response.status_code, 403)

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="billing@example.com",
    )
    def test_send_invoice_email_fails_when_customer_email_missing(self):
        invoice = self._create_invoice_with_item()
        invoice.customer.email = ""
        invoice.customer.save(update_fields=["email", "updated_at"])
        self.client.login(username="finance_u", password="TempPass123!")

        response = self.client.post(reverse("invoice-send-email", args=[invoice.pk]))

        self.assertEqual(response.status_code, 302)
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, Invoice.STATUS_DRAFT)
        log = EmailDeliveryLog.objects.latest("attempted_at")
        self.assertEqual(log.status, EmailDeliveryLog.STATUS_FAILED)
        self.assertIn("Customer email is missing", log.error_message)
        self.assertEqual(len(mail.outbox), 0)

    def test_customer_dashboard_shows_only_own_invoices(self):
        own_invoice = self._create_invoice_with_item()
        own_invoice.status = Invoice.STATUS_SENT
        own_invoice.save(update_fields=["status", "updated_at"])
        other_customer = Customer.objects.create(
            name="Other Corp",
            email="othercorp@example.com",
            created_by=self.finance_user,
        )
        Invoice.objects.create(
            invoice_number="INV-2099-9001",
            customer=other_customer,
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate(),
            due_date=timezone.localdate() + timedelta(days=5),
            currency="SGD",
            subtotal=Decimal("50.00"),
            tax_amount=Decimal("4.50"),
            total_amount=Decimal("54.50"),
            created_by=self.finance_user,
        )

        self.client.login(username="customer_u", password="TempPass123!")
        response = self.client.get(reverse("customer-invoice-dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, own_invoice.invoice_number)
        self.assertNotContains(response, "INV-2099-9001")

    def test_customer_dashboard_excludes_past_due_draft_invoice_and_keeps_it_draft(self):
        hidden_draft = self._create_basic_invoice(
            invoice_number="INV-2099-9010",
            status=Invoice.STATUS_DRAFT,
            due_date=timezone.localdate() - timedelta(days=1),
            issue_date=timezone.localdate() - timedelta(days=8),
        )
        visible_sent = self._create_basic_invoice(
            invoice_number="INV-2099-9011",
            status=Invoice.STATUS_SENT,
            due_date=timezone.localdate() + timedelta(days=3),
        )

        self.client.login(username="customer_u", password="TempPass123!")
        response = self.client.get(reverse("customer-invoice-dashboard"))

        self.assertEqual(response.status_code, 200)
        hidden_draft.refresh_from_db()
        self.assertEqual(hidden_draft.status, Invoice.STATUS_DRAFT)
        self.assertNotIn(hidden_draft, list(response.context["action_required_invoices"]))
        self.assertNotIn(hidden_draft, list(response.context["paid_invoices"]))
        self.assertNotContains(response, f'href="/invoices/my/{hidden_draft.pk}/"', html=False)
        self.assertContains(response, visible_sent.invoice_number)

    def test_customer_can_view_own_invoice_detail_and_download_pdf(self):
        own_invoice = self._create_invoice_with_item()
        own_invoice.status = Invoice.STATUS_SENT
        own_invoice.save(update_fields=["status", "updated_at"])
        self.client.login(username="customer_u", password="TempPass123!")

        detail_response = self.client.get(reverse("customer-invoice-detail", args=[own_invoice.pk]))
        pdf_response = self.client.get(reverse("customer-invoice-download-pdf", args=[own_invoice.pk]))

        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, own_invoice.invoice_number)
        self.assertContains(detail_response, "GST %")
        self.assertContains(detail_response, "GST</th>", html=False)
        self.assertContains(detail_response, "Ask About This Invoice")
        self.assertContains(detail_response, reverse("customer-invoice-support-ticket-create", args=[own_invoice.pk]))
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response["Content-Type"], "application/pdf")

    def test_public_invoice_view_uses_gst_wording(self):
        invoice = self._create_invoice_with_item()

        response = self.client.get(reverse("invoice-public-view", args=[invoice.public_view_token]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "GST %")
        self.assertContains(response, "GST</th>", html=False)

    def test_customer_invoice_detail_shows_reminder_history_empty_state(self):
        own_invoice = self._create_invoice_with_item()
        own_invoice.status = Invoice.STATUS_SENT
        own_invoice.save(update_fields=["status", "updated_at"])
        self.client.login(username="customer_u", password="TempPass123!")

        response = self.client.get(reverse("customer-invoice-detail", args=[own_invoice.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Reminder History")
        self.assertContains(response, "No reminders sent yet.")

    def test_customer_cannot_view_other_customer_invoice(self):
        other_customer = Customer.objects.create(
            name="Other Corp",
            email="othercorp@example.com",
            created_by=self.finance_user,
        )
        other_invoice = Invoice.objects.create(
            invoice_number="INV-2099-9002",
            customer=other_customer,
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate(),
            due_date=timezone.localdate() + timedelta(days=5),
            currency="SGD",
            subtotal=Decimal("50.00"),
            tax_amount=Decimal("4.50"),
            total_amount=Decimal("54.50"),
            created_by=self.finance_user,
        )

        self.client.login(username="customer_u", password="TempPass123!")
        detail_response = self.client.get(reverse("customer-invoice-detail", args=[other_invoice.pk]))
        pdf_response = self.client.get(reverse("customer-invoice-download-pdf", args=[other_invoice.pk]))

        self.assertEqual(detail_response.status_code, 404)
        self.assertEqual(pdf_response.status_code, 404)

    def test_finance_can_preview_and_confirm_invoice_csv_import(self):
        csv_content = (
            "seller_id,shop_title,OrderID,paymentMethod,email,customerName,qty,serviceName,bookedDate,vanidayShare\n"
            "S1,Acme Salon,ORD-001,Credit Card,billing@acme.com,Acme Customer,2,Hair Cut,2026-05-01 10:00,120.00\n"
            "S1,Acme Salon,ORD-002,Credit Card,billing@acme.com,Acme Customer,1,Nail Service,2026-05-02 11:00,80.00\n"
        ).encode("utf-8")
        upload_file = SimpleUploadedFile("vaniday_sample.csv", csv_content, content_type="text/csv")

        self.client.login(username="finance_u", password="TempPass123!")
        preview_response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": upload_file},
        )

        self.assertEqual(preview_response.status_code, 200)
        self.assertContains(preview_response, "Preview Summary")
        self.assertContains(preview_response, "Confirm Import")
        preview = preview_response.context["preview"]
        import_token = preview["import_token"]

        confirm_response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "confirm", "import_token": import_token},
        )
        self.assertEqual(confirm_response.status_code, 302)

        imported_invoice = Invoice.objects.filter(notes__contains="vaniday_sample.csv").first()
        self.assertIsNotNone(imported_invoice)
        self.assertEqual(imported_invoice.items.count(), 2)
        self.assertEqual(InvoiceSourceRow.objects.filter(source_file_name="vaniday_sample.csv").count(), 2)
        self.assertEqual(imported_invoice.total_amount, Decimal("200.00"))

        job = ImportJob.objects.latest("id")
        self.assertEqual(job.module, ImportJob.MODULE_INVOICING)
        self.assertEqual(job.status, ImportJob.STATUS_COMPLETED)
        self.assertEqual(job.saved_rows, 2)

    def test_invoice_csv_preview_shows_validation_errors_for_missing_fields(self):
        csv_content = (
            "seller_id,shop_title,OrderID,email,qty,serviceName,bookedDate,vanidayShare\n"
            "S1,,ORD-001,,1,Hair Cut,2026-05-01 10:00,\n"
        ).encode("utf-8")
        upload_file = SimpleUploadedFile("invalid_vaniday.csv", csv_content, content_type="text/csv")

        self.client.login(username="finance_u", password="TempPass123!")
        preview_response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": upload_file},
        )

        self.assertEqual(preview_response.status_code, 200)
        self.assertContains(preview_response, "Validation Issues")
        self.assertContains(preview_response, "Missing merchant/customer name")
        self.assertContains(preview_response, "Missing customer email")
        self.assertContains(preview_response, "Missing invoice amount")

    def test_staff_cannot_access_invoice_csv_upload(self):
        self.client.login(username="staff_u", password="TempPass123!")
        response = self.client.get(reverse("invoice-csv-upload"))
        self.assertEqual(response.status_code, 403)


class InvoiceTemplateSettingsTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp(prefix="invoice-template-tests-")
        self.addCleanup(shutil.rmtree, self.media_root, ignore_errors=True)
        self.media_override = override_settings(MEDIA_ROOT=self.media_root)
        self.media_override.enable()
        self.addCleanup(self.media_override.disable)

        self.finance_user = User.objects.create_user(username="finance_template", password="TempPass123!")
        self.finance_user.role_profile.role = FINANCE
        self.finance_user.role_profile.save(update_fields=["role", "updated_at"])

        self.admin_user = User.objects.create_user(username="admin_template", password="TempPass123!")
        self.admin_user.role_profile.role = ADMIN
        self.admin_user.role_profile.save(update_fields=["role", "updated_at"])

        self.superadmin_user = User.objects.create_user(username="superadmin_template", password="TempPass123!")
        self.superadmin_user.role_profile.role = SUPERADMIN
        self.superadmin_user.role_profile.save(update_fields=["role", "updated_at"])

        self.hr_user = User.objects.create_user(username="hr_template", password="TempPass123!")
        self.hr_user.role_profile.role = HR
        self.hr_user.role_profile.save(update_fields=["role", "updated_at"])

        self.staff_user = User.objects.create_user(username="staff_template", password="TempPass123!")
        self.staff_user.role_profile.role = STAFF
        self.staff_user.role_profile.save(update_fields=["role", "updated_at"])

        self.customer_user = User.objects.create_user(username="customer_template", password="TempPass123!")
        self.customer_user.role_profile.role = CUSTOMER
        self.customer_user.role_profile.save(update_fields=["role", "updated_at"])

        self.customer = Customer.objects.create(
            name="Template Customer",
            email="template.customer@example.com",
            billing_address="88 Customer Road\nSingapore 123456",
            created_by=self.finance_user,
        )
        self.invoice = Invoice.objects.create(
            invoice_number="INV-TEMPLATE-1001",
            customer=self.customer,
            status=Invoice.STATUS_SENT,
            issue_date=timezone.localdate(),
            due_date=timezone.localdate() + timedelta(days=14),
            currency="SGD",
            subtotal=Decimal("100.00"),
            tax_amount=Decimal("9.00"),
            total_amount=Decimal("109.00"),
            created_by=self.finance_user,
        )
        InvoiceItem.objects.create(
            invoice=self.invoice,
            description="Template Service",
            quantity=Decimal("1.00"),
            unit_price=Decimal("100.00"),
            tax_rate=Decimal("9.00"),
            line_total=Decimal("109.00"),
        )

    def _build_logo_upload(self, *, filename="invoice-logo.png", image_format="PNG", size=(200, 80)):
        image = PilImage.new("RGB", size, color=(24, 104, 171))
        output = BytesIO()
        image.save(output, format=image_format)
        output.seek(0)
        content_type = "image/png" if image_format == "PNG" else "image/jpeg"
        return SimpleUploadedFile(filename, output.read(), content_type=content_type)

    def _save_template_settings(
        self,
        *,
        company_name="Custom Invoice Company Pte Ltd",
        company_address="77 Custom Avenue\nSingapore 654321",
        logo=None,
        logo_size=InvoiceTemplateSettings.LOGO_SIZE_MEDIUM,
        logo_position=InvoiceTemplateSettings.LOGO_POSITION_LEFT,
        address_position=InvoiceTemplateSettings.ADDRESS_POSITION_LEFT,
        **extra_fields,
    ):
        template_settings = InvoiceTemplateSettings.load()
        template_settings.company_display_name = company_name
        template_settings.company_address = company_address
        template_settings.logo_size = logo_size
        template_settings.logo_position = logo_position
        template_settings.address_position = address_position
        for field_name, value in extra_fields.items():
            setattr(template_settings, field_name, value)
        if logo is not None:
            template_settings.logo = logo
        template_settings.save()
        return template_settings

    def _save_complete_bank_details(self):
        return PaymentBankDetails.objects.update_or_create(
            pk=1,
            defaults={
                "account_name": "Configured Account Name",
                "bank_name": "Configured Bank",
                "account_number": "123456789",
                "paynow_id": "PAYNOW123",
                "bic": "BANKSGSG",
                "instructions": "Use invoice number as reference.",
            },
        )[0]

    def test_authorized_user_can_open_invoice_template_settings(self):
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse("invoice-template-settings"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Invoice Template Settings")

    def test_unauthorized_users_cannot_access_invoice_template_settings(self):
        for user in [self.customer_user, self.hr_user, self.staff_user]:
            with self.subTest(role=user.role_profile.role):
                self.client.force_login(user)

                response = self.client.get(reverse("invoice-template-settings"))

                self.assertEqual(response.status_code, 403)

    def test_invoice_template_settings_can_be_saved(self):
        self.client.force_login(self.finance_user)

        response = self.client.post(
            reverse("invoice-template-settings"),
            data={
                "company_display_name": "Student Demo Company Pte Ltd",
                "company_address": "1 Example Street\nSingapore 111111",
                "logo_size": InvoiceTemplateSettings.LOGO_SIZE_LARGE,
                "logo_position": InvoiceTemplateSettings.LOGO_POSITION_RIGHT,
                "address_position": InvoiceTemplateSettings.ADDRESS_POSITION_RIGHT,
            },
        )

        self.assertEqual(response.status_code, 302)
        template_settings = InvoiceTemplateSettings.load()
        self.assertEqual(template_settings.company_display_name, "Student Demo Company Pte Ltd")
        self.assertEqual(template_settings.company_address, "1 Example Street\nSingapore 111111")
        self.assertEqual(template_settings.logo_size, InvoiceTemplateSettings.LOGO_SIZE_LARGE)
        self.assertEqual(template_settings.logo_position, InvoiceTemplateSettings.LOGO_POSITION_RIGHT)
        self.assertEqual(template_settings.address_position, InvoiceTemplateSettings.ADDRESS_POSITION_RIGHT)
        self.assertEqual(template_settings.updated_by, self.finance_user)
        self.assertTrue(
            AuditLog.objects.filter(
                action="invoice.template_settings.updated",
                target_type="invoice_template_settings",
                target_id=str(template_settings.id),
            ).exists()
        )

    def test_authorized_users_can_update_new_invoice_template_fields(self):
        authorized_users = [self.finance_user, self.admin_user, self.superadmin_user]
        for user in authorized_users:
            with self.subTest(role=user.role_profile.role):
                self.client.force_login(user)

                response = self.client.post(
                    reverse("invoice-template-settings"),
                    data={
                        "company_display_name": f"{user.username} Company Pte Ltd",
                        "company_address": "1 Example Street\nSingapore 111111",
                        "company_email": f"{user.username}@example.com",
                        "company_phone": "+65 6123 4567",
                        "company_registration_number": "202600001Z",
                        "registered_office_text": "50 Office Road, Singapore 050050",
                        "default_payment_term_days": 21,
                        "invoice_payment_notes": "Configured note one\nConfigured note two",
                        "header_text": "Configured header",
                        "footer_text": "Configured footer",
                        "logo_size": InvoiceTemplateSettings.LOGO_SIZE_MEDIUM,
                        "logo_position": InvoiceTemplateSettings.LOGO_POSITION_LEFT,
                        "address_position": InvoiceTemplateSettings.ADDRESS_POSITION_LEFT,
                    },
                )

                self.assertEqual(response.status_code, 302)
                template_settings = InvoiceTemplateSettings.load()
                self.assertEqual(template_settings.company_email, f"{user.username}@example.com")
                self.assertEqual(template_settings.company_phone, "+65 6123 4567")
                self.assertEqual(template_settings.company_registration_number, "202600001Z")
                self.assertEqual(template_settings.registered_office_text, "50 Office Road, Singapore 050050")
                self.assertEqual(template_settings.default_payment_term_days, 21)
                self.assertEqual(template_settings.invoice_payment_notes, "Configured note one\nConfigured note two")
                self.assertEqual(template_settings.header_text, "Configured header")
                self.assertEqual(template_settings.footer_text, "Configured footer")
                self.assertEqual(template_settings.updated_by, user)

    def test_invalid_default_payment_term_is_rejected(self):
        self.client.force_login(self.finance_user)

        response = self.client.post(
            reverse("invoice-template-settings"),
            data={
                "company_display_name": "Student Demo Company Pte Ltd",
                "company_address": "1 Example Street\nSingapore 111111",
                "default_payment_term_days": 0,
                "logo_size": InvoiceTemplateSettings.LOGO_SIZE_MEDIUM,
                "logo_position": InvoiceTemplateSettings.LOGO_POSITION_LEFT,
                "address_position": InvoiceTemplateSettings.ADDRESS_POSITION_LEFT,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ensure this value is greater than or equal to 1.")

    def test_invalid_logo_type_is_rejected(self):
        self.client.force_login(self.finance_user)
        invalid_logo = SimpleUploadedFile("invoice-logo.txt", b"not-an-image", content_type="text/plain")

        response = self.client.post(
            reverse("invoice-template-settings"),
            data={
                "company_display_name": "Student Demo Company Pte Ltd",
                "company_address": "1 Example Street\nSingapore 111111",
                "logo": invalid_logo,
                "logo_size": InvoiceTemplateSettings.LOGO_SIZE_MEDIUM,
                "logo_position": InvoiceTemplateSettings.LOGO_POSITION_LEFT,
                "address_position": InvoiceTemplateSettings.ADDRESS_POSITION_LEFT,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Upload a PNG or JPEG image.")

    @override_settings(INVOICE_TEMPLATE_LOGO_MAX_UPLOAD_BYTES=1024)
    def test_oversized_logo_is_rejected(self):
        self.client.force_login(self.finance_user)
        oversized_logo = SimpleUploadedFile(
            "oversized.png",
            b"x" * 2048,
            content_type="image/png",
        )

        response = self.client.post(
            reverse("invoice-template-settings"),
            data={
                "company_display_name": "Student Demo Company Pte Ltd",
                "company_address": "1 Example Street\nSingapore 111111",
                "logo": oversized_logo,
                "logo_size": InvoiceTemplateSettings.LOGO_SIZE_MEDIUM,
                "logo_position": InvoiceTemplateSettings.LOGO_POSITION_LEFT,
                "address_position": InvoiceTemplateSettings.ADDRESS_POSITION_LEFT,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Logo exceeds the maximum file size")

    def test_settings_page_does_not_show_link_for_missing_existing_logo_file(self):
        template_settings = InvoiceTemplateSettings.load()
        template_settings.logo = "invoice_branding/logos/ocbc.png"
        template_settings.save(update_fields=["logo", "updated_at"])
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse("invoice-template-settings"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Open uploaded logo")

    def test_post_with_missing_existing_logo_file_saves_other_settings_and_clears_stale_logo(self):
        template_settings = InvoiceTemplateSettings.load()
        template_settings.logo = "invoice_branding/logos/ocbc.png"
        template_settings.save(update_fields=["logo", "updated_at"])
        self.client.force_login(self.finance_user)

        response = self.client.post(
            reverse("invoice-template-settings"),
            data={
                "company_display_name": "Recovered Company Pte Ltd",
                "company_address": "8 Recovery Road\nSingapore 888888",
                "logo_size": InvoiceTemplateSettings.LOGO_SIZE_MEDIUM,
                "logo_position": InvoiceTemplateSettings.LOGO_POSITION_LEFT,
                "address_position": InvoiceTemplateSettings.ADDRESS_POSITION_LEFT,
            },
        )

        self.assertEqual(response.status_code, 302)
        template_settings.refresh_from_db()
        self.assertEqual(template_settings.company_display_name, "Recovered Company Pte Ltd")
        self.assertEqual(template_settings.company_address, "8 Recovery Road\nSingapore 888888")
        self.assertEqual(template_settings.logo.name, "")
        self.assertEqual(template_settings.updated_by, self.finance_user)

    def test_replacement_logo_upload_saves_when_existing_logo_file_is_missing(self):
        template_settings = InvoiceTemplateSettings.load()
        template_settings.logo = "invoice_branding/logos/ocbc.png"
        template_settings.save(update_fields=["logo", "updated_at"])
        self.client.force_login(self.finance_user)

        response = self.client.post(
            reverse("invoice-template-settings"),
            data={
                "company_display_name": "Replacement Logo Company Pte Ltd",
                "company_address": "9 Replacement Road\nSingapore 999999",
                "logo": self._build_logo_upload(filename="replacement-logo.png"),
                "logo_size": InvoiceTemplateSettings.LOGO_SIZE_MEDIUM,
                "logo_position": InvoiceTemplateSettings.LOGO_POSITION_LEFT,
                "address_position": InvoiceTemplateSettings.ADDRESS_POSITION_LEFT,
            },
        )

        self.assertEqual(response.status_code, 302)
        template_settings.refresh_from_db()
        self.assertTrue(template_settings.logo.name.endswith(".png"))
        self.assertTrue(template_settings.logo.storage.exists(template_settings.logo.name))
        self.assertEqual(template_settings.company_display_name, "Replacement Logo Company Pte Ltd")

    def test_clear_logo_checkbox_still_clears_existing_valid_logo(self):
        template_settings = self._save_template_settings(logo=self._build_logo_upload())
        self.assertTrue(template_settings.has_logo_file())
        self.client.force_login(self.finance_user)

        response = self.client.post(
            reverse("invoice-template-settings"),
            data={
                "company_display_name": "Clear Logo Company Pte Ltd",
                "company_address": "10 Clear Road\nSingapore 101010",
                "logo-clear": "on",
                "logo_size": InvoiceTemplateSettings.LOGO_SIZE_MEDIUM,
                "logo_position": InvoiceTemplateSettings.LOGO_POSITION_LEFT,
                "address_position": InvoiceTemplateSettings.ADDRESS_POSITION_LEFT,
            },
        )

        self.assertEqual(response.status_code, 302)
        template_settings.refresh_from_db()
        self.assertEqual(template_settings.logo.name, "")

    def test_pdf_generation_still_works_with_no_template_settings(self):
        InvoiceTemplateSettings.objects.all().delete()

        pdf_bytes = generate_invoice_pdf(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        context = build_export_context(self.invoice)
        self.assertEqual(context["company_name"], settings.COMPANY_NAME)
        self.assertEqual(context["company_address"], settings.COMPANY_ADDRESS)
        self.assertEqual(context["invoice_logo_path"], "")

    def test_pdf_generation_uses_saved_company_name_and_address(self):
        template_settings = InvoiceTemplateSettings.load()
        template_settings.company_display_name = "Custom Invoice Company Pte Ltd"
        template_settings.company_address = "77 Custom Avenue\nSingapore 654321"
        template_settings.save()

        pdf_bytes = generate_invoice_pdf(self.invoice)
        context = build_export_context(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(context["company_name"], "Custom Invoice Company Pte Ltd")
        self.assertEqual(context["company_address"], "77 Custom Avenue\nSingapore 654321")

    def test_pdf_generation_uses_no_logo_fallback_when_existing_logo_file_is_missing(self):
        template_settings = InvoiceTemplateSettings.load()
        template_settings.logo = "invoice_branding/logos/ocbc.png"
        template_settings.save(update_fields=["logo", "updated_at"])

        pdf_bytes = generate_invoice_pdf(self.invoice)
        context = build_export_context(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(context["invoice_logo_path"], "")
        self.assertIsNone(_build_invoice_logo(template_settings, context["invoice_branding"]))

    def test_pdf_generation_uses_saved_company_business_information(self):
        self._save_template_settings(
            company_email="configured.finance@example.com",
            company_phone="+65 6987 6543",
            company_registration_number="202655555M",
            registered_office_text="99 Configured Office Road\nSingapore 099999",
            default_payment_term_days=45,
        )

        pdf_bytes = generate_invoice_pdf(self.invoice)
        context = build_export_context(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(context["company_email"], "configured.finance@example.com")
        self.assertEqual(context["company_phone"], "+65 6987 6543")
        self.assertEqual(context["company_reg_no"], "202655555M")
        self.assertEqual(context["registered_office_text"], "99 Configured Office Road\nSingapore 099999")
        self.assertEqual(context["invoice_payment_term_days"], 45)
        self.assertEqual(context["invoice_attention_email"], "configured.finance@example.com")

    def test_pdf_generation_uses_saved_payment_notes_header_and_footer(self):
        self._save_template_settings(
            invoice_payment_notes="Configured payment note A\nConfigured payment note B",
            header_text="Configured header A\nConfigured header B",
            footer_text="Configured footer A\nConfigured footer B",
        )

        pdf_bytes = generate_invoice_pdf(self.invoice)
        context = build_export_context(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(context["invoice_payment_notes"], "Configured payment note A\nConfigured payment note B")
        self.assertEqual(context["invoice_header_text"], "Configured header A\nConfigured header B")
        self.assertEqual(context["invoice_footer_text"], "Configured footer A\nConfigured footer B")

    @override_settings(
        COMPANY_EMAIL="fallback.finance@example.com",
        COMPANY_PHONE="+65 6000 9999",
        COMPANY_REG_NO="199900001A",
        REGISTERED_OFFICE_TEXT="Fallback registered office",
        INVOICE_PAYMENT_TERM_DAYS=30,
        INVOICE_PAYMENT_NOTES="Fallback payment note",
    )
    def test_blank_template_fields_fall_back_to_django_settings(self):
        self._save_template_settings(
            company_email="",
            company_phone="",
            company_registration_number="",
            registered_office_text="",
            default_payment_term_days=None,
            invoice_payment_notes="",
            header_text="",
            footer_text="",
        )

        context = build_export_context(self.invoice)

        self.assertEqual(context["company_email"], "fallback.finance@example.com")
        self.assertEqual(context["company_phone"], "+65 6000 9999")
        self.assertEqual(context["company_reg_no"], "199900001A")
        self.assertEqual(context["registered_office_text"], "Fallback registered office")
        self.assertEqual(context["invoice_payment_term_days"], 30)
        self.assertEqual(context["invoice_payment_notes"], "Fallback payment note")
        self.assertEqual(context["invoice_header_text"], "")
        self.assertEqual(context["invoice_footer_text"], "")

    def test_bank_transfer_details_still_come_from_payment_bank_details(self):
        self._save_template_settings(invoice_payment_notes="Configured invoice note")
        self._save_complete_bank_details()

        context = build_export_context(self.invoice)

        self.assertEqual(context["bank_transfer_details"]["account_name"], "Configured Account Name")
        self.assertEqual(context["bank_transfer_details"]["bank_name"], "Configured Bank")
        self.assertEqual(context["bank_transfer_details"]["account_number"], "123456789")
        self.assertEqual(context["bank_transfer_details"]["paynow_id"], "PAYNOW123")
        self.assertEqual(context["bank_transfer_details"]["bic"], "BANKSGSG")
        self.assertEqual(context["bank_transfer_details"]["instructions"], "Use invoice number as reference.")

    def test_unpaid_issued_invoice_with_balance_shows_bank_details(self):
        self._save_complete_bank_details()
        self.invoice.status = Invoice.STATUS_SENT
        self.invoice.save(update_fields=["status", "updated_at"])

        context = build_export_context(self.invoice)
        payment_summary = _resolve_invoice_payment_summary(self.invoice)
        pdf_bytes = generate_invoice_pdf(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertTrue(payment_summary["show_payment_instructions"])
        self.assertIsNotNone(context["bank_transfer_details"])
        self.assertEqual(payment_summary["amount_due"], self.invoice.total_amount)

    def test_overdue_invoice_with_balance_shows_bank_details(self):
        self._save_complete_bank_details()
        self.invoice.status = Invoice.STATUS_OVERDUE
        self.invoice.save(update_fields=["status", "updated_at"])

        payment_summary = _resolve_invoice_payment_summary(self.invoice)
        pdf_bytes = generate_invoice_pdf(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertTrue(payment_summary["show_payment_instructions"])
        self.assertEqual(payment_summary["amount_due"], self.invoice.total_amount)

    def test_paid_invoice_with_zero_balance_hides_bank_details_and_shows_paid_in_full(self):
        self._save_complete_bank_details()
        self._save_template_settings(invoice_payment_notes="Please pay by bank transfer.")
        self.invoice.status = Invoice.STATUS_PAID
        self.invoice.save(update_fields=["status", "updated_at"])

        context = build_export_context(self.invoice)
        payment_summary = _resolve_invoice_payment_summary(self.invoice)
        pdf_bytes = generate_invoice_pdf(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertFalse(payment_summary["show_payment_instructions"])
        self.assertEqual(payment_summary["amount_due"], Decimal("0.00"))
        self.assertEqual(payment_summary["payment_status_message"], "Paid in Full")
        self.assertEqual(_payment_note_lines_for_pdf(context, payment_summary), [])

    def test_refunded_invoice_hides_bank_details(self):
        self._save_complete_bank_details()
        self._save_template_settings(invoice_payment_notes="Please pay by bank transfer.")
        self.invoice.status = Invoice.STATUS_REFUNDED
        self.invoice.save(update_fields=["status", "updated_at"])

        context = build_export_context(self.invoice)
        payment_summary = _resolve_invoice_payment_summary(self.invoice)
        pdf_bytes = generate_invoice_pdf(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertFalse(payment_summary["show_payment_instructions"])
        self.assertEqual(payment_summary["payment_status_message"], "Refunded")
        self.assertEqual(_payment_note_lines_for_pdf(context, payment_summary), [])

    def test_computer_generated_statement_is_rendered_once(self):
        self._save_template_settings(
            invoice_payment_notes=(
                f"{COMPUTER_GENERATED_INVOICE_STATEMENT}\nPlease include your invoice number as reference."
            ),
            footer_text=COMPUTER_GENERATED_INVOICE_STATEMENT,
        )
        self.invoice.status = Invoice.STATUS_SENT
        self.invoice.save(update_fields=["status", "updated_at"])
        context = build_export_context(self.invoice)
        payment_summary = _resolve_invoice_payment_summary(self.invoice)

        rendered_lines = (
            _payment_note_lines_for_pdf(context, payment_summary)
            + _invoice_text_lines(context["invoice_footer_text"])
            + [COMPUTER_GENERATED_INVOICE_STATEMENT]
        )

        self.assertEqual(rendered_lines.count(COMPUTER_GENERATED_INVOICE_STATEMENT), 1)
        self.assertIn("Please include your invoice number as reference.", rendered_lines)

    def test_registered_office_prefix_is_rendered_once(self):
        self._save_template_settings(
            registered_office_text="Registered Office: Registered Office: 7 Office Road, Singapore 070707"
        )

        office_line = _registered_office_line(build_export_context(self.invoice))

        self.assertEqual(office_line.count("Registered Office:"), 1)
        self.assertEqual(office_line, "Registered Office: 7 Office Road, Singapore 070707")

    def test_configured_payment_notes_render_when_payment_is_due(self):
        self._save_template_settings(invoice_payment_notes="Configured payment note")
        self.invoice.status = Invoice.STATUS_SENT
        self.invoice.save(update_fields=["status", "updated_at"])
        context = build_export_context(self.invoice)
        payment_summary = _resolve_invoice_payment_summary(self.invoice)

        self.assertEqual(_payment_note_lines_for_pdf(context, payment_summary), ["Configured payment note"])

    def test_payment_notes_conflicting_with_configured_term_are_skipped(self):
        self._save_template_settings(
            default_payment_term_days=30,
            invoice_payment_notes=(
                "We will payout within 10 days from Invoice Date.\n"
                "Please include your invoice number as reference."
            ),
        )
        self.invoice.status = Invoice.STATUS_SENT
        self.invoice.save(update_fields=["status", "updated_at"])
        context = build_export_context(self.invoice)
        payment_summary = _resolve_invoice_payment_summary(self.invoice)

        self.assertEqual(
            _payment_note_lines_for_pdf(context, payment_summary),
            ["Please include your invoice number as reference."],
        )

    def test_pdf_context_labels_customer_information_as_bill_to(self):
        context = build_export_context(self.invoice)

        self.assertEqual(context["bill_to_label"], "Bill To")
        self.assertEqual(context["customer_display"], "Template Customer")

    def test_pdf_generation_works_without_logo_when_business_fields_are_configured(self):
        self._save_template_settings(
            company_email="configured.finance@example.com",
            company_phone="+65 6987 6543",
            company_registration_number="202655555M",
        )

        pdf_bytes = generate_invoice_pdf(self.invoice)
        context = build_export_context(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(context["invoice_logo_path"], "")

    def test_pdf_generation_does_not_change_invoice_totals_or_status(self):
        self._save_template_settings(default_payment_term_days=45)
        original_values = {
            "status": self.invoice.status,
            "issue_date": self.invoice.issue_date,
            "due_date": self.invoice.due_date,
            "subtotal": self.invoice.subtotal,
            "tax_amount": self.invoice.tax_amount,
            "total_amount": self.invoice.total_amount,
        }

        pdf_bytes = generate_invoice_pdf(self.invoice)
        self.invoice.refresh_from_db()

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(self.invoice.status, original_values["status"])
        self.assertEqual(self.invoice.issue_date, original_values["issue_date"])
        self.assertEqual(self.invoice.due_date, original_values["due_date"])
        self.assertEqual(self.invoice.subtotal, original_values["subtotal"])
        self.assertEqual(self.invoice.tax_amount, original_values["tax_amount"])
        self.assertEqual(self.invoice.total_amount, original_values["total_amount"])

    def test_pdf_generation_works_with_logo_settings(self):
        template_settings = self._save_template_settings(
            logo=self._build_logo_upload(),
            logo_size=InvoiceTemplateSettings.LOGO_SIZE_SMALL,
            logo_position=InvoiceTemplateSettings.LOGO_POSITION_CENTRE,
            address_position=InvoiceTemplateSettings.ADDRESS_POSITION_RIGHT,
        )

        pdf_bytes = generate_invoice_pdf(self.invoice)
        context = build_export_context(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertTrue(context["invoice_logo_path"].endswith(".png"))
        self.assertEqual(
            context["invoice_template_settings"].logo_position,
            InvoiceTemplateSettings.LOGO_POSITION_CENTRE,
        )
        self.assertEqual(
            context["invoice_template_settings"].address_position,
            InvoiceTemplateSettings.ADDRESS_POSITION_RIGHT,
        )

    def test_pdf_generation_with_logo_position_left_places_logo_left(self):
        template_settings = self._save_template_settings(
            logo=self._build_logo_upload(),
            logo_position=InvoiceTemplateSettings.LOGO_POSITION_LEFT,
        )

        pdf_bytes = generate_invoice_pdf(self.invoice)
        branding = _resolve_invoice_branding(template_settings)
        logo = _build_invoice_logo(template_settings, branding)
        logo_row = _build_logo_row(logo, branding, column_width=78 * mm)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(logo_row._invoice_logo_position, InvoiceTemplateSettings.LOGO_POSITION_LEFT)
        self.assertEqual(logo_row._invoice_logo_alignment, "LEFT")
        context = build_export_context(self.invoice)
        self.assertEqual(context["invoice_branding"]["logo_position"], InvoiceTemplateSettings.LOGO_POSITION_LEFT)
        self.assertEqual(context["invoice_branding"]["logo_size"], InvoiceTemplateSettings.LOGO_SIZE_MEDIUM)
        self.assertEqual(context["invoice_branding"]["address_position"], InvoiceTemplateSettings.ADDRESS_POSITION_LEFT)

    def test_pdf_generation_with_logo_position_centre_places_logo_centre(self):
        template_settings = self._save_template_settings(
            logo=self._build_logo_upload(),
            logo_position=InvoiceTemplateSettings.LOGO_POSITION_CENTRE,
        )

        pdf_bytes = generate_invoice_pdf(self.invoice)
        branding = _resolve_invoice_branding(template_settings)
        logo = _build_invoice_logo(template_settings, branding)
        logo_row = _build_logo_row(logo, branding, column_width=78 * mm)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(logo_row._invoice_logo_position, InvoiceTemplateSettings.LOGO_POSITION_CENTRE)
        self.assertEqual(logo_row._invoice_logo_alignment, "CENTER")
        context = build_export_context(self.invoice)
        self.assertEqual(context["invoice_branding"]["logo_position"], InvoiceTemplateSettings.LOGO_POSITION_CENTRE)
        self.assertEqual(context["invoice_branding"]["logo_size"], InvoiceTemplateSettings.LOGO_SIZE_MEDIUM)
        self.assertEqual(context["invoice_branding"]["address_position"], InvoiceTemplateSettings.ADDRESS_POSITION_LEFT)

    def test_pdf_generation_with_logo_position_right_places_logo_right(self):
        template_settings = self._save_template_settings(
            logo=self._build_logo_upload(),
            logo_position=InvoiceTemplateSettings.LOGO_POSITION_RIGHT,
        )

        pdf_bytes = generate_invoice_pdf(self.invoice)
        branding = _resolve_invoice_branding(template_settings)
        logo = _build_invoice_logo(template_settings, branding)
        logo_row = _build_logo_row(logo, branding, column_width=78 * mm)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(logo_row._invoice_logo_position, InvoiceTemplateSettings.LOGO_POSITION_RIGHT)
        self.assertEqual(logo_row._invoice_logo_alignment, "RIGHT")
        context = build_export_context(self.invoice)
        self.assertEqual(context["invoice_branding"]["logo_position"], InvoiceTemplateSettings.LOGO_POSITION_RIGHT)
        self.assertEqual(context["invoice_branding"]["logo_size"], InvoiceTemplateSettings.LOGO_SIZE_MEDIUM)
        self.assertEqual(context["invoice_branding"]["address_position"], InvoiceTemplateSettings.ADDRESS_POSITION_LEFT)

    def test_pdf_generation_with_logo_sizes_changes_rendered_logo_width(self):
        width_by_size = {}
        for logo_size in [
            InvoiceTemplateSettings.LOGO_SIZE_SMALL,
            InvoiceTemplateSettings.LOGO_SIZE_MEDIUM,
            InvoiceTemplateSettings.LOGO_SIZE_LARGE,
        ]:
            with self.subTest(logo_size=logo_size):
                template_settings = self._save_template_settings(
                    logo=self._build_logo_upload(filename=f"{logo_size}.png"),
                    logo_size=logo_size,
                )
                pdf_bytes = generate_invoice_pdf(self.invoice)
                branding = _resolve_invoice_branding(template_settings)
                logo = _build_invoice_logo(template_settings, branding)
                width_by_size[logo_size] = logo._invoice_logo_width
                self.assertTrue(pdf_bytes.startswith(b"%PDF"))
                self.assertEqual(branding["logo_size"], logo_size)

        self.assertLess(width_by_size[InvoiceTemplateSettings.LOGO_SIZE_SMALL], width_by_size[InvoiceTemplateSettings.LOGO_SIZE_MEDIUM])
        self.assertLess(width_by_size[InvoiceTemplateSettings.LOGO_SIZE_MEDIUM], width_by_size[InvoiceTemplateSettings.LOGO_SIZE_LARGE])

    def test_pdf_generation_with_address_position_left_places_address_left(self):
        template_settings = self._save_template_settings(
            company_name="Address Left Demo Co",
            logo=self._build_logo_upload(),
            address_position=InvoiceTemplateSettings.ADDRESS_POSITION_LEFT,
        )

        pdf_bytes = generate_invoice_pdf(self.invoice)
        branding = _resolve_invoice_branding(template_settings)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(branding["address_position"], InvoiceTemplateSettings.ADDRESS_POSITION_LEFT)
        self.assertEqual(branding["address_alignment"], "LEFT")
        context = build_export_context(self.invoice)
        self.assertEqual(context["invoice_branding"]["address_position"], InvoiceTemplateSettings.ADDRESS_POSITION_LEFT)
        self.assertEqual(context["invoice_branding"]["logo_position"], InvoiceTemplateSettings.LOGO_POSITION_LEFT)

    def test_pdf_generation_with_address_position_right_places_address_right(self):
        template_settings = self._save_template_settings(
            company_name="Address Right Demo Co",
            logo=self._build_logo_upload(),
            address_position=InvoiceTemplateSettings.ADDRESS_POSITION_RIGHT,
        )

        pdf_bytes = generate_invoice_pdf(self.invoice)
        branding = _resolve_invoice_branding(template_settings)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(branding["address_position"], InvoiceTemplateSettings.ADDRESS_POSITION_RIGHT)
        self.assertEqual(branding["address_alignment"], "RIGHT")
        context = build_export_context(self.invoice)
        self.assertEqual(context["invoice_branding"]["address_position"], InvoiceTemplateSettings.ADDRESS_POSITION_RIGHT)
        self.assertEqual(context["invoice_branding"]["logo_position"], InvoiceTemplateSettings.LOGO_POSITION_LEFT)

    @override_settings(INVOICE_PAYMENT_NOTES="Line one\\nLine two")
    def test_generated_pdf_renders_escaped_newlines_as_real_line_breaks(self):
        context = build_export_context(self.invoice)
        pdf_bytes = generate_invoice_pdf(self.invoice)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertEqual(context["invoice_payment_notes"], "Line one\nLine two")
        self.assertNotIn("\\n", context["invoice_payment_notes"])


class InvoiceCollectionReportingTests(TestCase):
    def setUp(self):
        self.finance_user = User.objects.create_user(username="finance_collect", password="TempPass123!")
        self.finance_user.role_profile.role = FINANCE
        self.finance_user.role_profile.save(update_fields=["role", "updated_at"])
        self.customer = Customer.objects.create(
            name="Collection Customer",
            email="collection@example.com",
            created_by=self.finance_user,
        )

    def _month_start(self, months_ago=0):
        month_start = timezone.localdate().replace(day=1)
        for _ in range(months_ago):
            month_start = (month_start - timedelta(days=1)).replace(day=1)
        return month_start

    def _aware_datetime(self, day_value, hour=10):
        return timezone.make_aware(datetime.combine(day_value, time(hour=hour)))

    def _create_invoice(self, *, invoice_number, status, issue_date, due_date, total_amount="109.00"):
        return Invoice.objects.create(
            invoice_number=invoice_number,
            customer=self.customer,
            status=status,
            issue_date=issue_date,
            due_date=due_date,
            currency="SGD",
            subtotal=Decimal(total_amount) - Decimal("9.00"),
            tax_amount=Decimal("9.00"),
            total_amount=Decimal(total_amount),
            created_by=self.finance_user,
        )

    def test_invoice_dashboard_counts_collection_by_payment_date(self):
        issue_month_start = self._month_start(2)
        payment_month_start = self._month_start(1)
        current_month_start = self._month_start(0)
        invoice = self._create_invoice(
            invoice_number="INV-COLL-2001",
            status=Invoice.STATUS_PAID,
            issue_date=issue_month_start,
            due_date=issue_month_start + timedelta(days=7),
        )
        PaymentRecord.objects.create(
            invoice=invoice,
            payment_reference="PAY-COLL-2001",
            provider=PaymentRecord.PROVIDER_STRIPE,
            status=PaymentRecord.STATUS_SUCCEEDED,
            amount=Decimal("109.00"),
            currency="SGD",
            paid_at=self._aware_datetime(payment_month_start + timedelta(days=4)),
        )
        Invoice.objects.filter(pk=invoice.pk).update(
            updated_at=self._aware_datetime(current_month_start + timedelta(days=2), hour=15)
        )

        self.client.force_login(self.finance_user)
        response = self.client.get(reverse("invoice-dashboard"), data={"category": "outstanding"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["collected_month"], Decimal("0.00"))
        self.assertEqual(response.context["collected_year"], Decimal("109.00"))

    def test_invoice_dashboard_excludes_failed_cancelled_and_unpaid_invoices_from_collection(self):
        month_start = self._month_start(0)
        unpaid_invoice = self._create_invoice(
            invoice_number="INV-COLL-2002",
            status=Invoice.STATUS_SENT,
            issue_date=month_start,
            due_date=month_start + timedelta(days=5),
            total_amount="150.00",
        )
        failed_invoice = self._create_invoice(
            invoice_number="INV-COLL-2003",
            status=Invoice.STATUS_SENT,
            issue_date=month_start,
            due_date=month_start + timedelta(days=6),
            total_amount="80.00",
        )
        cancelled_invoice = self._create_invoice(
            invoice_number="INV-COLL-2004",
            status=Invoice.STATUS_OVERDUE,
            issue_date=month_start,
            due_date=month_start + timedelta(days=3),
            total_amount="70.00",
        )
        PaymentRecord.objects.create(
            invoice=failed_invoice,
            payment_reference="PAY-COLL-2003",
            provider=PaymentRecord.PROVIDER_STRIPE,
            status=PaymentRecord.STATUS_FAILED,
            amount=Decimal("80.00"),
            currency="SGD",
        )
        PaymentRecord.objects.create(
            invoice=cancelled_invoice,
            payment_reference="PAY-COLL-2004",
            provider=PaymentRecord.PROVIDER_STRIPE,
            status=PaymentRecord.STATUS_CANCELLED,
            amount=Decimal("70.00"),
            currency="SGD",
        )

        self.client.force_login(self.finance_user)
        response = self.client.get(reverse("invoice-dashboard"), data={"category": "outstanding"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["collected_month"], Decimal("0.00"))
        self.assertEqual(response.context["outstanding_amount"], Decimal("300.00"))
        self.assertContains(response, unpaid_invoice.invoice_number)

    def test_invoice_dashboard_counts_confirmed_bank_transfer_using_paid_at(self):
        today = timezone.localdate()
        invoice = self._create_invoice(
            invoice_number="INV-COLL-2005",
            status=Invoice.STATUS_PAID,
            issue_date=self._month_start(1),
            due_date=self._month_start(1) + timedelta(days=8),
            total_amount="54.50",
        )
        PaymentRecord.objects.create(
            invoice=invoice,
            payment_reference="PAY-COLL-2005",
            provider=PaymentRecord.PROVIDER_MANUAL,
            status=PaymentRecord.STATUS_SUCCEEDED,
            amount=Decimal("54.50"),
            currency="SGD",
            paid_at=self._aware_datetime(today),
        )

        self.client.force_login(self.finance_user)
        response = self.client.get(reverse("invoice-dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["collected_month"], Decimal("54.50"))


class InvoiceDashboardUiTests(TestCase):
    def setUp(self):
        self.finance_user = User.objects.create_user(username="finance_dashboard", password="TempPass123!")
        self.finance_user.role_profile.role = FINANCE
        self.finance_user.role_profile.save(update_fields=["role", "updated_at"])

        self.admin_user = User.objects.create_user(username="admin_dashboard", password="TempPass123!")
        self.admin_user.role_profile.role = ADMIN
        self.admin_user.role_profile.save(update_fields=["role", "updated_at"])

        self.superadmin_user = User.objects.create_user(username="super_dashboard", password="TempPass123!")
        self.superadmin_user.role_profile.role = SUPERADMIN
        self.superadmin_user.role_profile.save(update_fields=["role", "updated_at"])

        self.hr_user = User.objects.create_user(username="hr_dashboard", password="TempPass123!")
        self.hr_user.role_profile.role = HR
        self.hr_user.role_profile.save(update_fields=["role", "updated_at"])

        self.staff_user = User.objects.create_user(username="staff_dashboard", password="TempPass123!")
        self.staff_user.role_profile.role = STAFF
        self.staff_user.role_profile.save(update_fields=["role", "updated_at"])

        self.customer_user = User.objects.create_user(username="customer_dashboard", password="TempPass123!")
        self.customer_user.role_profile.role = CUSTOMER
        self.customer_user.role_profile.save(update_fields=["role", "updated_at"])

        self.customer = Customer.objects.create(
            name="Dashboard Customer",
            email="dashboard@example.com",
            created_by=self.finance_user,
        )

    def _aware_datetime(self, day_value, hour=10):
        return timezone.make_aware(datetime.combine(day_value, time(hour=hour)))

    def _create_invoice(self, *, invoice_number, status, total_amount, issue_date, due_date):
        total_amount_decimal = Decimal(total_amount)
        return Invoice.objects.create(
            invoice_number=invoice_number,
            customer=self.customer,
            status=status,
            issue_date=issue_date,
            due_date=due_date,
            currency="SGD",
            subtotal=total_amount_decimal - Decimal("9.00"),
            tax_amount=Decimal("9.00"),
            total_amount=total_amount_decimal,
            created_by=self.finance_user,
        )

    def _create_submitted_bank_transfer(self, invoice, *, payment_reference="BANK-DASH-VERIFY-001", status=None):
        return PaymentRecord.objects.create(
            invoice=invoice,
            payment_reference=payment_reference,
            provider=PaymentRecord.PROVIDER_MANUAL,
            status=status or PaymentRecord.STATUS_PENDING,
            amount=invoice.total_amount,
            currency=invoice.currency,
            manual_customer_amount=invoice.total_amount,
            manual_customer_transfer_date=timezone.localdate(),
            manual_customer_bank_reference=f"{payment_reference}-CUSTOMER-REF",
            manual_customer_submitted_at=timezone.now(),
        )

    def test_superadmin_admin_and_finance_can_access_invoice_dashboard(self):
        for user in [self.superadmin_user, self.admin_user, self.finance_user]:
            self.client.force_login(user)
            response = self.client.get(reverse("invoice-dashboard"))
            self.assertEqual(response.status_code, 200)
            self.client.logout()

    def test_customer_hr_and_staff_cannot_access_invoice_dashboard(self):
        for user in [self.customer_user, self.hr_user, self.staff_user]:
            self.client.force_login(user)
            response = self.client.get(reverse("invoice-dashboard"))
            self.assertEqual(response.status_code, 403)
            self.client.logout()

    def test_finance_sees_invoice_support_ticket_action_counts(self):
        open_ticket = SupportTicket.objects.create(
            category=SupportTicket.CATEGORY_INVOICE,
            subject="Open invoice ticket",
            message="Needs Finance review.",
            status=SupportTicket.STATUS_OPEN,
            created_by=self.customer_user,
        )
        SupportTicket.objects.create(
            category=SupportTicket.CATEGORY_PAYMENT,
            subject="Payment ticket in progress",
            message="Finance is reviewing this.",
            status=SupportTicket.STATUS_IN_PROGRESS,
            created_by=self.customer_user,
        )
        SupportTicket.objects.create(
            category=SupportTicket.CATEGORY_INVOICE,
            subject="Resolved invoice ticket",
            message="Already handled.",
            status=SupportTicket.STATUS_RESOLVED,
            created_by=self.customer_user,
        )
        SupportTicket.objects.create(
            category=SupportTicket.CATEGORY_PAYROLL,
            subject="Payroll ticket",
            message="Not a Finance invoice/payment ticket.",
            status=SupportTicket.STATUS_OPEN,
            created_by=self.staff_user,
        )
        open_ticket.created_at = timezone.now() - timedelta(days=settings.SUPPORT_TICKET_SLA_DAYS + 1)
        open_ticket.save(update_fields=["created_at"])
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse("invoice-dashboard"))

        self.assertEqual(response.status_code, 200)
        summary = response.context["invoice_support_ticket_summary"]
        self.assertEqual(summary["open_count"], 1)
        self.assertEqual(summary["in_progress_count"], 1)
        self.assertEqual(summary["overdue_count"], 1)
        self.assertEqual(summary["active_count"], 2)
        self.assertContains(response, "Invoice Support Tickets")
        self.assertContains(response, "Open")
        self.assertContains(response, "In Progress")
        self.assertContains(response, "Overdue")
        self.assertContains(response, reverse("finance-support-ticket-list"))
        self.assertNotContains(response, "No invoice support tickets require action.")

    def test_admin_and_superadmin_see_invoice_support_ticket_action_counts(self):
        SupportTicket.objects.create(
            category=SupportTicket.CATEGORY_INVOICE,
            subject="Admin visible invoice ticket",
            message="Needs Finance action.",
            status=SupportTicket.STATUS_OPEN,
            created_by=self.customer_user,
        )

        for user in [self.admin_user, self.superadmin_user]:
            self.client.force_login(user)
            response = self.client.get(reverse("invoice-dashboard"))
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.context["invoice_support_ticket_summary"]["open_count"], 1)
            self.assertContains(response, "Invoice Support Tickets")
            self.assertContains(response, reverse("finance-support-ticket-list"))
            self.client.logout()

    def test_invoice_dashboard_support_ticket_empty_state(self):
        SupportTicket.objects.create(
            category=SupportTicket.CATEGORY_INVOICE,
            subject="Resolved only",
            message="This should not require action.",
            status=SupportTicket.STATUS_RESOLVED,
            created_by=self.customer_user,
        )
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse("invoice-dashboard"))

        self.assertEqual(response.status_code, 200)
        summary = response.context["invoice_support_ticket_summary"]
        self.assertEqual(summary["active_count"], 0)
        self.assertContains(response, "No invoice support tickets require action.")

    def test_invoice_dashboard_outstanding_amount_counts_only_issued_unpaid_invoices(self):
        today = timezone.localdate()
        self._create_invoice(
            invoice_number="INV-DASH-OUT-1",
            status=Invoice.STATUS_DRAFT,
            total_amount="109.00",
            issue_date=today,
            due_date=today + timedelta(days=7),
        )
        self._create_invoice(
            invoice_number="INV-DASH-OUT-2",
            status=Invoice.STATUS_SENT,
            total_amount="150.00",
            issue_date=today - timedelta(days=2),
            due_date=today + timedelta(days=5),
        )
        self._create_invoice(
            invoice_number="INV-DASH-OUT-3",
            status=Invoice.STATUS_VIEWED,
            total_amount="200.00",
            issue_date=today - timedelta(days=4),
            due_date=today + timedelta(days=3),
        )
        self._create_invoice(
            invoice_number="INV-DASH-OUT-4",
            status=Invoice.STATUS_OVERDUE,
            total_amount="80.00",
            issue_date=today - timedelta(days=20),
            due_date=today - timedelta(days=2),
        )
        self._create_invoice(
            invoice_number="INV-DASH-OUT-5",
            status=Invoice.STATUS_PAID,
            total_amount="218.00",
            issue_date=today - timedelta(days=6),
            due_date=today + timedelta(days=1),
        )
        self._create_invoice(
            invoice_number="INV-DASH-OUT-6",
            status=Invoice.STATUS_REFUNDED,
            total_amount="54.50",
            issue_date=today - timedelta(days=8),
            due_date=today - timedelta(days=1),
        )

        self.client.force_login(self.finance_user)
        response = self.client.get(reverse("invoice-dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["outstanding_amount"], Decimal("430.00"))
        self.assertEqual(response.context["draft_count"], 1)
        self.assertContains(response, "Draft Invoices")
        self.assertContains(response, "Sent, viewed, and overdue invoices still waiting for payment.")
        self.assertNotContains(response, "Draft, sent, viewed, and overdue invoices still waiting for payment.")

    def test_invoice_dashboard_renders_operational_sections_and_preserves_calculations(self):
        today = timezone.localdate()
        previous_month_day = (today.replace(day=1) - timedelta(days=3)).replace(day=12)
        draft_invoice = self._create_invoice(
            invoice_number="INV-DASH-1001",
            status=Invoice.STATUS_DRAFT,
            total_amount="109.00",
            issue_date=today,
            due_date=today + timedelta(days=7),
        )
        sent_invoice = self._create_invoice(
            invoice_number="INV-DASH-1002",
            status=Invoice.STATUS_SENT,
            total_amount="150.00",
            issue_date=today - timedelta(days=2),
            due_date=today + timedelta(days=5),
        )
        viewed_invoice = self._create_invoice(
            invoice_number="INV-DASH-1003",
            status=Invoice.STATUS_VIEWED,
            total_amount="200.00",
            issue_date=today - timedelta(days=4),
            due_date=today + timedelta(days=3),
        )
        overdue_invoice = self._create_invoice(
            invoice_number="INV-DASH-1004",
            status=Invoice.STATUS_OVERDUE,
            total_amount="80.00",
            issue_date=today - timedelta(days=20),
            due_date=today - timedelta(days=2),
        )
        self._create_invoice(
            invoice_number="INV-DASH-1005",
            status=Invoice.STATUS_REFUNDED,
            total_amount="54.50",
            issue_date=today - timedelta(days=8),
            due_date=today - timedelta(days=1),
        )
        paid_this_month = self._create_invoice(
            invoice_number="INV-DASH-1006",
            status=Invoice.STATUS_PAID,
            total_amount="218.00",
            issue_date=today - timedelta(days=6),
            due_date=today + timedelta(days=1),
        )
        paid_previous_month = self._create_invoice(
            invoice_number="INV-DASH-1007",
            status=Invoice.STATUS_PAID,
            total_amount="327.00",
            issue_date=previous_month_day - timedelta(days=5),
            due_date=previous_month_day + timedelta(days=7),
        )
        PaymentRecord.objects.create(
            invoice=paid_this_month,
            payment_reference="PAY-DASH-1006",
            provider=PaymentRecord.PROVIDER_STRIPE,
            status=PaymentRecord.STATUS_SUCCEEDED,
            amount=Decimal("218.00"),
            currency="SGD",
            paid_at=self._aware_datetime(today, hour=11),
        )
        PaymentRecord.objects.create(
            invoice=paid_previous_month,
            payment_reference="PAY-DASH-1007",
            provider=PaymentRecord.PROVIDER_MANUAL,
            status=PaymentRecord.STATUS_SUCCEEDED,
            amount=Decimal("327.00"),
            currency="SGD",
            paid_at=self._aware_datetime(previous_month_day, hour=9),
        )
        EmailDeliveryLog.objects.create(
            recipient_email=self.customer.email,
            subject="Invoice email",
            template_key="invoice_email_v1",
            status=EmailDeliveryLog.STATUS_FAILED,
            related_object_type="invoice",
            related_object_id=str(sent_invoice.pk),
            error_message="SMTP timeout",
        )
        ImportJob.objects.create(
            module=ImportJob.MODULE_INVOICING,
            source_file_name="invoice_upload_july.xlsx",
            status=ImportJob.STATUS_COMPLETED_WITH_ERRORS,
            total_rows=5,
            valid_rows=3,
            invalid_rows=2,
            saved_rows=3,
            initiated_by=self.finance_user,
        )

        self.client.force_login(self.finance_user)
        response = self.client.get(reverse("invoice-dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Invoice Dashboard")
        self.assertContains(
            response,
            "Use this dashboard to manage daily invoice work, review follow-up priorities, and act on delivery or import issues.",
        )
        self.assertContains(response, "Reporting period:")
        self.assertContains(response, "Last updated:")
        self.assertContains(response, "Collected This Month")
        self.assertContains(response, "Outstanding Amount")
        self.assertContains(response, "Overdue Amount")
        self.assertContains(response, "Invoices Requiring Follow-up")
        self.assertContains(response, "Sent, viewed, and overdue invoices still waiting for payment.")
        self.assertContains(response, "Draft Invoices")
        self.assertContains(response, "Pending Payment")
        self.assertContains(response, "Viewed Invoices")
        self.assertContains(response, "Total Invoices")
        self.assertContains(response, reverse("invoice-create"))
        self.assertContains(response, reverse("invoice-csv-upload"))
        self.assertContains(response, reverse("invoice-list"))
        self.assertContains(response, reverse("invoice-customer-create"))
        self.assertContains(response, reverse("invoice-template-settings"))
        self.assertContains(response, "Invoice Attention")
        self.assertContains(response, "Overdue invoices")
        self.assertContains(response, "Draft invoices not sent")
        self.assertContains(response, "Failed invoice email deliveries")
        self.assertContains(response, "Import validation issues")
        self.assertContains(response, "Invoice Category Queue")
        self.assertContains(response, "Viewing: Bank Transfer to Verify")
        self.assertNotContains(response, draft_invoice.invoice_number)
        self.assertNotContains(response, viewed_invoice.invoice_number)
        self.assertNotContains(response, overdue_invoice.invoice_number)
        self.assertEqual(response.context["collected_month"], Decimal("218.00"))
        self.assertEqual(response.context["collected_year"], Decimal("545.00"))
        self.assertEqual(response.context["outstanding_amount"], Decimal("430.00"))
        self.assertEqual(response.context["overdue_amount"], Decimal("80.00"))
        self.assertEqual(response.context["total_invoices"], 7)
        self.assertEqual(response.context["invoices_requiring_follow_up_count"], 4)
        self.assertEqual(response.context["draft_count"], 1)
        self.assertEqual(response.context["pending_payment_count"], 1)
        self.assertEqual(response.context["viewed_count"], 1)
        self.assertEqual(response.context["refunded_count"], 1)
        self.assertEqual(response.context["failed_invoice_email_count"], 1)
        self.assertEqual(response.context["import_validation_issue_count"], 2)
        self.assertEqual(response.context["selected_invoice_category_key"], "bank_transfer_to_verify")
        self.assertEqual(response.context["selected_invoice_category"]["label"], "Bank Transfer to Verify")
        self.assertEqual(list(response.context["category_invoices"]), [])

    def test_invoice_dashboard_shows_submitted_bank_transfer_verification_card(self):
        invoice = self._create_invoice(
            invoice_number="INV-DASH-BANK-VERIFY",
            status=Invoice.STATUS_OVERDUE,
            total_amount="109.00",
            issue_date=timezone.localdate() - timedelta(days=20),
            due_date=timezone.localdate() - timedelta(days=2),
        )
        self._create_submitted_bank_transfer(invoice)
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse("invoice-dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["submitted_bank_transfer_count"], 1)
        self.assertContains(response, "Bank Transfers To Verify")
        self.assertContains(response, "Customer-submitted bank transfer notices waiting for Finance verification.")
        self.assertContains(response, "Bank transfers awaiting verification")
        self.assertContains(response, "Review transfers")
        self.assertContains(response, reverse("payment-stripe-report"))
        self.assertContains(response, "Viewing: Bank Transfer to Verify")
        self.assertContains(response, invoice.invoice_number)
        self.assertContains(response, reverse("invoice-detail", args=[invoice.pk]))

    def test_invoice_dashboard_action_labels_match_status(self):
        today = timezone.localdate()
        draft_invoice = self._create_invoice(
            invoice_number="INV-DASH-ACT-1",
            status=Invoice.STATUS_DRAFT,
            total_amount="109.00",
            issue_date=today,
            due_date=today + timedelta(days=7),
        )
        sent_invoice = self._create_invoice(
            invoice_number="INV-DASH-ACT-2",
            status=Invoice.STATUS_SENT,
            total_amount="150.00",
            issue_date=today - timedelta(days=1),
            due_date=today + timedelta(days=2),
        )
        self._create_invoice(
            invoice_number="INV-DASH-ACT-3",
            status=Invoice.STATUS_VIEWED,
            total_amount="200.00",
            issue_date=today - timedelta(days=2),
            due_date=today + timedelta(days=1),
        )
        self._create_invoice(
            invoice_number="INV-DASH-ACT-4",
            status=Invoice.STATUS_OVERDUE,
            total_amount="80.00",
            issue_date=today - timedelta(days=20),
            due_date=today - timedelta(days=2),
        )

        self._create_submitted_bank_transfer(sent_invoice)
        self.client.force_login(self.finance_user)
        response = self.client.get(reverse("invoice-dashboard"), data={"category": "requires_follow_up"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("invoice-edit", args=[draft_invoice.pk]))
        self.assertContains(response, "Edit Draft")
        self.assertContains(response, "Send Invoice")
        self.assertContains(response, reverse("invoice-detail", args=[sent_invoice.pk]))
        self.assertContains(response, "View Invoice")
        self.assertContains(response, "Send Reminder")

    def test_invoice_dashboard_defaults_to_bank_transfer_category_and_rejects_invalid_category(self):
        today = timezone.localdate()
        bank_invoice = self._create_invoice(
            invoice_number="INV-DASH-DEFAULT-BANK",
            status=Invoice.STATUS_SENT,
            total_amount="109.00",
            issue_date=today - timedelta(days=1),
            due_date=today + timedelta(days=2),
        )
        outstanding_invoice = self._create_invoice(
            invoice_number="INV-DASH-DEFAULT-OUT",
            status=Invoice.STATUS_SENT,
            total_amount="150.00",
            issue_date=today,
            due_date=today + timedelta(days=5),
        )
        self._create_submitted_bank_transfer(bank_invoice)

        self.client.force_login(self.finance_user)
        response = self.client.get(reverse("invoice-dashboard"))
        invalid_response = self.client.get(reverse("invoice-dashboard"), data={"category": "all"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_invoice_category_key"], "bank_transfer_to_verify")
        self.assertContains(response, "Viewing: Bank Transfer to Verify")
        self.assertContains(response, bank_invoice.invoice_number)
        self.assertNotContains(response, outstanding_invoice.invoice_number)
        self.assertEqual(invalid_response.context["selected_invoice_category_key"], "bank_transfer_to_verify")
        self.assertContains(invalid_response, "Viewing: Bank Transfer to Verify")
        self.assertContains(invalid_response, bank_invoice.invoice_number)
        self.assertNotContains(invalid_response, outstanding_invoice.invoice_number)

    def test_invoice_dashboard_filters_outstanding_category(self):
        today = timezone.localdate()
        draft_invoice = self._create_invoice(
            invoice_number="INV-DASH-FILTER-DRAFT",
            status=Invoice.STATUS_DRAFT,
            total_amount="109.00",
            issue_date=today,
            due_date=today + timedelta(days=7),
        )
        sent_invoice = self._create_invoice(
            invoice_number="INV-DASH-FILTER-SENT",
            status=Invoice.STATUS_SENT,
            total_amount="150.00",
            issue_date=today - timedelta(days=2),
            due_date=today + timedelta(days=5),
        )
        viewed_invoice = self._create_invoice(
            invoice_number="INV-DASH-FILTER-VIEWED",
            status=Invoice.STATUS_VIEWED,
            total_amount="200.00",
            issue_date=today - timedelta(days=3),
            due_date=today + timedelta(days=4),
        )
        overdue_invoice = self._create_invoice(
            invoice_number="INV-DASH-FILTER-OVERDUE",
            status=Invoice.STATUS_OVERDUE,
            total_amount="80.00",
            issue_date=today - timedelta(days=20),
            due_date=today - timedelta(days=2),
        )
        paid_invoice = self._create_invoice(
            invoice_number="INV-DASH-FILTER-PAID",
            status=Invoice.STATUS_PAID,
            total_amount="218.00",
            issue_date=today,
            due_date=today + timedelta(days=7),
        )

        self.client.force_login(self.finance_user)
        response = self.client.get(reverse("invoice-dashboard"), data={"category": "outstanding"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_invoice_category_key"], "outstanding")
        self.assertContains(response, "Viewing: Outstanding")
        self.assertContains(response, sent_invoice.invoice_number)
        self.assertContains(response, viewed_invoice.invoice_number)
        self.assertContains(response, overdue_invoice.invoice_number)
        self.assertContains(response, reverse("invoice-detail", args=[sent_invoice.pk]))
        self.assertNotContains(response, draft_invoice.invoice_number)
        self.assertNotContains(response, paid_invoice.invoice_number)

    def test_invoice_dashboard_filters_requires_follow_up_category(self):
        today = timezone.localdate()
        draft_invoice = self._create_invoice(
            invoice_number="INV-DASH-FOLLOW-DRAFT",
            status=Invoice.STATUS_DRAFT,
            total_amount="109.00",
            issue_date=today,
            due_date=today + timedelta(days=7),
        )
        sent_invoice = self._create_invoice(
            invoice_number="INV-DASH-FOLLOW-SENT",
            status=Invoice.STATUS_SENT,
            total_amount="150.00",
            issue_date=today - timedelta(days=2),
            due_date=today + timedelta(days=5),
        )
        paid_invoice = self._create_invoice(
            invoice_number="INV-DASH-FOLLOW-PAID",
            status=Invoice.STATUS_PAID,
            total_amount="218.00",
            issue_date=today,
            due_date=today + timedelta(days=7),
        )

        self.client.force_login(self.finance_user)
        response = self.client.get(reverse("invoice-dashboard"), data={"category": "requires_follow_up"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Viewing: Requires Follow Up")
        self.assertContains(response, draft_invoice.invoice_number)
        self.assertContains(response, sent_invoice.invoice_number)
        self.assertNotContains(response, paid_invoice.invoice_number)
        self.assertContains(response, reverse("invoice-edit", args=[draft_invoice.pk]))
        self.assertContains(response, "Edit Draft")
        self.assertContains(response, "Send Invoice")

    def test_invoice_dashboard_filters_overdue_category(self):
        today = timezone.localdate()
        overdue_invoice = self._create_invoice(
            invoice_number="INV-DASH-ONLY-OVERDUE",
            status=Invoice.STATUS_OVERDUE,
            total_amount="80.00",
            issue_date=today - timedelta(days=20),
            due_date=today - timedelta(days=2),
        )
        sent_invoice = self._create_invoice(
            invoice_number="INV-DASH-NOT-OVERDUE",
            status=Invoice.STATUS_SENT,
            total_amount="150.00",
            issue_date=today,
            due_date=today + timedelta(days=5),
        )

        self.client.force_login(self.finance_user)
        response = self.client.get(reverse("invoice-dashboard"), data={"category": "overdue"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Viewing: Overdue")
        self.assertContains(response, overdue_invoice.invoice_number)
        self.assertNotContains(response, sent_invoice.invoice_number)

    def test_invoice_dashboard_bank_transfer_category_handles_overlap_and_payment_status(self):
        today = timezone.localdate()
        overlapping_invoice = self._create_invoice(
            invoice_number="INV-DASH-BANK-OVERLAP",
            status=Invoice.STATUS_OVERDUE,
            total_amount="109.00",
            issue_date=today - timedelta(days=20),
            due_date=today - timedelta(days=2),
        )
        succeeded_invoice = self._create_invoice(
            invoice_number="INV-DASH-BANK-SUCCEEDED",
            status=Invoice.STATUS_PAID,
            total_amount="150.00",
            issue_date=today - timedelta(days=4),
            due_date=today + timedelta(days=2),
        )
        self._create_submitted_bank_transfer(overlapping_invoice, payment_reference="BANK-DASH-OVERLAP-001")
        self._create_submitted_bank_transfer(overlapping_invoice, payment_reference="BANK-DASH-OVERLAP-002")
        self._create_submitted_bank_transfer(
            succeeded_invoice,
            payment_reference="BANK-DASH-SUCCEEDED",
            status=PaymentRecord.STATUS_SUCCEEDED,
        )

        self.client.force_login(self.finance_user)
        bank_response = self.client.get(reverse("invoice-dashboard"), data={"category": "bank_transfer_to_verify"})
        overdue_response = self.client.get(reverse("invoice-dashboard"), data={"category": "overdue"})

        self.assertEqual(bank_response.status_code, 200)
        self.assertEqual(bank_response.context["submitted_bank_transfer_count"], 1)
        self.assertEqual(bank_response.context["category_invoices"].count(), 1)
        self.assertContains(bank_response, overlapping_invoice.invoice_number, count=1)
        self.assertContains(bank_response, reverse("invoice-detail", args=[overlapping_invoice.pk]))
        self.assertNotContains(bank_response, succeeded_invoice.invoice_number)
        self.assertContains(overdue_response, overlapping_invoice.invoice_number)

    def test_invoice_dashboard_shows_empty_states_when_no_invoice_data_exists(self):
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse("invoice-dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No invoice issues need attention right now.")
        self.assertContains(response, "No invoice support tickets require action.")
        self.assertContains(response, "No invoices match Bank Transfer to Verify.")


class InvoiceFileUploadTests(TestCase):
    def setUp(self):
        self.finance_user = User.objects.create_user(username="finance_upload", password="TempPass123!")
        self.finance_user.role_profile.role = FINANCE
        self.finance_user.role_profile.save()

        self.admin_user = User.objects.create_user(username="admin_upload", password="TempPass123!")
        self.admin_user.role_profile.role = ADMIN
        self.admin_user.role_profile.save()

        self.superadmin_user = User.objects.create_user(username="super_upload", password="TempPass123!")
        self.superadmin_user.role_profile.role = SUPERADMIN
        self.superadmin_user.role_profile.save()

        self.hr_user = User.objects.create_user(username="hr_upload", password="TempPass123!")
        self.hr_user.role_profile.role = HR
        self.hr_user.role_profile.save()

        self.staff_user = User.objects.create_user(username="staff_upload", password="TempPass123!")
        self.staff_user.role_profile.role = STAFF
        self.staff_user.role_profile.save()

        self.customer_user = User.objects.create_user(username="customer_upload", password="TempPass123!")
        self.customer_user.role_profile.role = CUSTOMER
        self.customer_user.role_profile.save()

    def _sample_csv_rows(self):
        return [
            {
                "seller_id": "S1",
                "shop_title": "Acme Salon",
                "OrderID": "ORD-001",
                "paymentMethod": "Credit Card",
                "email": "billing@acme.com",
                "customerName": "Acme Customer",
                "qty": "2",
                "serviceName": "Hair Cut",
                "bookedDate": "2026-05-01 10:00:00",
                "vanidayShare": "120.00",
            },
            {
                "seller_id": "S1",
                "shop_title": "Acme Salon",
                "OrderID": "ORD-002",
                "paymentMethod": "Credit Card",
                "email": "billing@acme.com",
                "customerName": "Acme Customer",
                "qty": "1",
                "serviceName": "Nail Service",
                "bookedDate": "2026-05-02 11:00:00",
                "vanidayShare": "80.00",
            },
        ]

    def _build_csv_upload(self, rows=None, filename="vaniday_sample.csv"):
        rows = rows or self._sample_csv_rows()
        headers = [
            "seller_id",
            "shop_title",
            "OrderID",
            "paymentMethod",
            "email",
            "customerName",
            "qty",
            "serviceName",
            "bookedDate",
            "vanidayShare",
        ]
        lines = [",".join(headers)]
        for row in rows:
            lines.append(",".join(str(row.get(header, "")) for header in headers))
        content = ("\n".join(lines) + "\n").encode("utf-8")
        return SimpleUploadedFile(filename, content, content_type="text/csv")

    def _build_excel_upload(self, rows=None, filename="vaniday_sample.xlsx", headers=None):
        workbook = Workbook()
        sheet = workbook.active
        header_row = headers or [
            "seller_id",
            "shop_title",
            "OrderID",
            "paymentMethod",
            "email",
            "customerName",
            "qty",
            "serviceName",
            "bookedDate",
            "vanidayShare",
        ]
        sheet.append(header_row)
        for row in rows or self._sample_csv_rows():
            if isinstance(row, dict):
                sheet.append([row.get(header, "") for header in header_row])
            else:
                sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return SimpleUploadedFile(
            filename,
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _parse_single_csv_row(self, **overrides):
        row = self._sample_csv_rows()[0].copy()
        row.update(overrides)
        return parse_invoice_csv(self._build_csv_upload(rows=[row], filename="single_row.csv"))

    def _assert_single_csv_row_invalid(self, expected_error, **overrides):
        parsed = self._parse_single_csv_row(**overrides)
        self.assertEqual(len(parsed["valid_rows"]), 0)
        self.assertEqual(len(parsed["invalid_rows"]), 1)
        self.assertIn(expected_error, parsed["invalid_rows"][0]["errors"])
        return parsed["invalid_rows"][0]

    def _preview_csv_rows(self, rows, filename="duplicate_test.csv"):
        return self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": self._build_csv_upload(rows=rows, filename=filename)},
        )

    def _confirm_preview(self, preview_response):
        import_token = preview_response.context["preview"]["import_token"]
        return self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "confirm", "import_token": import_token},
        )

    def _preview_and_confirm_csv_rows(self, rows, filename="duplicate_test.csv"):
        preview_response = self._preview_csv_rows(rows, filename=filename)
        self.assertEqual(preview_response.status_code, 200)
        confirm_response = self._confirm_preview(preview_response)
        self.assertEqual(confirm_response.status_code, 302)
        return preview_response, confirm_response

    def test_valid_invoice_upload_row_passes(self):
        parsed = self._parse_single_csv_row()

        self.assertEqual(len(parsed["valid_rows"]), 1)
        self.assertEqual(len(parsed["invalid_rows"]), 0)
        valid_row = parsed["valid_rows"][0]
        self.assertEqual(valid_row["source"]["order_id"], "ORD-001")
        self.assertEqual(valid_row["customer_name"], "Acme Salon")
        self.assertEqual(valid_row["email"], "billing@acme.com")
        self.assertEqual(valid_row["amount"], "120.00")

    def test_missing_email_is_invalid_for_invoice_upload(self):
        self._assert_single_csv_row_invalid("Missing customer email.", email="")

    def test_invalid_email_format_is_invalid_for_invoice_upload(self):
        self._assert_single_csv_row_invalid("Customer email format is invalid.", email="not-an-email")

    def test_missing_customer_or_merchant_name_is_invalid_for_invoice_upload(self):
        self._assert_single_csv_row_invalid(
            "Missing merchant/customer name (shop_title or customerName).",
            shop_title="",
            customerName="",
        )

    def test_missing_order_reference_is_invalid_for_invoice_upload(self):
        self._assert_single_csv_row_invalid("Missing OrderID/order reference.", OrderID="")

    def test_missing_service_name_is_invalid_for_invoice_upload(self):
        self._assert_single_csv_row_invalid("Missing serviceName/item description.", serviceName="")

    def test_missing_booked_date_is_invalid_for_invoice_upload(self):
        self._assert_single_csv_row_invalid("Missing bookedDate.", bookedDate="")

    def test_invalid_booked_date_is_invalid_for_invoice_upload(self):
        self._assert_single_csv_row_invalid("bookedDate format is invalid.", bookedDate="not-a-date")

    def test_missing_invoice_amount_is_invalid_for_invoice_upload(self):
        self._assert_single_csv_row_invalid("Missing invoice amount (vanidayShare).", vanidayShare="")

    def test_non_numeric_invoice_amount_is_invalid_for_invoice_upload(self):
        self._assert_single_csv_row_invalid(
            "Invoice amount (vanidayShare) must be numeric.",
            vanidayShare="not-a-number",
        )

    def test_mixed_valid_and_invalid_csv_rows_show_correct_counts(self):
        rows = self._sample_csv_rows()
        rows.append(
            {
                "seller_id": "S3",
                "shop_title": "",
                "OrderID": "ORD-003",
                "paymentMethod": "Credit Card",
                "email": "",
                "customerName": "",
                "qty": "1",
                "serviceName": "",
                "bookedDate": "not-a-date",
                "vanidayShare": "",
            }
        )
        self.client.login(username="finance_upload", password="TempPass123!")

        response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": self._build_csv_upload(rows=rows, filename="mixed_rows.csv")},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Preview Summary")
        self.assertContains(response, "Validation Issues")
        preview = response.context["preview"]
        self.assertEqual(preview["total_rows"], 3)
        self.assertEqual(len(preview["valid_rows"]), 2)
        self.assertEqual(len(preview["invalid_rows"]), 1)
        invalid_errors = preview["invalid_rows"][0]["errors"]
        self.assertIn("Missing customer email.", invalid_errors)
        self.assertIn("Missing invoice amount (vanidayShare).", invalid_errors)

    def test_reuploading_same_file_marks_rows_as_duplicates(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        rows = self._sample_csv_rows()

        self._preview_and_confirm_csv_rows(rows, filename="first_import.csv")
        first_invoice_count = Invoice.objects.count()
        first_item_count = InvoiceItem.objects.count()

        second_preview = self._preview_csv_rows(rows, filename="first_import_again.csv")

        self.assertEqual(second_preview.status_code, 200)
        preview = second_preview.context["preview"]
        self.assertEqual(len(preview["valid_rows"]), 0)
        self.assertEqual(len(preview["invalid_rows"]), 2)
        self.assertContains(second_preview, "This order/service row has already been imported.")

        second_confirm = self._confirm_preview(second_preview)

        self.assertEqual(second_confirm.status_code, 302)
        self.assertEqual(Invoice.objects.count(), first_invoice_count)
        self.assertEqual(InvoiceItem.objects.count(), first_item_count)
        latest_job = ImportJob.objects.latest("id")
        self.assertEqual(latest_job.saved_rows, 0)

    def test_mixed_valid_invalid_and_duplicate_rows_show_correct_counts(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        imported_row = self._sample_csv_rows()[0]
        new_row = self._sample_csv_rows()[1].copy()
        invalid_row = self._sample_csv_rows()[0].copy()
        invalid_row.update({"OrderID": "ORD-INVALID", "email": "", "serviceName": ""})
        self._preview_and_confirm_csv_rows([imported_row], filename="initial_valid.csv")

        response = self._preview_csv_rows([imported_row, new_row, invalid_row], filename="mixed_duplicate.csv")

        self.assertEqual(response.status_code, 200)
        preview = response.context["preview"]
        self.assertEqual(preview["total_rows"], 3)
        self.assertEqual(len(preview["valid_rows"]), 1)
        self.assertEqual(len(preview["invalid_rows"]), 2)
        errors_by_order = {
            row["source"]["order_id"]: row["errors"]
            for row in preview["invalid_rows"]
        }
        self.assertIn("This order/service row has already been imported.", errors_by_order["ORD-001"])
        self.assertIn("Missing customer email.", errors_by_order["ORD-INVALID"])

    def test_confirm_import_skips_duplicate_rows_and_imports_new_rows(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        imported_row = self._sample_csv_rows()[0]
        new_row = self._sample_csv_rows()[1]
        self._preview_and_confirm_csv_rows([imported_row], filename="initial_valid.csv")
        first_invoice_count = Invoice.objects.count()
        first_item_count = InvoiceItem.objects.count()

        preview_response = self._preview_csv_rows([imported_row, new_row], filename="duplicate_plus_new.csv")
        confirm_response = self._confirm_preview(preview_response)

        self.assertEqual(confirm_response.status_code, 302)
        self.assertEqual(Invoice.objects.count(), first_invoice_count + 1)
        self.assertEqual(InvoiceItem.objects.count(), first_item_count + 1)
        self.assertEqual(InvoiceItem.objects.filter(description__icontains="Nail Service").count(), 1)
        latest_job = ImportJob.objects.latest("id")
        self.assertEqual(latest_job.saved_rows, 1)
        self.assertEqual(latest_job.invalid_rows, 1)

    def test_corrected_previously_invalid_row_can_be_imported_later(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        valid_row = self._sample_csv_rows()[0]
        invalid_row = self._sample_csv_rows()[1].copy()
        invalid_row.update({"shop_title": "", "customerName": ""})
        self._preview_and_confirm_csv_rows([valid_row, invalid_row], filename="with_invalid.csv")
        first_invoice_count = Invoice.objects.count()
        first_item_count = InvoiceItem.objects.count()

        corrected_row = invalid_row.copy()
        corrected_row.update({"shop_title": "Acme Salon", "customerName": "Acme Customer"})
        corrected_preview = self._preview_csv_rows([valid_row, corrected_row], filename="corrected_full.csv")

        self.assertEqual(corrected_preview.status_code, 200)
        preview = corrected_preview.context["preview"]
        self.assertEqual(len(preview["valid_rows"]), 1)
        self.assertEqual(len(preview["invalid_rows"]), 1)
        self.assertEqual(preview["valid_rows"][0]["source"]["order_id"], "ORD-002")
        self.assertIn("This order/service row has already been imported.", preview["invalid_rows"][0]["errors"])

        confirm_response = self._confirm_preview(corrected_preview)

        self.assertEqual(confirm_response.status_code, 302)
        self.assertEqual(Invoice.objects.count(), first_invoice_count + 1)
        self.assertEqual(InvoiceItem.objects.count(), first_item_count + 1)
        self.assertEqual(InvoiceItem.objects.filter(description__icontains="Nail Service").count(), 1)

    def test_same_order_id_with_different_service_name_can_be_imported(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        imported_row = self._sample_csv_rows()[0]
        separate_service_row = imported_row.copy()
        separate_service_row["serviceName"] = "Color Treatment"
        self._preview_and_confirm_csv_rows([imported_row], filename="single_service.csv")

        preview_response = self._preview_csv_rows([separate_service_row], filename="separate_service.csv")

        self.assertEqual(preview_response.status_code, 200)
        self.assertEqual(len(preview_response.context["preview"]["valid_rows"]), 1)
        self.assertEqual(len(preview_response.context["preview"]["invalid_rows"]), 0)

    def test_same_order_id_and_service_name_with_different_amount_can_be_imported(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        imported_row = self._sample_csv_rows()[0]
        adjusted_amount_row = imported_row.copy()
        adjusted_amount_row["vanidayShare"] = "125.00"
        self._preview_and_confirm_csv_rows([imported_row], filename="original_amount.csv")

        preview_response = self._preview_csv_rows([adjusted_amount_row], filename="adjusted_amount.csv")

        self.assertEqual(preview_response.status_code, 200)
        self.assertEqual(len(preview_response.context["preview"]["valid_rows"]), 1)
        self.assertEqual(len(preview_response.context["preview"]["invalid_rows"]), 0)

    def test_finance_can_upload_valid_csv_file(self):
        self.client.login(username="finance_upload", password="TempPass123!")

        response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": self._build_csv_upload()},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Preview Summary")
        self.assertContains(response, "vaniday_sample.csv")

    def test_finance_can_upload_valid_xlsx_file(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        rows = [
            [
                "S1",
                "Acme Salon",
                "ORD-001",
                "Credit Card",
                "billing@acme.com",
                "Acme Customer",
                2,
                "Hair Cut",
                timezone.datetime(2026, 5, 1, 10, 0),
                120,
            ],
            [
                "S1",
                "Acme Salon",
                "ORD-002",
                "Credit Card",
                "billing@acme.com",
                "Acme Customer",
                1,
                "Nail Service",
                timezone.datetime(2026, 5, 2, 11, 0),
                80.5,
            ],
            [None, None, None, None, None, None, None, None, None, None],
        ]

        response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": self._build_excel_upload(rows=rows)},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Preview Summary")
        self.assertContains(response, "Invoice Upload Preview")
        preview = response.context["preview"]
        self.assertEqual(preview["total_rows"], 2)
        self.assertEqual(len(preview["valid_rows"]), 2)
        self.assertEqual(len(preview["invalid_rows"]), 0)
        self.assertEqual(preview["valid_rows"][0]["amount"], "120.00")
        self.assertEqual(preview["valid_rows"][1]["amount"], "80.50")

    def test_admin_and_superadmin_can_access_invoice_upload(self):
        for username in ["admin_upload", "super_upload"]:
            with self.subTest(username=username):
                self.client.login(username=username, password="TempPass123!")
                response = self.client.get(reverse("invoice-csv-upload"))
                self.assertEqual(response.status_code, 200)
                self.client.logout()

    def test_hr_customer_and_staff_cannot_access_invoice_upload(self):
        for username in ["hr_upload", "customer_upload", "staff_upload"]:
            with self.subTest(username=username):
                self.client.login(username=username, password="TempPass123!")
                response = self.client.get(reverse("invoice-csv-upload"))
                self.assertEqual(response.status_code, 403)
                self.client.logout()

    def test_unsupported_extension_is_rejected(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        upload_file = SimpleUploadedFile("bad.pdf", b"%PDF-1.4", content_type="application/pdf")

        response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": upload_file},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Upload a CSV or Excel (.xlsx) file.")

    def test_corrupted_xlsx_is_rejected_safely(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        upload_file = SimpleUploadedFile(
            "broken.xlsx",
            b"this-is-not-a-real-workbook",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": upload_file},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Excel workbook is corrupted or unreadable.")

    def test_empty_excel_workbook_is_rejected(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        workbook = Workbook()
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        upload_file = SimpleUploadedFile(
            "empty.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": upload_file},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Excel workbook is empty.")

    def test_missing_required_excel_headers_are_reported(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["invoice_number", "issue_date", "total"])
        sheet.append(["INV-001", "2026-05-01", 120])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        upload_file = SimpleUploadedFile(
            "wrong_headers.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": upload_file},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Missing required columns for:")

    def test_excel_values_are_normalized_like_csv(self):
        csv_rows = [
            {
                "seller_id": "S1",
                "shop_title": "Acme Salon",
                "OrderID": "12345",
                "paymentMethod": "Credit Card",
                "email": "billing@acme.com",
                "customerName": "Acme Customer",
                "qty": "2",
                "serviceName": "Hair Cut",
                "bookedDate": "2026-05-01 10:00:00",
                "vanidayShare": "120.50",
            }
        ]
        excel_rows = [
            [
                "S1",
                "Acme Salon",
                12345,
                "Credit Card",
                "billing@acme.com",
                "Acme Customer",
                2,
                "Hair Cut",
                timezone.datetime(2026, 5, 1, 10, 0),
                120.5,
            ]
        ]

        parsed_csv = parse_invoice_csv(self._build_csv_upload(rows=csv_rows))
        parsed_excel = parse_invoice_excel(self._build_excel_upload(rows=excel_rows))

        csv_row = parsed_csv["valid_rows"][0]
        excel_row = parsed_excel["valid_rows"][0]
        self.assertEqual(excel_row["source"]["order_id"], "12345")
        self.assertEqual(csv_row["amount"], excel_row["amount"])
        self.assertEqual(csv_row["quantity"], excel_row["quantity"])
        self.assertEqual(csv_row["group_period"], excel_row["group_period"])
        self.assertEqual(csv_row["item_description"], excel_row["item_description"])

    def test_excel_preview_shows_valid_and_invalid_rows(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        rows = [
            [
                "S1",
                "Acme Salon",
                "ORD-001",
                "Credit Card",
                "billing@acme.com",
                "Acme Customer",
                2,
                "Hair Cut",
                timezone.datetime(2026, 5, 1, 10, 0),
                120,
            ],
            [
                "S1",
                "",
                "ORD-002",
                "Credit Card",
                "",
                "",
                1,
                "Nail Service",
                timezone.datetime(2026, 5, 2, 11, 0),
                "",
            ],
        ]

        response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": self._build_excel_upload(rows=rows)},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Validation Issues")
        preview = response.context["preview"]
        self.assertEqual(len(preview["valid_rows"]), 1)
        self.assertEqual(len(preview["invalid_rows"]), 1)

    def test_confirmation_saves_valid_excel_rows_only(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        rows = [
            [
                "S1",
                "Acme Salon",
                "ORD-001",
                "Credit Card",
                "billing@acme.com",
                "Acme Customer",
                2,
                "Hair Cut",
                timezone.datetime(2026, 5, 1, 10, 0),
                120,
            ],
            [
                "S1",
                "Acme Salon",
                "ORD-002",
                "Credit Card",
                "billing@acme.com",
                "Acme Customer",
                1,
                "Nail Service",
                timezone.datetime(2026, 5, 2, 11, 0),
                80,
            ],
            [
                "S1",
                "",
                "ORD-003",
                "Credit Card",
                "",
                "",
                1,
                "Bad Row",
                timezone.datetime(2026, 5, 3, 12, 0),
                "",
            ],
        ]
        preview_response = self.client.post(
            reverse("invoice-csv-upload"),
            data={
                "action": "preview",
                "csv_file": self._build_excel_upload(rows=rows, filename="vaniday_batch.xlsx"),
            },
        )
        self.assertEqual(preview_response.status_code, 200)
        import_token = preview_response.context["preview"]["import_token"]

        confirm_response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "confirm", "import_token": import_token},
        )

        self.assertEqual(confirm_response.status_code, 302)
        imported_invoice = Invoice.objects.get(notes__contains="vaniday_batch.xlsx")
        self.assertEqual(imported_invoice.items.count(), 2)
        self.assertFalse(imported_invoice.items.filter(description__icontains="Bad Row").exists())
        self.assertEqual(imported_invoice.total_amount, Decimal("200.00"))
        self.assertEqual(InvoiceSourceRow.objects.filter(source_file_name="vaniday_batch.xlsx").count(), 3)
        job = ImportJob.objects.latest("id")
        self.assertEqual(job.saved_rows, 2)
        self.assertEqual(job.invalid_rows, 1)

    def test_excel_duplicate_rows_keep_existing_grouping_behavior(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        duplicate_row = [
            "S1",
            "Acme Salon",
            "ORD-001",
            "Credit Card",
            "billing@acme.com",
            "Acme Customer",
            1,
            "Hair Cut",
            timezone.datetime(2026, 5, 1, 10, 0),
            120,
        ]
        preview_response = self.client.post(
            reverse("invoice-csv-upload"),
            data={
                "action": "preview",
                "csv_file": self._build_excel_upload(rows=[duplicate_row, duplicate_row], filename="duplicates.xlsx"),
            },
        )
        import_token = preview_response.context["preview"]["import_token"]

        self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "confirm", "import_token": import_token},
        )

        imported_invoice = Invoice.objects.get(notes__contains="duplicates.xlsx")
        self.assertEqual(imported_invoice.items.count(), 2)
        self.assertEqual(imported_invoice.total_amount, Decimal("240.00"))

    def test_new_upload_replaces_previous_preview_in_same_session(self):
        self.client.login(username="finance_upload", password="TempPass123!")
        first_response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": self._build_csv_upload(filename="first.csv")},
        )
        first_token = first_response.context["preview"]["import_token"]

        second_rows = [
            {
                "seller_id": "S2",
                "shop_title": "Beta Salon",
                "OrderID": "ORD-101",
                "paymentMethod": "Credit Card",
                "email": "billing@beta.com",
                "customerName": "Beta Customer",
                "qty": "1",
                "serviceName": "Facial",
                "bookedDate": "2026-06-10 09:00:00",
                "vanidayShare": "50.00",
            }
        ]
        second_response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "preview", "csv_file": self._build_csv_upload(rows=second_rows, filename="second.csv")},
        )
        second_token = second_response.context["preview"]["import_token"]

        expired_response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "confirm", "import_token": first_token},
            follow=True,
        )
        valid_response = self.client.post(
            reverse("invoice-csv-upload"),
            data={"action": "confirm", "import_token": second_token},
        )

        self.assertEqual(expired_response.status_code, 200)
        self.assertContains(expired_response, "Import preview expired. Please upload the file again.")
        self.assertEqual(valid_response.status_code, 302)
        imported_invoice = Invoice.objects.get(notes__contains="second.csv")
        self.assertEqual(imported_invoice.total_amount, Decimal("50.00"))
