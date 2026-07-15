from __future__ import annotations

import re
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from payments.services import get_bank_transfer_details

from .models import Invoice, InvoiceTemplateSettings


LOGO_SIZE_DIMENSIONS_MM = {
    InvoiceTemplateSettings.LOGO_SIZE_SMALL: (28 * mm, 14 * mm),
    InvoiceTemplateSettings.LOGO_SIZE_MEDIUM: (36 * mm, 20 * mm),
    InvoiceTemplateSettings.LOGO_SIZE_LARGE: (48 * mm, 28 * mm),
}

LOGO_ALIGNMENT_MAP = {
    InvoiceTemplateSettings.LOGO_POSITION_LEFT: "LEFT",
    InvoiceTemplateSettings.LOGO_POSITION_CENTRE: "CENTER",
    InvoiceTemplateSettings.LOGO_POSITION_RIGHT: "RIGHT",
}

ADDRESS_ALIGNMENT_MAP = {
    InvoiceTemplateSettings.ADDRESS_POSITION_LEFT: "LEFT",
    InvoiceTemplateSettings.ADDRESS_POSITION_RIGHT: "RIGHT",
}
ISSUED_INVOICE_STATUSES = {
    Invoice.STATUS_SENT,
    Invoice.STATUS_VIEWED,
    Invoice.STATUS_OVERDUE,
}
COMPUTER_GENERATED_INVOICE_STATEMENT = (
    "This is a computer generated invoice and therefore no signature is required."
)
REGISTERED_OFFICE_PREFIX_RE = re.compile(r"^\s*registered\s+office\s*:\s*", re.IGNORECASE)
TERM_CONFLICT_RE = re.compile(r"\bwithin\s+(\d+)\s+days?\b.*\binvoice\s+date\b", re.IGNORECASE)


def _normalize_pdf_multiline_text(value: str) -> str:
    normalized = str(value or "")
    normalized = normalized.replace("\r\n", "\n").replace("\r", "\n")
    normalized = normalized.replace("\\r\\n", "\n").replace("\\n", "\n")
    return normalized


def _normalize_pdf_line(value: str) -> str:
    return " ".join(str(value or "").strip().split())


def _strip_registered_office_prefix(value: str) -> str:
    normalized = _normalize_pdf_multiline_text(value).strip()
    previous = None
    while normalized and normalized != previous:
        previous = normalized
        normalized = REGISTERED_OFFICE_PREFIX_RE.sub("", normalized).strip()
    return normalized


def _registered_office_line(context: dict) -> str:
    registered_office = _strip_registered_office_prefix(context["registered_office_text"])
    if not registered_office:
        return "Registered Office:"
    return f"Registered Office: {registered_office}"


def _is_computer_generated_statement(line: str) -> bool:
    normalized_line = _normalize_pdf_line(line).rstrip(".").lower()
    normalized_statement = _normalize_pdf_line(COMPUTER_GENERATED_INVOICE_STATEMENT).rstrip(".").lower()
    return normalized_line == normalized_statement


def _payment_note_conflicts_with_term(line: str, payment_term_days: int) -> bool:
    match = TERM_CONFLICT_RE.search(_normalize_pdf_line(line))
    return bool(match and int(match.group(1)) != int(payment_term_days))


def _invoice_text_lines(value: str, *, skip_computer_statement: bool = True) -> list[str]:
    lines = []
    for line in _normalize_pdf_multiline_text(value).split("\n"):
        normalized_line = _normalize_pdf_line(line)
        if not normalized_line:
            continue
        if skip_computer_statement and _is_computer_generated_statement(normalized_line):
            continue
        lines.append(normalized_line)
    return lines


def _build_customer_display(invoice: Invoice) -> str:
    customer_display = invoice.customer.name
    if invoice.customer.tax_number:
        customer_display = f"{customer_display} ({invoice.customer.tax_number})"
    return customer_display


def _resolve_invoice_payment_summary(invoice: Invoice) -> dict:
    amount_paid = Decimal("0.00")
    if invoice.status == Invoice.STATUS_PAID:
        amount_paid = invoice.total_amount
    amount_due = invoice.total_amount - amount_paid
    show_payment_instructions = (
        invoice.status in ISSUED_INVOICE_STATUSES
        and amount_due > Decimal("0.00")
        and invoice.status not in {Invoice.STATUS_PAID, Invoice.STATUS_REFUNDED}
    )
    payment_status_message = ""
    if amount_due <= Decimal("0.00"):
        payment_status_message = "Paid in Full"
    elif invoice.status == Invoice.STATUS_REFUNDED:
        payment_status_message = "Refunded"
    return {
        "amount_paid": amount_paid,
        "amount_due": amount_due,
        "show_payment_instructions": show_payment_instructions,
        "payment_status_message": payment_status_message,
    }


def _payment_note_lines_for_pdf(context: dict, payment_summary: dict) -> list[str]:
    if not payment_summary["show_payment_instructions"]:
        return []
    lines = []
    for line in _invoice_text_lines(context["invoice_payment_notes"]):
        if _payment_note_conflicts_with_term(line, context["invoice_payment_term_days"]):
            continue
        lines.append(line)
    return lines


def _setting_text_value(
    template_settings: InvoiceTemplateSettings | None,
    field_name: str,
    fallback: str,
    *,
    multiline: bool = False,
) -> str:
    value = ""
    if template_settings is not None:
        value = str(getattr(template_settings, field_name, "") or "").strip()
    if not value:
        value = str(fallback or "").strip()
    if multiline:
        return _normalize_pdf_multiline_text(value)
    return value


def _resolve_invoice_payment_term_days(template_settings: InvoiceTemplateSettings | None) -> int:
    if template_settings is not None and template_settings.default_payment_term_days:
        return int(template_settings.default_payment_term_days)

    try:
        fallback_days = int(getattr(settings, "INVOICE_PAYMENT_TERM_DAYS", 30))
    except (TypeError, ValueError):
        return 30
    if fallback_days < 1 or fallback_days > 365:
        return 30
    return fallback_days


def _normalize_logo_position(value: str) -> str:
    normalized = (value or "").strip().lower()
    if normalized == "center":
        normalized = InvoiceTemplateSettings.LOGO_POSITION_CENTRE
    if normalized in LOGO_ALIGNMENT_MAP:
        return normalized
    return InvoiceTemplateSettings.LOGO_POSITION_LEFT


def _normalize_logo_size(value: str) -> str:
    normalized = (value or "").strip().lower()
    if normalized in LOGO_SIZE_DIMENSIONS_MM:
        return normalized
    return InvoiceTemplateSettings.LOGO_SIZE_MEDIUM


def _normalize_address_position(value: str) -> str:
    normalized = (value or "").strip().lower()
    if normalized in ADDRESS_ALIGNMENT_MAP:
        return normalized
    return InvoiceTemplateSettings.ADDRESS_POSITION_LEFT


def _resolve_invoice_branding(template_settings: InvoiceTemplateSettings | None) -> dict:
    company_name = ""
    company_address = ""
    logo_position = InvoiceTemplateSettings.LOGO_POSITION_LEFT
    logo_size = InvoiceTemplateSettings.LOGO_SIZE_MEDIUM
    address_position = InvoiceTemplateSettings.ADDRESS_POSITION_LEFT

    if template_settings is not None:
        company_name = (template_settings.company_display_name or "").strip()
        company_address = _normalize_pdf_multiline_text(template_settings.company_address).strip()
        logo_position = _normalize_logo_position(template_settings.logo_position)
        logo_size = _normalize_logo_size(template_settings.logo_size)
        address_position = _normalize_address_position(template_settings.address_position)

    return {
        "company_name": company_name,
        "company_address": company_address,
        "logo_position": logo_position,
        "logo_alignment": LOGO_ALIGNMENT_MAP[logo_position],
        "logo_size": logo_size,
        "logo_dimensions": LOGO_SIZE_DIMENSIONS_MM[logo_size],
        "address_position": address_position,
        "address_alignment": ADDRESS_ALIGNMENT_MAP[address_position],
    }


def _get_invoice_logo_path(template_settings: InvoiceTemplateSettings | None) -> str:
    if template_settings is None or not template_settings.logo:
        return ""
    try:
        if not template_settings.logo.storage.exists(template_settings.logo.name):
            return ""
        return template_settings.logo.path
    except (NotImplementedError, ValueError, OSError):
        return ""


def _build_invoice_logo(template_settings: InvoiceTemplateSettings | None, branding: dict):
    logo_path = _get_invoice_logo_path(template_settings)
    if not logo_path:
        return None

    try:
        width_px, height_px = ImageReader(logo_path).getSize()
    except Exception:
        return None

    max_width, max_height = branding["logo_dimensions"]
    scale = min(max_width / width_px, max_height / height_px)
    logo = Image(logo_path, width=width_px * scale, height=height_px * scale)
    logo.hAlign = branding["logo_alignment"]
    logo._invoice_logo_position = branding["logo_position"]
    logo._invoice_logo_size = branding["logo_size"]
    logo._invoice_logo_width = width_px * scale
    logo._invoice_logo_height = height_px * scale
    return logo


def _build_logo_row(logo_flowable: Image, branding: dict, *, column_width: float) -> Table:
    logo_table = Table([[logo_flowable]], colWidths=[column_width])
    logo_table.setStyle(
        TableStyle(
            [
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (0, 0), (-1, -1), branding["logo_alignment"]),
            ]
        )
    )
    logo_table._invoice_logo_position = branding["logo_position"]
    logo_table._invoice_logo_alignment = branding["logo_alignment"]
    logo_table._invoice_logo_width = logo_flowable._invoice_logo_width
    logo_table._invoice_logo_height = logo_flowable._invoice_logo_height
    return logo_table


def build_export_context(invoice: Invoice) -> dict:
    items = list(invoice.items.all())
    template_settings = InvoiceTemplateSettings.current()
    branding = _resolve_invoice_branding(template_settings)
    default_company_name = getattr(settings, "COMPANY_NAME", "Vaniday Singapore Pte Ltd")
    default_company_address = _normalize_pdf_multiline_text(getattr(settings, "COMPANY_ADDRESS", ""))
    company_name = branding["company_name"] or default_company_name
    company_address = branding["company_address"] or default_company_address
    bank_transfer_details = get_bank_transfer_details()
    if bank_transfer_details:
        bank_transfer_details = {
            **bank_transfer_details,
            "instructions": _normalize_pdf_multiline_text(bank_transfer_details.get("instructions", "")),
        }
    return {
        "company_name": company_name,
        "company_email": _setting_text_value(
            template_settings,
            "company_email",
            getattr(settings, "COMPANY_EMAIL", ""),
        ),
        "company_phone": _setting_text_value(
            template_settings,
            "company_phone",
            getattr(settings, "COMPANY_PHONE", ""),
        ),
        "company_address": company_address,
        "company_reg_no": _setting_text_value(
            template_settings,
            "company_registration_number",
            getattr(settings, "COMPANY_REG_NO", ""),
        ),
        "registered_office_text": _setting_text_value(
            template_settings,
            "registered_office_text",
            getattr(settings, "REGISTERED_OFFICE_TEXT", ""),
            multiline=True,
        ),
        "invoice_payment_term_days": _resolve_invoice_payment_term_days(template_settings),
        "bank_transfer_details": bank_transfer_details,
        "invoice_payment_notes": _setting_text_value(
            template_settings,
            "invoice_payment_notes",
            getattr(settings, "INVOICE_PAYMENT_NOTES", ""),
            multiline=True,
        ),
        "invoice_header_text": _setting_text_value(template_settings, "header_text", "", multiline=True),
        "invoice_footer_text": _setting_text_value(template_settings, "footer_text", "", multiline=True),
        "invoice_attention_email": _setting_text_value(
            template_settings,
            "company_email",
            getattr(settings, "COMPANY_EMAIL", "finance@vaniday.com"),
        ),
        "invoice_template_settings": template_settings,
        "invoice_branding": branding,
        "invoice_logo_path": _get_invoice_logo_path(template_settings),
        "bill_to_label": "Bill To",
        "customer_display": _build_customer_display(invoice),
        "invoice": invoice,
        "items": items,
    }


def generate_invoice_pdf(invoice: Invoice) -> bytes:
    context = build_export_context(invoice)
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    body = ParagraphStyle(
        "InvoiceBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
    )
    small = ParagraphStyle(
        "InvoiceSmall",
        parent=body,
        fontSize=8,
        leading=10,
    )
    small_right = ParagraphStyle(
        "InvoiceSmallRight",
        parent=small,
        alignment=2,
    )
    heading = ParagraphStyle(
        "InvoiceHeading",
        parent=body,
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
    )
    heading_right = ParagraphStyle(
        "InvoiceHeadingRight",
        parent=heading,
        alignment=2,
    )
    invoice_title = ParagraphStyle(
        "InvoiceTitle",
        parent=body,
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=24,
    )
    right = ParagraphStyle(
        "InvoiceRight",
        parent=body,
        alignment=2,
    )
    right_bold = ParagraphStyle(
        "InvoiceRightBold",
        parent=right,
        fontName="Helvetica-Bold",
    )
    right_title = ParagraphStyle(
        "InvoiceRightTitle",
        parent=right_bold,
        fontSize=22,
        leading=24,
    )
    body_right = ParagraphStyle(
        "InvoiceBodyRight",
        parent=body,
        alignment=2,
    )

    template_settings = context["invoice_template_settings"]
    branding = context["invoice_branding"]
    address_is_right = branding["address_position"] == InvoiceTemplateSettings.ADDRESS_POSITION_RIGHT
    company_heading_style = heading_right if address_is_right else heading
    company_body_style = body_right if address_is_right else body
    logo_flowable = _build_invoice_logo(template_settings, branding)

    attention_email = context["invoice_attention_email"]
    office_line = _registered_office_line(context)
    attention_line = f"Attention: {attention_email}"
    header_left = []
    if logo_flowable is not None:
        header_left.append([_build_logo_row(logo_flowable, branding, column_width=78 * mm)])
    header_left.extend(
        [
            [Paragraph(f"<b>{context['company_name']}</b>", company_heading_style)],
            [Paragraph(f"Reg No. {context['company_reg_no']}", company_body_style)],
            [Paragraph(context["company_address"].replace("\n", "<br/>"), company_body_style)],
            [Paragraph(f"Email: {context['company_email']}", company_body_style)],
            [Paragraph(f"Phone: {context['company_phone']}", company_body_style)],
        ]
    )
    left_header_table = Table(header_left, colWidths=[78 * mm])
    left_header_table.setStyle(
        TableStyle(
            [
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    right_header_data = [
        [Paragraph("INVOICE", right_title)],
        [Paragraph(f"<b>{context['bill_to_label']}</b>", right)],
        [Paragraph(context["customer_display"], right_bold)],
        [Paragraph("<b>Invoice Date</b>", right)],
        [Paragraph(str(invoice.issue_date.strftime("%d %b %Y")), right)],
        [Paragraph("<b>Invoice Number</b>", right)],
        [Paragraph(invoice.invoice_number, right)],
    ]
    right_header_table = Table(right_header_data, colWidths=[82 * mm])
    right_header_table.setStyle(
        TableStyle(
            [
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
            ]
        )
    )

    story = [
        Paragraph(office_line, small_right),
        Paragraph(attention_line, small_right),
    ]
    for line in _invoice_text_lines(context["invoice_header_text"], skip_computer_statement=False):
        story.append(Paragraph(line, small_right))
    story.extend(
        [
            Spacer(1, 6),
            Table([[left_header_table, right_header_table]], colWidths=[78 * mm, 82 * mm]),
            Spacer(1, 10),
        ]
    )

    amount_header = f"Amount {invoice.currency}"
    table_data = [["Description", "Quantity", "Unit Price", amount_header]]
    for item in context["items"]:
        description = (item.description or "").replace("\n", "<br/>")
        unit_value = f"{item.unit_price:.2f}"
        line_value = f"{item.line_total:.2f}"
        table_data.append(
            [
                Paragraph(description, body),
                str(item.quantity),
                unit_value,
                line_value,
            ]
        )
    if len(table_data) == 1:
        table_data.append(["No items", "-", "-", "-"])

    table = Table(table_data, colWidths=[95 * mm, 20 * mm, 22 * mm, 23 * mm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("LINEABOVE", (0, 0), (-1, 0), 0.8, colors.black),
                ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.black),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LINEBELOW", (0, -1), (-1, -1), 0.8, colors.black),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 10))

    payment_summary = _resolve_invoice_payment_summary(invoice)
    amount_paid = payment_summary["amount_paid"]
    amount_due = payment_summary["amount_due"]
    totals = [
        ["Subtotal", f"{invoice.subtotal:.2f}"],
        ["GST", f"{invoice.tax_amount:.2f}"],
        [f"TOTAL {invoice.currency}", f"{invoice.total_amount:.2f}"],
        ["Less Amount Paid", f"{amount_paid:.2f}"],
        [f"AMOUNT DUE {invoice.currency}", f"{amount_due:.2f}"],
    ]
    totals_table = Table(totals, colWidths=[124 * mm, 36 * mm])
    totals_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("LINEABOVE", (0, 2), (-1, 2), 0.8, colors.black),
                ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.black),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(totals_table)

    story.append(Spacer(1, 8))
    if payment_summary["payment_status_message"]:
        story.append(Paragraph(f"<b>{payment_summary['payment_status_message']}</b>", body))
    else:
        story.append(Paragraph(f"Due Date: {invoice.due_date.strftime('%d %b %Y')}", body))
        story.append(Paragraph(f"Payment Term: {context['invoice_payment_term_days']} Days", body))
    bank_transfer_details = context["bank_transfer_details"]
    if bank_transfer_details and payment_summary["show_payment_instructions"]:
        story.append(Spacer(1, 6))
        story.append(Paragraph("<b>Bank Transfer Details</b>", body))
        story.append(Paragraph("Please arrange payment via bank transfer to:", body))
        detail_lines = [
            ("Account Name", bank_transfer_details.get("account_name", "")),
            ("Bank", bank_transfer_details.get("bank_name", "")),
            ("Account Number", bank_transfer_details.get("account_number", "")),
            ("PayNow ID", bank_transfer_details.get("paynow_id", "")),
            ("BIC/SWIFT", bank_transfer_details.get("bic", "")),
        ]
        for label, value in detail_lines:
            if value:
                story.append(Paragraph(f"<b>{label}:</b> {value}", body))
        for line in bank_transfer_details.get("instructions", "").split("\n"):
            if line.strip():
                story.append(Paragraph(line, body))

    story.append(Spacer(1, 4))
    for line in _payment_note_lines_for_pdf(context, payment_summary):
        story.append(Paragraph(line, small))

    if invoice.notes:
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"Notes: {invoice.notes}", body))

    story.append(Spacer(1, 6))
    footer_lines = _invoice_text_lines(context["invoice_footer_text"])
    for line in footer_lines:
        story.append(Paragraph(line, small))

    if footer_lines:
        story.append(Spacer(1, 4))
    story.append(
        Paragraph(
            "This is a computer generated invoice and therefore no signature is required.",
            small,
        )
    )

    doc.build(story)
    return buffer.getvalue()


def generate_invoice_excel(invoice: Invoice) -> bytes:
    context = build_export_context(invoice)
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    header_fill = PatternFill(fill_type="solid", start_color="0D6EFD", end_color="0D6EFD")
    header_font = Font(color="FFFFFF", bold=True)

    ws["A1"] = context["company_name"]
    ws["A2"] = context["company_address"]
    ws["A3"] = f"Email: {context['company_email']}"
    ws["A4"] = f"Phone: {context['company_phone']}"

    ws["E1"] = "Invoice #"
    ws["F1"] = invoice.invoice_number
    ws["E2"] = "Issue Date"
    ws["F2"] = str(invoice.issue_date)
    ws["E3"] = "Due Date"
    ws["F3"] = str(invoice.due_date)
    ws["E4"] = "Status"
    ws["F4"] = invoice.get_status_display()

    ws["A6"] = "Bill To"
    ws["A7"] = invoice.customer.name
    ws["A8"] = invoice.customer.billing_address or "-"
    ws["A9"] = invoice.customer.email

    headers = ["Description", "Qty", "Unit Price", "GST %", "Line Total"]
    row = 11
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row += 1
    for item in context["items"]:
        ws.cell(row=row, column=1, value=item.description)
        ws.cell(row=row, column=2, value=float(item.quantity))
        ws.cell(row=row, column=3, value=float(item.unit_price))
        ws.cell(row=row, column=4, value=float(item.tax_rate))
        ws.cell(row=row, column=5, value=float(item.line_total))
        row += 1

    if row == 12:
        ws.cell(row=row, column=1, value="No items")
        row += 1

    row += 1
    ws.cell(row=row, column=4, value="Subtotal").font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(invoice.subtotal))
    row += 1
    ws.cell(row=row, column=4, value="GST").font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(invoice.tax_amount))
    row += 1
    ws.cell(row=row, column=4, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(invoice.total_amount)).font = Font(bold=True)

    if invoice.notes:
        row += 2
        ws.cell(row=row, column=1, value="Notes").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value=invoice.notes)

    widths = {"A": 45, "B": 10, "C": 15, "D": 10, "E": 15, "F": 25}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
