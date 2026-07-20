from __future__ import annotations

import csv
from collections import defaultdict
from datetime import datetime, timedelta
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO, StringIO
from typing import Any
from zipfile import BadZipFile

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.db.models import F
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .models import Customer, Invoice, InvoiceItem, InvoiceSourceRow

ZERO = Decimal("0.00")
TWOPLACES = Decimal("0.01")
OVERDUE_ELIGIBLE_STATUSES = {
    Invoice.STATUS_SENT,
    Invoice.STATUS_VIEWED,
}


def _to_money(value: Decimal) -> Decimal:
    return value.quantize(TWOPLACES, rounding=ROUND_HALF_UP)


def generate_invoice_number() -> str:
    year = timezone.localdate().year
    prefix = f"INV-{year}-"
    last_invoice = (
        Invoice.objects.filter(invoice_number__startswith=prefix)
        .order_by("-invoice_number")
        .only("invoice_number")
        .first()
    )
    if not last_invoice:
        return f"{prefix}0001"

    try:
        last_seq = int(last_invoice.invoice_number.split("-")[-1])
    except (ValueError, IndexError):
        last_seq = Invoice.objects.filter(invoice_number__startswith=prefix).count()
    return f"{prefix}{last_seq + 1:04d}"


def calculate_item_amounts(item: InvoiceItem) -> tuple[Decimal, Decimal, Decimal]:
    base_amount = _to_money(Decimal(item.quantity) * Decimal(item.unit_price))
    tax_amount = _to_money(base_amount * (Decimal(item.tax_rate) / Decimal("100")))
    line_total = _to_money(base_amount + tax_amount)
    return base_amount, tax_amount, line_total


def recalculate_invoice_totals(invoice: Invoice) -> Invoice:
    subtotal = ZERO
    tax_total = ZERO
    total = ZERO
    for item in invoice.items.all():
        base_amount, tax_amount, line_total = calculate_item_amounts(item)
        if item.line_total != line_total:
            item.line_total = line_total
            item.save(update_fields=["line_total", "updated_at"])
        subtotal += base_amount
        tax_total += tax_amount
        total += line_total

    invoice.subtotal = _to_money(subtotal)
    invoice.tax_amount = _to_money(tax_total)
    invoice.total_amount = _to_money(total)
    invoice.save(update_fields=["subtotal", "tax_amount", "total_amount", "updated_at"])
    return invoice


def apply_overdue_status(invoice: Invoice, today: date | None = None) -> bool:
    today = today or timezone.localdate()
    if invoice.due_date < today and invoice.status in OVERDUE_ELIGIBLE_STATUSES:
        invoice.status = Invoice.STATUS_OVERDUE
        invoice.save(update_fields=["status", "updated_at"])
        return True
    return False


def refresh_overdue_invoices(today: date | None = None) -> int:
    today = today or timezone.localdate()
    return Invoice.objects.filter(
        due_date__lt=today,
        status__in=OVERDUE_ELIGIBLE_STATUSES,
    ).update(status=Invoice.STATUS_OVERDUE, updated_at=timezone.now())


def transition_invoice_status(invoice: Invoice, new_status: str) -> tuple[bool, str]:
    allowed_transitions = {
        Invoice.STATUS_DRAFT: {Invoice.STATUS_SENT, Invoice.STATUS_PAID},
        Invoice.STATUS_SENT: {Invoice.STATUS_VIEWED, Invoice.STATUS_PAID, Invoice.STATUS_OVERDUE},
        Invoice.STATUS_VIEWED: {Invoice.STATUS_PAID, Invoice.STATUS_OVERDUE},
        Invoice.STATUS_OVERDUE: {Invoice.STATUS_PAID},
        Invoice.STATUS_PAID: set(),
        Invoice.STATUS_PARTIALLY_REFUNDED: set(),
        Invoice.STATUS_REFUNDED: set(),
    }
    if new_status not in dict(Invoice.STATUS_CHOICES):
        return False, "Invalid status."
    if new_status == invoice.status:
        return False, "Invoice is already in the selected status."
    if new_status == Invoice.STATUS_PAID:
        return False, "Invoices can only be marked paid by Stripe confirmation or verified bank-transfer confirmation."

    allowed = allowed_transitions.get(invoice.status, set())
    if new_status not in allowed:
        return False, f"Cannot move status from {invoice.status} to {new_status}."

    if new_status == Invoice.STATUS_OVERDUE and invoice.due_date >= timezone.localdate():
        return False, "Invoice cannot be marked overdue before its due date."

    invoice.status = new_status
    invoice.save(update_fields=["status", "updated_at"])
    return True, "Status updated."


@transaction.atomic
def mark_invoice_viewed(invoice: Invoice) -> Invoice:
    Invoice.objects.filter(pk=invoice.pk).update(view_count=F("view_count") + 1)
    invoice.refresh_from_db()
    if invoice.viewed_at is None:
        invoice.viewed_at = timezone.now()
    if invoice.status == Invoice.STATUS_SENT:
        invoice.status = Invoice.STATUS_VIEWED
    invoice.save(update_fields=["viewed_at", "status", "updated_at"])
    apply_overdue_status(invoice)
    return invoice


CSV_HEADER_ALIASES = {
    "seller_id": ["seller_id", "sellerid"],
    "shop_title": ["shop_title", "shoptitle", "merchant", "merchant_name"],
    "order_id": ["orderid", "order_id"],
    "partner_type_name": ["partnertypename", "partner_type_name"],
    "payment_method": ["paymentmethod", "payment_method"],
    "product_type": ["producttype", "product_type"],
    "customer_id": ["customerid", "customer_id"],
    "status": ["status"],
    "order_status": ["orderstatus", "order_status"],
    "email": ["email"],
    "customer_name": ["customername", "customer_name"],
    "contact_no": ["contactno", "contact_no"],
    "qty": ["qty", "quantity"],
    "service_name": ["servicename", "service_name"],
    "booked_date": ["bookeddate", "booked_date"],
    "service_duration": ["service_duration", "serviceduration"],
    "staff_id": ["staffid", "staff_id"],
    "staff_name": ["staffname", "staff_name"],
    "total_revenue": ["total_revenue", "totalrevenue"],
    "credit_card": ["credit_card", "creditcard"],
    "shipping_amount": ["shippingamount", "shipping_amount"],
    "reward_point": ["reward_point", "rewardpoint"],
    "vaniday_commission": ["vanidaycommission", "vaniday_commission"],
    "vaniday_share": ["vanidayshare", "vaniday_share"],
    "cashback_fee": ["cashback_fee", "cashbackfee"],
    "cashback_discount": ["cashback_discount", "cashbackdiscount"],
    "cashback_date": ["cashback_date", "cashbackdate"],
    "salon_share": ["salonshare", "salon_share"],
}

INVOICE_IMPORT_REQUIRED_HEADERS = {
    "order reference": ["order_id"],
    "merchant/customer name": ["shop_title", "customer_name"],
    "customer email": ["email"],
    "service description": ["service_name"],
    "booked date": ["booked_date"],
    "invoice amount": ["vaniday_share"],
}
DUPLICATE_INVOICE_IMPORT_MESSAGE = "This order/service row has already been imported."


def _normalize_key(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum())


def _stringify_import_value(raw_value: Any) -> str:
    if raw_value in (None, ""):
        return ""
    if isinstance(raw_value, datetime):
        if raw_value.time() == datetime.min.time():
            return raw_value.strftime("%Y-%m-%d")
        return raw_value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(raw_value, date):
        return raw_value.strftime("%Y-%m-%d")
    if isinstance(raw_value, bool):
        return "true" if raw_value else "false"
    if isinstance(raw_value, Decimal):
        return format(raw_value, "f").rstrip("0").rstrip(".") or "0"
    if isinstance(raw_value, int):
        return str(raw_value)
    if isinstance(raw_value, float):
        if raw_value.is_integer():
            return str(int(raw_value))
        return format(raw_value, "f").rstrip("0").rstrip(".")
    return str(raw_value).strip()


def _parse_decimal(raw_value: Any) -> Decimal | None:
    if raw_value is None:
        return None
    text = str(raw_value).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except Exception:
        return None


def _parse_int(raw_value: Any) -> int | None:
    number = _parse_decimal(raw_value)
    if number is None:
        return None
    try:
        return int(number)
    except Exception:
        return None


def _parse_datetime(raw_value: Any) -> datetime | None:
    if raw_value is None:
        return None
    text = str(raw_value).strip()
    if not text:
        return None
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    ]
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if timezone.is_naive(parsed):
                return timezone.make_aware(parsed, timezone.get_current_timezone())
            return parsed
        except ValueError:
            continue
    return None


def _parse_date(raw_value: Any) -> date | None:
    parsed = _parse_datetime(raw_value)
    return parsed.date() if parsed else None


def _resolve_field(normalized_row: dict[str, Any], canonical_field: str) -> str:
    for alias in CSV_HEADER_ALIASES.get(canonical_field, []):
        if alias in normalized_row:
            return str(normalized_row.get(alias, "")).strip()
    return ""


def _coalesce(*values: str) -> str:
    for value in values:
        cleaned = (value or "").strip()
        if cleaned:
            return cleaned
    return ""


def _normalize_duplicate_text(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _normalize_duplicate_datetime(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        parsed = value
    else:
        parsed = _parse_datetime(value)
        if parsed is None:
            try:
                parsed = datetime.fromisoformat(str(value).strip())
            except (TypeError, ValueError):
                parsed = None
    if parsed is None:
        return ""
    if timezone.is_aware(parsed):
        parsed = timezone.localtime(parsed)
    return parsed.replace(microsecond=0).isoformat()


def _normalize_duplicate_amount(value: Any) -> str:
    amount = _parse_decimal(value)
    if amount is None:
        return ""
    return f"{_to_money(amount):.2f}"


def _invoice_source_duplicate_key(
    *,
    order_id: Any,
    customer_id: Any,
    email: Any,
    service_name: Any,
    booked_date: Any,
    vaniday_share: Any,
) -> tuple[str, str, str, str, str] | None:
    normalized_order_id = _normalize_duplicate_text(order_id)
    normalized_customer = _normalize_duplicate_text(customer_id) or _normalize_duplicate_text(_sanitize_email(str(email or "")))
    normalized_service = _normalize_duplicate_text(service_name)
    normalized_booked_date = _normalize_duplicate_datetime(booked_date)
    normalized_amount = _normalize_duplicate_amount(vaniday_share)

    key_parts = [
        normalized_order_id,
        normalized_customer,
        normalized_service,
        normalized_booked_date,
        normalized_amount,
    ]
    if not all(key_parts):
        return None
    return tuple(key_parts)


def _invoice_row_duplicate_key(row: dict[str, Any]) -> tuple[str, str, str, str, str] | None:
    source = row["source"]
    return _invoice_source_duplicate_key(
        order_id=source["order_id"],
        customer_id=source["customer_id"],
        email=source["email"],
        service_name=source["service_name"],
        booked_date=row["booked_at"],
        vaniday_share=row["amount"],
    )


def _invoice_source_item_signature(source_row: InvoiceSourceRow) -> tuple[str, str]:
    description = _build_item_description(
        source_row.service_name,
        source_row.order_id,
        source_row.booked_date,
        source_row.product_type,
    )[:255]
    return description, _normalize_duplicate_amount(source_row.vaniday_share)


def _existing_imported_invoice_row_keys() -> set[tuple[str, str, str, str, str]]:
    duplicate_keys = set()
    imported_item_signatures = {
        (description, _normalize_duplicate_amount(line_total))
        for description, line_total in InvoiceItem.objects.values_list("description", "line_total")
    }
    source_rows = InvoiceSourceRow.objects.only(
        "order_id",
        "shop_title",
        "customer_name",
        "customer_id",
        "email",
        "product_type",
        "service_name",
        "booked_date",
        "vaniday_share",
    )
    for source_row in source_rows.iterator():
        if _invoice_source_item_signature(source_row) not in imported_item_signatures:
            continue
        if not _coalesce(source_row.shop_title, source_row.customer_name):
            continue
        duplicate_key = _invoice_source_duplicate_key(
            order_id=source_row.order_id,
            customer_id=source_row.customer_id,
            email=source_row.email,
            service_name=source_row.service_name,
            booked_date=source_row.booked_date,
            vaniday_share=source_row.vaniday_share,
        )
        if duplicate_key is not None:
            duplicate_keys.add(duplicate_key)
    return duplicate_keys


def _invoice_group_period(booked_date: datetime | None) -> str:
    reference = booked_date.date() if booked_date else timezone.localdate()
    return reference.strftime("%Y-%m")


def _build_item_description(service_name: str, order_id: str, booked_date: datetime | None, product_type: str) -> str:
    base_service = _coalesce(service_name, product_type, "Imported Service")
    parts = [base_service]
    if order_id:
        parts.append(f"OrderID: {order_id}")
    if booked_date:
        parts.append(f"Booked: {booked_date.strftime('%d %b %Y')}")
    return " | ".join(parts)


def _sanitize_email(raw_email: str) -> str:
    email = (raw_email or "").strip().lower()
    if not email:
        return ""
    try:
        validate_email(email)
    except ValidationError:
        return ""
    return email


def _build_fallback_email(customer_name: str, seller_id: str) -> str:
    safe_name = "".join(ch.lower() if ch.isalnum() else "-" for ch in customer_name).strip("-")
    safe_name = safe_name or "merchant"
    safe_seller = "".join(ch.lower() if ch.isalnum() else "" for ch in seller_id)
    suffix = safe_seller[:12] or "import"
    return f"{safe_name}-{suffix}@import.local"


def _max_invoice_import_rows() -> int:
    return int(getattr(settings, "INVOICE_IMPORT_MAX_ROWS", 1000))


def _validate_invoice_headers(headers: list[str]) -> None:
    normalized_headers = {_normalize_key(str(header)) for header in headers if str(header).strip()}
    missing_groups = []
    for group_label, canonical_fields in INVOICE_IMPORT_REQUIRED_HEADERS.items():
        valid_aliases = {
            alias
            for field in canonical_fields
            for alias in CSV_HEADER_ALIASES.get(field, [])
        }
        if not normalized_headers.intersection(valid_aliases):
            missing_groups.append(group_label)

    if missing_groups:
        missing_labels = "; ".join(missing_groups)
        raise ValueError(
            "Missing required columns for: "
            f"{missing_labels}. Required upload columns are OrderID, shop_title or customerName, "
            "email, serviceName, bookedDate, and vanidayShare."
        )


def _map_invoice_source_row(source_row: dict[str, Any]) -> dict[str, str]:
    normalized_source = {
        _normalize_key(str(key)): _stringify_import_value(value)
        for key, value in source_row.items()
        if key is not None and str(key).strip()
    }
    return {
        "seller_id": _resolve_field(normalized_source, "seller_id"),
        "shop_title": _resolve_field(normalized_source, "shop_title"),
        "order_id": _resolve_field(normalized_source, "order_id"),
        "partner_type_name": _resolve_field(normalized_source, "partner_type_name"),
        "payment_method": _resolve_field(normalized_source, "payment_method"),
        "product_type": _resolve_field(normalized_source, "product_type"),
        "customer_id": _resolve_field(normalized_source, "customer_id"),
        "status": _resolve_field(normalized_source, "status"),
        "order_status": _resolve_field(normalized_source, "order_status"),
        "email": _resolve_field(normalized_source, "email"),
        "customer_name": _resolve_field(normalized_source, "customer_name"),
        "contact_no": _resolve_field(normalized_source, "contact_no"),
        "qty": _resolve_field(normalized_source, "qty"),
        "service_name": _resolve_field(normalized_source, "service_name"),
        "booked_date": _resolve_field(normalized_source, "booked_date"),
        "service_duration": _resolve_field(normalized_source, "service_duration"),
        "staff_id": _resolve_field(normalized_source, "staff_id"),
        "staff_name": _resolve_field(normalized_source, "staff_name"),
        "total_revenue": _resolve_field(normalized_source, "total_revenue"),
        "credit_card": _resolve_field(normalized_source, "credit_card"),
        "shipping_amount": _resolve_field(normalized_source, "shipping_amount"),
        "reward_point": _resolve_field(normalized_source, "reward_point"),
        "vaniday_commission": _resolve_field(normalized_source, "vaniday_commission"),
        "vaniday_share": _resolve_field(normalized_source, "vaniday_share"),
        "cashback_fee": _resolve_field(normalized_source, "cashback_fee"),
        "cashback_discount": _resolve_field(normalized_source, "cashback_discount"),
        "cashback_date": _resolve_field(normalized_source, "cashback_date"),
        "salon_share": _resolve_field(normalized_source, "salon_share"),
    }


def _transform_invoice_source_row(index: int, mapped: dict[str, str]) -> dict[str, Any]:
    customer_name = _coalesce(mapped["shop_title"], mapped["customer_name"])
    email_value = _sanitize_email(mapped["email"])
    amount_value = _parse_decimal(mapped["vaniday_share"])
    quantity_value = _parse_decimal(mapped["qty"]) or Decimal("1.00")
    booked_at = _parse_datetime(mapped["booked_date"])

    row_errors = []
    if not mapped["order_id"]:
        row_errors.append("Missing OrderID/order reference.")
    if not customer_name:
        row_errors.append("Missing merchant/customer name (shop_title or customerName).")
    if not mapped["email"]:
        row_errors.append("Missing customer email.")
    elif not email_value:
        row_errors.append("Customer email format is invalid.")
    if not mapped["service_name"]:
        row_errors.append("Missing serviceName/item description.")
    if not mapped["booked_date"]:
        row_errors.append("Missing bookedDate.")
    elif booked_at is None:
        row_errors.append("bookedDate format is invalid.")
    if not mapped["vaniday_share"]:
        row_errors.append("Missing invoice amount (vanidayShare).")
    elif amount_value is None:
        row_errors.append("Invoice amount (vanidayShare) must be numeric.")
    elif amount_value < ZERO:
        row_errors.append("Amount must be zero or positive.")
    if quantity_value <= ZERO:
        row_errors.append("Quantity must be greater than zero.")

    return {
        "row_number": index,
        "source": mapped,
        "customer_name": customer_name,
        "email": email_value,
        "amount": str(_to_money(amount_value or ZERO)),
        "quantity": str(_to_money(quantity_value)),
        "booked_at": booked_at.isoformat() if booked_at else "",
        "item_description": _build_item_description(
            mapped["service_name"],
            mapped["order_id"],
            booked_at,
            mapped["product_type"],
        ),
        "group_period": _invoice_group_period(booked_at),
        "errors": row_errors,
    }


def _build_invoice_parse_result(headers: list[str], source_rows: list[tuple[int, dict[str, Any]]]) -> dict[str, Any]:
    _validate_invoice_headers(headers)

    imported_duplicate_keys = _existing_imported_invoice_row_keys()
    normalized_rows: list[dict[str, Any]] = []
    invalid_rows: list[dict[str, Any]] = []
    valid_rows: list[dict[str, Any]] = []

    for index, source_row in source_rows:
        transformed = _transform_invoice_source_row(index, _map_invoice_source_row(source_row))
        if not transformed["errors"] and _invoice_row_duplicate_key(transformed) in imported_duplicate_keys:
            transformed["errors"].append(DUPLICATE_INVOICE_IMPORT_MESSAGE)
        normalized_rows.append(transformed)
        if transformed["errors"]:
            invalid_rows.append(transformed)
        else:
            valid_rows.append(transformed)

    preview_groups = defaultdict(lambda: {"customer_name": "", "period": "", "rows": 0, "amount_total": ZERO})
    for row in valid_rows:
        key = f"{row['customer_name']}|{row['email'] or 'no-email'}|{row['group_period']}"
        group = preview_groups[key]
        group["customer_name"] = row["customer_name"]
        group["period"] = row["group_period"]
        group["rows"] += 1
        group["amount_total"] += Decimal(row["amount"])

    return {
        "headers": headers,
        "total_rows": len(normalized_rows),
        "all_rows": normalized_rows,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "preview_groups": [
            {
                "customer_name": group["customer_name"],
                "period": group["period"],
                "rows": group["rows"],
                "amount_total": f"{_to_money(group['amount_total']):.2f}",
            }
            for group in preview_groups.values()
        ],
    }


def parse_invoice_csv(uploaded_file) -> dict[str, Any]:
    raw_bytes = uploaded_file.read()
    uploaded_file.seek(0)
    if not raw_bytes:
        raise ValueError("Uploaded file is empty.")

    decode_attempts = ["utf-8-sig", "utf-8", "latin-1"]
    decoded_text = None
    for encoding in decode_attempts:
        try:
            decoded_text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded_text is None:
        raise ValueError("CSV encoding is not supported. Please upload UTF-8 or Latin-1 CSV.")

    reader = csv.DictReader(StringIO(decoded_text))
    headers = reader.fieldnames or []
    if not headers:
        raise ValueError("CSV file has no header row.")

    source_rows = []
    max_rows = _max_invoice_import_rows()
    for index, source_row in enumerate(reader, start=2):
        if source_row is None:
            continue
        if all(_stringify_import_value(value) == "" for value in source_row.values()):
            continue
        source_rows.append((index, source_row))
        if len(source_rows) > max_rows:
            raise ValueError(f"Upload exceeds the maximum of {max_rows} data rows.")

    return _build_invoice_parse_result(headers, source_rows)


def parse_invoice_excel(uploaded_file) -> dict[str, Any]:
    raw_bytes = uploaded_file.read()
    uploaded_file.seek(0)
    if not raw_bytes:
        raise ValueError("Uploaded file is empty.")

    try:
        workbook = load_workbook(BytesIO(raw_bytes), data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise ValueError("Excel workbook is corrupted or unreadable.") from exc

    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    header_row_index = None
    header_row = None
    for row_index, row in enumerate(rows, start=1):
        if row and any(_stringify_import_value(cell) for cell in row):
            header_row_index = row_index
            header_row = row
            break

    if header_row is None:
        raise ValueError("Excel workbook is empty.")

    headers = [_stringify_import_value(cell) for cell in header_row]
    source_rows = []
    max_rows = _max_invoice_import_rows()
    for row_index, row in enumerate(rows[header_row_index:], start=header_row_index + 1):
        if row is None or all(_stringify_import_value(cell) == "" for cell in row):
            continue
        source_row = {
            headers[column_index]: row[column_index] if column_index < len(row) else None
            for column_index in range(len(headers))
            if headers[column_index]
        }
        source_rows.append((row_index, source_row))
        if len(source_rows) > max_rows:
            raise ValueError(f"Upload exceeds the maximum of {max_rows} data rows.")

    return _build_invoice_parse_result(headers, source_rows)


def parse_invoice_upload(uploaded_file) -> dict[str, Any]:
    file_name = (uploaded_file.name or "").strip().lower()
    if file_name.endswith(".csv"):
        return parse_invoice_csv(uploaded_file)
    if file_name.endswith(".xlsx"):
        return parse_invoice_excel(uploaded_file)
    raise ValueError("Upload a CSV or Excel (.xlsx) file.")


def import_invoice_rows_from_preview(
    valid_rows: list[dict[str, Any]],
    all_rows: list[dict[str, Any]],
    source_file_name: str,
    initiated_by=None,
) -> dict[str, int]:
    customers_cache: dict[str, Customer] = {}
    created_customers = 0
    created_invoices = 0
    created_items = 0
    stored_source_rows = 0
    existing_duplicate_keys = _existing_imported_invoice_row_keys()
    importable_valid_rows = [
        row
        for row in valid_rows
        if _invoice_row_duplicate_key(row) not in existing_duplicate_keys
    ]

    grouped_valid_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in importable_valid_rows:
        group_key = f"{row['customer_name']}|{row['email'] or 'no-email'}|{row['group_period']}"
        grouped_valid_rows[group_key].append(row)

    with transaction.atomic():
        for row in all_rows:
            if DUPLICATE_INVOICE_IMPORT_MESSAGE in row["errors"]:
                continue
            duplicate_key = _invoice_row_duplicate_key(row)
            if duplicate_key is not None and duplicate_key in existing_duplicate_keys:
                continue
            source = row["source"]
            InvoiceSourceRow.objects.create(
                seller_id=source["seller_id"],
                shop_title=source["shop_title"],
                order_id=source["order_id"],
                partner_type_name=source["partner_type_name"],
                payment_method=source["payment_method"],
                product_type=source["product_type"],
                customer_id=source["customer_id"],
                status=source["status"],
                order_status=source["order_status"],
                email=_sanitize_email(source["email"]),
                customer_name=source["customer_name"],
                contact_no=source["contact_no"],
                qty=_parse_decimal(source["qty"]),
                service_name=source["service_name"],
                booked_date=_parse_datetime(source["booked_date"]),
                service_duration=_parse_int(source["service_duration"]),
                staff_id=source["staff_id"],
                staff_name=source["staff_name"],
                total_revenue=_parse_decimal(source["total_revenue"]),
                credit_card=_parse_decimal(source["credit_card"]),
                shipping_amount=_parse_decimal(source["shipping_amount"]),
                reward_point=_parse_decimal(source["reward_point"]),
                vaniday_commission=_parse_decimal(source["vaniday_commission"]),
                vaniday_share=_parse_decimal(source["vaniday_share"]),
                cashback_fee=_parse_decimal(source["cashback_fee"]),
                cashback_discount=_parse_decimal(source["cashback_discount"]),
                cashback_date=_parse_date(source["cashback_date"]),
                salon_share=_parse_decimal(source["salon_share"]),
                source_file_name=source_file_name,
                raw_data=source,
            )
            stored_source_rows += 1

        for grouped_rows in grouped_valid_rows.values():
            sample = grouped_rows[0]
            customer_name = sample["customer_name"]
            email_value = sample["email"]
            seller_id = sample["source"]["seller_id"]
            customer_email = email_value or _build_fallback_email(customer_name, seller_id)

            if customer_email in customers_cache:
                customer = customers_cache[customer_email]
            else:
                customer = Customer.objects.filter(email__iexact=customer_email).first()
                if customer is None:
                    customer = Customer.objects.create(
                        name=customer_name[:255],
                        email=customer_email,
                        phone=sample["source"]["contact_no"][:30],
                        created_by=initiated_by,
                    )
                    created_customers += 1
                else:
                    updated_fields = []
                    if not customer.name and customer_name:
                        customer.name = customer_name[:255]
                        updated_fields.append("name")
                    if not customer.phone and sample["source"]["contact_no"]:
                        customer.phone = sample["source"]["contact_no"][:30]
                        updated_fields.append("phone")
                    if updated_fields:
                        updated_fields.append("updated_at")
                        customer.save(update_fields=updated_fields)
                customers_cache[customer_email] = customer

            booked_dates = [
                datetime.fromisoformat(row["booked_at"]).date()
                for row in grouped_rows
                if row["booked_at"]
            ]
            issue_date = min(booked_dates) if booked_dates else timezone.localdate()
            payment_term_days = getattr(settings, "INVOICE_PAYMENT_TERM_DAYS", 30)
            due_date = issue_date + timedelta(days=payment_term_days)

            invoice = Invoice.objects.create(
                invoice_number=generate_invoice_number(),
                customer=customer,
                status=Invoice.STATUS_DRAFT,
                issue_date=issue_date,
                due_date=due_date,
                currency="SGD",
                created_by=initiated_by,
                notes=f"Imported from file: {source_file_name}",
            )
            created_invoices += 1

            for row in grouped_rows:
                amount = Decimal(row["amount"])
                quantity = Decimal(row["quantity"])
                unit_price = _to_money(amount / quantity) if quantity > ZERO else amount
                InvoiceItem.objects.create(
                    invoice=invoice,
                    description=row["item_description"][:255],
                    quantity=quantity,
                    unit_price=unit_price,
                    tax_rate=ZERO,
                    line_total=amount,
                )
                created_items += 1

            recalculate_invoice_totals(invoice)

    return {
        "created_customers": created_customers,
        "created_invoices": created_invoices,
        "created_items": created_items,
        "saved_rows": len(importable_valid_rows),
        "stored_source_rows": stored_source_rows,
    }
